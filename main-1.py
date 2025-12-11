"""
Telegram Trucker Dispatch Assistant Bot

Features:
- Load parsing and tracking (PU/DEL workflow)
- Real-time ETA calculations via OSRM routing
- Geofence alerts (auto-detect arrival/departure)
- Document reminders (PTI, BOL, POD)
- Appointment schedule alerts with thresholds
- Weekly load catalog with Excel export
- Live location tracking

Version: 2025-12-11_full_features_v1
"""

import asyncio
import hashlib
import html
import io
import json
import math
import os
import re
import time
from datetime import datetime, timezone, timedelta, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Set
from enum import Enum

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from timezonefinder import TimezoneFinder
from zoneinfo import ZoneInfo

from telegram import (
    Update,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.error import Forbidden, BadRequest, TelegramError
from telegram.ext import (
    Application,
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

BOT_VERSION = "2025-12-11_full_features_v1"


# ============================================================================
# CONSTANTS
# ============================================================================
EARTH_RADIUS_M = 6371000.0
METERS_PER_MILE = 1609.344
MILES_PER_METER = 1 / METERS_PER_MILE
MAX_HISTORY_RECORDS = 1000
MAX_DELETEALL_MESSAGES = 2000
EXCEL_SHEET_NAME_MAX_LEN = 31
DELETE_MESSAGE_DELAY_SEC = 0.02
REMINDER_CHECK_INTERVAL_SEC = 60
GEOFENCE_CHECK_INTERVAL_SEC = 30


# ============================================================================
# ENVIRONMENT HELPERS
# ============================================================================
def _strip_quotes(s: str) -> str:
    s = (s or "").strip()
    if len(s) >= 2 and ((s[0] == s[-1] == '"') or (s[0] == s[-1] == "'")):
        return s[1:-1].strip()
    return s


def env_str(name: str, default: str = "") -> str:
    v = os.environ.get(name)
    return _strip_quotes(v) if v is not None else default


def env_int(name: str, default: int) -> int:
    v = env_str(name, "")
    if not v:
        return default
    try:
        return int(v)
    except ValueError:
        return default


def env_float(name: str, default: float) -> float:
    v = env_str(name, "")
    if not v:
        return default
    try:
        return float(v)
    except ValueError:
        return default


def env_bool(name: str, default: bool = False) -> bool:
    v = env_str(name, "")
    if not v:
        return default
    return v.lower() in ("1", "true", "yes", "y", "on")


def env_list_int(name: str, default: List[int]) -> List[int]:
    v = env_str(name, "")
    if not v:
        return default
    try:
        return [int(x.strip()) for x in v.split(",") if x.strip()]
    except ValueError:
        return default


# ============================================================================
# CONFIGURATION
# ============================================================================
TOKEN = env_str("TELEGRAM_TOKEN", "")
CLAIM_CODE = env_str("CLAIM_CODE", "")

STATE_FILE = Path(env_str("STATE_FILE", "state.json"))
STATE_FALLBACK = Path("/tmp/dispatch_bot_state.json")

TRIGGERS: Set[str] = {t.strip().lower() for t in env_str("TRIGGERS", "eta,1717").split(",") if t.strip()}

NOMINATIM_USER_AGENT = env_str("NOMINATIM_USER_AGENT", "dispatch-eta-bot/1.0")
NOMINATIM_MIN_INTERVAL = env_float("NOMINATIM_MIN_INTERVAL", 1.1)

ETA_ALL_MAX = env_int("ETA_ALL_MAX", 6)
ALERT_TTL_SECONDS = env_int("ALERT_TTL_SECONDS", 25)
DELETEALL_DEFAULT = env_int("DELETEALL_DEFAULT", 300)

# Geofence settings
GEOFENCE_MILES = env_float("GEOFENCE_MILES", 5.0)
GEOFENCE_METERS = GEOFENCE_MILES * METERS_PER_MILE

# Reminder settings
REMINDER_DOC_AFTER_MIN = env_int("REMINDER_DOC_AFTER_MIN", 15)
REMINDER_THRESHOLDS_MIN = env_list_int("REMINDER_THRESHOLDS_MIN", [60, 30, 10])

# Schedule grace period
SCHEDULE_GRACE_MIN = env_int("SCHEDULE_GRACE_MIN", 30)

DEBUG = env_bool("DEBUG", False)


def log(msg: str) -> None:
    if DEBUG:
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] [bot] {msg}", flush=True)


# ============================================================================
# GLOBALS
# ============================================================================
TF = TimezoneFinder()
NOM_URL = "https://nominatim.openstreetmap.org/search"
OSRM_URL = "https://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}"

_state_lock = asyncio.Lock()
_geo_lock = asyncio.Lock()
_geo_last = 0.0
_background_tasks: Set[asyncio.Task] = set()


# ============================================================================
# TIME HELPERS
# ============================================================================
def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def now_iso() -> str:
    return now_utc().isoformat()


def safe_tz(name: str) -> timezone:
    if not name:
        return timezone.utc
    try:
        return ZoneInfo(name)
    except Exception:
        return timezone.utc


def h(x: Any) -> str:
    return html.escape("" if x is None else str(x), quote=False)


def local_stamp(tz_name: str) -> str:
    tz = safe_tz(tz_name or "UTC")
    return now_utc().astimezone(tz).strftime("%Y-%m-%d %H:%M")


def week_key(dt: datetime) -> str:
    iso = dt.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def money(x: Optional[float]) -> str:
    if x is None:
        return "-"
    try:
        return f"${float(x):,.0f}"
    except (ValueError, TypeError):
        return str(x)


def parse_appointment_time(time_str: str, tz_name: str) -> Optional[datetime]:
    """Parse appointment time string to datetime."""
    if not time_str:
        return None
    
    time_str = time_str.strip()
    tz = safe_tz(tz_name)
    now = now_utc().astimezone(tz)
    
    formats = [
        "%m/%d/%Y %H:%M",
        "%m-%d-%Y %H:%M",
        "%Y-%m-%d %H:%M",
        "%m/%d/%y %H:%M",
        "%m/%d %H:%M",
        "%H:%M",
    ]
    
    for fmt in formats:
        try:
            parsed = datetime.strptime(time_str, fmt)
            if fmt == "%H:%M":
                parsed = parsed.replace(year=now.year, month=now.month, day=now.day)
            elif fmt == "%m/%d %H:%M":
                parsed = parsed.replace(year=now.year)
            return parsed.replace(tzinfo=tz)
        except ValueError:
            continue
    
    time_match = re.search(r'(\d{1,2}):(\d{2})', time_str)
    if time_match:
        hour, minute = int(time_match.group(1)), int(time_match.group(2))
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    
    return None


def minutes_until(target: datetime) -> float:
    delta = target - now_utc()
    return delta.total_seconds() / 60


# ============================================================================
# STATE MANAGEMENT
# ============================================================================
def _migrate_state(st: dict) -> Tuple[dict, bool]:
    changed = False

    if st.get("owner_id") is None and st.get("owner") is not None:
        st["owner_id"] = st.get("owner")
        changed = True
    if st.get("owner") is None and st.get("owner_id") is not None:
        st["owner"] = st.get("owner_id")
        changed = True

    if (not st.get("allowed_chats")) and st.get("allowed"):
        st["allowed_chats"] = st.get("allowed")
        changed = True
    if (not st.get("allowed")) and st.get("allowed_chats"):
        st["allowed"] = st.get("allowed_chats")
        changed = True

    if st.get("last_location") is None and st.get("last") is not None:
        ll = st.get("last") or {}
        st["last_location"] = {
            "lat": ll.get("lat"),
            "lon": ll.get("lon"),
            "tz": ll.get("tz"),
            "updated_at": ll.get("at") or ll.get("updated_at"),
        }
        changed = True

    if (not st.get("geocode_cache")) and st.get("gc"):
        st["geocode_cache"] = st.get("gc")
        changed = True

    if (not st.get("history")) and st.get("hist"):
        st["history"] = st.get("hist")
        changed = True

    st.setdefault("owner_id", None)
    st.setdefault("allowed_chats", [])
    st.setdefault("last_location", None)
    st.setdefault("job", None)
    st.setdefault("focus_i", 0)
    st.setdefault("geocode_cache", {})
    st.setdefault("history", [])
    st.setdefault("last_finished", None)
    st.setdefault("panel_messages", {})
    st.setdefault("reminders_sent", {})
    st.setdefault("geofence_state", {})

    st["owner"] = st.get("owner_id")
    st["allowed"] = st.get("allowed_chats")
    st["gc"] = st.get("geocode_cache")
    st["hist"] = st.get("history")

    return st, changed


def load_state() -> dict:
    global STATE_FILE

    if (not STATE_FILE.exists()) and STATE_FALLBACK.exists():
        STATE_FILE = STATE_FALLBACK

    if STATE_FILE.exists():
        try:
            st = json.loads(STATE_FILE.read_text(encoding="utf-8"))
            if not isinstance(st, dict):
                st = {}
        except (json.JSONDecodeError, OSError) as e:
            log(f"Error loading state: {e}")
            st = {}
    else:
        st = {}

    st, changed = _migrate_state(st)
    if changed:
        try:
            save_state(st)
        except OSError:
            pass
    return st


def save_state(st: dict) -> None:
    global STATE_FILE

    def _write(path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = path.with_suffix(".tmp")
        tmp.write_text(json.dumps(st, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(path)

    try:
        _write(STATE_FILE)
    except OSError as e:
        log(f"save_state failed: {e}")
        STATE_FILE = STATE_FALLBACK
        _write(STATE_FILE)


def is_owner(update: Update, st: dict) -> bool:
    u = update.effective_user
    return bool(u and st.get("owner_id") and u.id == st["owner_id"])


def chat_allowed(update: Update, st: dict) -> bool:
    chat = update.effective_chat
    if not chat:
        return False
    if chat.type == "private":
        return is_owner(update, st)
    return chat.id in set(st.get("allowed_chats") or [])


def get_broadcast_chats(st: dict) -> List[int]:
    chats = list(st.get("allowed_chats") or [])
    owner_id = st.get("owner_id")
    if owner_id and owner_id not in chats:
        chats.append(owner_id)
    return chats


# ============================================================================
# GEOCODE / ROUTING
# ============================================================================
def hav_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * EARTH_RADIUS_M * math.asin(math.sqrt(a))


def hav_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    return hav_m(lat1, lon1, lat2, lon2) * MILES_PER_METER


def fallback_seconds(dist_m: float) -> float:
    km = dist_m / 1000.0
    speed = 55 if km < 80 else (85 if km < 320 else 105)
    return (km / speed) * 3600.0


def fmt_dur(seconds: float) -> str:
    seconds = max(0, int(seconds))
    m = seconds // 60
    h_ = m // 60
    m = m % 60
    return f"{h_}h {m}m" if h_ else f"{m}m"


def fmt_mi(meters: float) -> str:
    mi = meters * MILES_PER_METER
    return f"{mi:.1f} mi" if mi < 10 else f"{mi:.0f} mi"


def addr_variants(addr: str) -> List[str]:
    a = " ".join((addr or "").split())
    if not a:
        return []
    
    out = [a]
    parts = [p.strip() for p in a.split(",") if p.strip()]
    
    if len(parts) >= 2:
        out.append(", ".join(parts[1:]))
    
    cleaned = re.sub(r"\b(?:suite|ste|unit|#)\s*[\w\-]+\b", "", a, flags=re.I).strip()
    if cleaned:
        out.append(cleaned)
    
    if len(parts) >= 2:
        out.append(", ".join(parts[-2:]))
    
    if "usa" not in a.lower():
        out.append(a + ", USA")

    seen: Set[str] = set()
    res: List[str] = []
    for x in out:
        x = " ".join(x.split())
        if x and x not in seen:
            seen.add(x)
            res.append(x)
    return res


async def geocode_cached(st: dict, addr: str) -> Optional[Tuple[float, float, str]]:
    cache = st.get("geocode_cache") or {}
    
    if addr in cache and isinstance(cache[addr], dict):
        try:
            v = cache[addr]
            return float(v["lat"]), float(v["lon"]), (v.get("tz") or "UTC")
        except (KeyError, ValueError, TypeError):
            pass

    if not NOMINATIM_USER_AGENT:
        return None

    headers = {"User-Agent": NOMINATIM_USER_AGENT}
    
    async with httpx.AsyncClient(timeout=15, headers=headers) as client:
        for q in addr_variants(addr):
            async with _geo_lock:
                global _geo_last
                wait = (_geo_last + NOMINATIM_MIN_INTERVAL) - time.monotonic()
                if wait > 0:
                    await asyncio.sleep(wait)
                
                try:
                    r = await client.get(NOM_URL, params={"q": q, "format": "jsonv2", "limit": 1})
                except httpx.RequestError as e:
                    log(f"Geocode error: {e}")
                    continue
                finally:
                    _geo_last = time.monotonic()

            if r.status_code >= 400:
                continue
            
            try:
                js = r.json() or []
            except json.JSONDecodeError:
                continue
                
            if not js:
                continue

            try:
                lat, lon = float(js[0]["lat"]), float(js[0]["lon"])
            except (KeyError, ValueError, TypeError, IndexError):
                continue
                
            tz = TF.timezone_at(lat=lat, lng=lon) or "UTC"
            cache_entry = {"lat": lat, "lon": lon, "tz": tz}

            async with _state_lock:
                st2 = load_state()
                gc = st2.get("geocode_cache") or {}
                gc[addr] = cache_entry
                st2["geocode_cache"] = gc
                st2["gc"] = gc
                save_state(st2)

            st["geocode_cache"] = st.get("geocode_cache") or {}
            st["geocode_cache"][addr] = cache_entry

            return lat, lon, tz

    return None


async def route(origin: Tuple[float, float], dest: Tuple[float, float]) -> Optional[Tuple[float, float]]:
    url = OSRM_URL.format(lon1=origin[1], lat1=origin[0], lon2=dest[1], lat2=dest[0])
    
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(url, params={"overview": "false"})
            if r.status_code >= 400:
                return None
            js = r.json() or {}
            routes = js.get("routes") or []
            if not routes:
                return None
            return float(routes[0]["distance"]), float(routes[0]["duration"])
    except Exception as e:
        log(f"Route error: {e}")
        return None


async def eta_to(st: dict, origin: Tuple[float, float], label: str, addr: str) -> dict:
    g = await geocode_cached(st, addr)
    if not g:
        return {"ok": False, "err": f"Couldn't locate {label}."}
    
    dest = (g[0], g[1])
    r = await route(origin, dest)
    
    if r:
        return {"ok": True, "m": r[0], "s": r[1], "method": "osrm", "tz": g[2], "lat": g[0], "lon": g[1]}
    
    dist = hav_m(origin[0], origin[1], dest[0], dest[1])
    return {"ok": True, "m": dist, "s": fallback_seconds(dist), "method": "approx", "tz": g[2], "lat": g[0], "lon": g[1]}


async def estimate_miles(st: dict, job: dict) -> Optional[float]:
    pu = job.get("pu")
    if not pu or not pu.get("addr"):
        return None
        
    addrs = [pu["addr"]] + [d["addr"] for d in (job.get("del") or []) if d.get("addr")]
    
    coords: List[Tuple[float, float]] = []
    for a in addrs:
        g = await geocode_cached(st, a)
        if not g:
            return None
        coords.append((g[0], g[1]))
    
    if len(coords) < 2:
        return 0.0
    
    total_m = 0.0
    for a, b in zip(coords, coords[1:]):
        r = await route(a, b)
        total_m += r[0] if r else hav_m(a[0], a[1], b[0], b[1])
    
    return total_m * MILES_PER_METER


# ============================================================================
# LOAD PARSING
# ============================================================================
RATE_RE = re.compile(r"\b(?:RATE|PAY)\b\s*:\s*\$?\s*([0-9][0-9,]*(?:\.[0-9]{1,2})?)", re.I)
MILES_RE = re.compile(r"\b(?:LOADED|MILES)\b\s*:\s*([0-9][0-9,]*)", re.I)
PU_TIME_RE = re.compile(r"^\s*PU time:\s*(.+)$", re.I)
DEL_TIME_RE = re.compile(r"^\s*DEL time:\s*(.+)$", re.I)
PU_ADDR_RE = re.compile(r"^\s*PU Address\s*:\s*(.*)$", re.I)
DEL_ADDR_RE = re.compile(r"^\s*DEL Address(?:\s*\d+)?\s*:\s*(.*)$", re.I)
LOAD_NUM_RE = re.compile(r"^\s*Load Number\s*:\s*(.+)$", re.I)
LOAD_DATE_RE = re.compile(r"^\s*Load Date\s*:\s*(.+)$", re.I)
PICKUP_RE = re.compile(r"^\s*Pickup\s*:\s*(.+)$", re.I)
DELIVERY_RE = re.compile(r"^\s*Delivery\s*:\s*(.+)$", re.I)
TIMEISH = re.compile(r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{2}-\d{2}|\d{1,2}:\d{2})\b")


def extract_rate_miles(text: str) -> Tuple[Optional[float], Optional[int]]:
    rate, miles = None, None
    m = RATE_RE.search(text)
    if m:
        try:
            rate = float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    m = MILES_RE.search(text)
    if m:
        try:
            miles = int(m.group(1).replace(",", ""))
        except ValueError:
            pass
    return rate, miles


def take_block(lines: List[str], i: int, first: str) -> Tuple[List[str], int]:
    out = []
    if first.strip():
        out.append(first.strip())
    j = i + 1
    while j < len(lines):
        s = lines[j].strip()
        if not s:
            break
        low = s.lower()
        if low.startswith(("pu time:", "del time:", "pu address", "del address", "pickup:", "delivery:")):
            break
        if set(s) <= {"-"} or set(s) <= {"="}:
            break
        out.append(s)
        j += 1
    return out, j


def init_job(job: dict) -> dict:
    job.setdefault("id", hashlib.sha1(str(time.time()).encode()).hexdigest()[:10])
    job.setdefault("meta", {})
    job.setdefault("created_at", now_iso())
    
    pu = job.setdefault("pu", {})
    pu.setdefault("addr", "")
    pu.setdefault("lines", [])
    pu.setdefault("time", None)
    pu.setdefault("status", {"arr": None, "load": None, "dep": None, "comp": None})
    pu.setdefault("docs", {"pti": False, "bol": False})
    pu.setdefault("geofence", {"entered": None, "exited": None})

    for d in job.setdefault("del", []):
        d.setdefault("addr", "")
        d.setdefault("lines", [])
        d.setdefault("time", None)
        d.setdefault("status", {"arr": None, "del": None, "dep": None, "comp": None, "skip": False})
        d.setdefault("docs", {"pod": False})
        d.setdefault("geofence", {"entered": None, "exited": None})
    
    return job


def normalize_job(job: Optional[dict]) -> Optional[dict]:
    if not job or not isinstance(job, dict):
        return None
    if "pu" not in job or "del" not in job:
        return None
    return init_job(job)


def parse_detailed(text: str) -> Optional[dict]:
    low = text.lower()
    if "pu address" not in low or "del address" not in low:
        return None

    lines = [ln.rstrip() for ln in text.splitlines()]
    pu_time, cur_del_time, pu_addr, pu_lines = None, None, None, None
    dels, load_num, load_date = [], None, None

    for i, ln in enumerate(lines):
        if m := LOAD_NUM_RE.match(ln):
            load_num = m.group(1).strip()
        elif m := LOAD_DATE_RE.match(ln):
            load_date = m.group(1).strip()
        elif m := PU_TIME_RE.match(ln):
            pu_time = m.group(1).strip()
        elif m := DEL_TIME_RE.match(ln):
            cur_del_time = m.group(1).strip()
        elif m := PU_ADDR_RE.match(ln):
            if not pu_addr:
                blk, _ = take_block(lines, i, m.group(1))
                if blk:
                    pu_lines, pu_addr = blk, ", ".join(blk)
        elif m := DEL_ADDR_RE.match(ln):
            blk, _ = take_block(lines, i, m.group(1))
            if blk:
                dels.append({"addr": ", ".join(blk), "lines": blk, "time": cur_del_time})

    if not pu_addr or not dels:
        return None

    rate, miles = extract_rate_miles(text)
    meta = {"rate": rate, "miles": miles}
    if load_num:
        meta["load_number"] = load_num
    if load_date:
        meta["load_date"] = load_date

    jid = hashlib.sha1((pu_addr + "|" + "|".join(d["addr"] for d in dels)).encode()).hexdigest()[:10]
    return init_job({
        "id": jid,
        "meta": meta,
        "pu": {"addr": pu_addr, "lines": pu_lines or [pu_addr], "time": pu_time},
        "del": dels,
    })


def parse_summary(text: str) -> Optional[dict]:
    low = text.lower()
    if "pickup:" not in low or "delivery:" not in low:
        return None

    meta, pu_addr, pu_time, load_date = {}, None, None, None
    dels, pending = [], None

    for ln in [x.strip() for x in text.splitlines() if x.strip()]:
        if m := LOAD_NUM_RE.match(ln):
            meta["load_number"] = m.group(1).strip()
        elif m := LOAD_DATE_RE.match(ln):
            load_date = m.group(1).strip()
        elif m := PICKUP_RE.match(ln):
            v = m.group(1).strip()
            if TIMEISH.search(v):
                pu_time = v
            else:
                pu_addr = v
        elif m := DELIVERY_RE.match(ln):
            v = m.group(1).strip()
            if TIMEISH.search(v):
                if pending and not pending.get("time"):
                    pending["time"] = v
                    pending = None
            else:
                pending = {"addr": v, "lines": [v], "time": None}
                dels.append(pending)

    if not pu_addr or not dels:
        return None

    rate, miles = extract_rate_miles(text)
    if rate:
        meta["rate"] = rate
    if miles:
        meta["miles"] = miles
    if load_date:
        meta["load_date"] = load_date

    jid = hashlib.sha1((str(meta.get("load_number", "")) + "|" + pu_addr).encode()).hexdigest()[:10]
    return init_job({
        "id": jid,
        "meta": meta,
        "pu": {"addr": pu_addr, "lines": [pu_addr], "time": pu_time},
        "del": dels,
    })


def parse_job(text: str) -> Optional[dict]:
    return parse_detailed(text) or parse_summary(text)


# ============================================================================
# WORKFLOW HELPERS
# ============================================================================
def pu_complete(job: dict) -> bool:
    return bool((job.get("pu") or {}).get("status", {}).get("comp"))


def next_incomplete(job: dict, start: int = 0) -> Optional[int]:
    for i, d in enumerate(job.get("del") or []):
        if i >= start and not (d.get("status") or {}).get("comp"):
            return i
    return None


def all_stops_complete(job: dict) -> bool:
    if not pu_complete(job):
        return False
    return all((d.get("status") or {}).get("comp") for d in (job.get("del") or []))


def focus(job: dict, st: dict) -> Tuple[str, int]:
    if not pu_complete(job):
        return "PU", 0
    
    dels = job.get("del") or []
    if not dels:
        return "DEL", 0
    
    i = max(0, min(int(st.get("focus_i") or 0), len(dels) - 1))
    
    if dels[i].get("status", {}).get("comp"):
        ni = next_incomplete(job, i + 1)
        if ni is not None:
            i = ni
    
    return "DEL", i


def load_id_text(job: dict) -> str:
    m = job.get("meta") or {}
    return f"Load {m['load_number']}" if m.get("load_number") else f"Job {job.get('id', '?')}"


def toggle_ts(obj: dict, key: str) -> bool:
    if obj.get(key):
        obj[key] = None
        return False
    obj[key] = now_iso()
    return True


def short_place(lines: List[str], addr: str) -> str:
    for x in reversed(lines or []):
        x = (x or "").strip()
        if x and len(x) <= 70:
            return x
    return (addr or "").strip()[:70]


# ============================================================================
# GEOFENCE SYSTEM
# ============================================================================
class GeofenceEvent(Enum):
    ENTER = "enter"
    EXIT = "exit"


async def check_geofence(st: dict, job: dict, lat: float, lon: float) -> List[Tuple[str, int, GeofenceEvent, str]]:
    """Check geofence for all stops."""
    events = []
    gf_state = st.setdefault("geofence_state", {})
    
    # Check PU
    pu = job.get("pu") or {}
    if pu.get("addr") and not (pu.get("status") or {}).get("comp"):
        g = await geocode_cached(st, pu["addr"])
        if g:
            dist = hav_miles(lat, lon, g[0], g[1])
            key = "PU:0"
            was_inside = gf_state.get(key, False)
            is_inside = dist <= GEOFENCE_MILES
            
            if is_inside and not was_inside:
                events.append(("PU", 0, GeofenceEvent.ENTER, pu["addr"]))
                gf_state[key] = True
            elif not is_inside and was_inside:
                events.append(("PU", 0, GeofenceEvent.EXIT, pu["addr"]))
                gf_state[key] = False
    
    # Check DEL stops
    for i, d in enumerate(job.get("del") or []):
        if d.get("addr") and not (d.get("status") or {}).get("comp"):
            g = await geocode_cached(st, d["addr"])
            if g:
                dist = hav_miles(lat, lon, g[0], g[1])
                key = f"DEL:{i}"
                was_inside = gf_state.get(key, False)
                is_inside = dist <= GEOFENCE_MILES
                
                if is_inside and not was_inside:
                    events.append(("DEL", i, GeofenceEvent.ENTER, d["addr"]))
                    gf_state[key] = True
                elif not is_inside and was_inside:
                    events.append(("DEL", i, GeofenceEvent.EXIT, d["addr"]))
                    gf_state[key] = False
    
    return events


async def process_geofence_events(ctx: ContextTypes.DEFAULT_TYPE, st: dict, job: dict, 
                                   events: List[Tuple[str, int, GeofenceEvent, str]]) -> None:
    """Process geofence events - send alerts and auto-update status."""
    if not events:
        return
    
    tz_name = (st.get("last_location") or {}).get("tz") or "UTC"
    ts = local_stamp(tz_name)
    load_label = load_id_text(job)
    chats = get_broadcast_chats(st)
    
    for stop_type, idx, event, addr in events:
        place = short_place([], addr)
        stop_label = "Pickup" if stop_type == "PU" else f"Delivery {idx + 1}"
        
        if event == GeofenceEvent.ENTER:
            msg = f"📍 <b>ARRIVED: {stop_label}</b>\n{h(place)}\n{h(ts)} — {h(load_label)}"
            # Auto-mark arrival
            if stop_type == "PU":
                ps = job.get("pu", {}).get("status", {})
                if not ps.get("arr"):
                    ps["arr"] = now_iso()
                    job["pu"]["status"] = ps
            else:
                dels = job.get("del") or []
                if idx < len(dels):
                    ds = dels[idx].get("status") or {}
                    if not ds.get("arr"):
                        ds["arr"] = now_iso()
                        dels[idx]["status"] = ds
        else:  # EXIT
            msg = f"🚚 <b>DEPARTED: {stop_label}</b>\n{h(place)}\n{h(ts)} — {h(load_label)}"
            # Auto-mark departure
            if stop_type == "PU":
                ps = job.get("pu", {}).get("status", {})
                if not ps.get("dep"):
                    ps["dep"] = now_iso()
                    job["pu"]["status"] = ps
            else:
                dels = job.get("del") or []
                if idx < len(dels):
                    ds = dels[idx].get("status") or {}
                    if not ds.get("dep"):
                        ds["dep"] = now_iso()
                        dels[idx]["status"] = ds
        
        for chat_id in chats:
            try:
                await ctx.bot.send_message(chat_id=chat_id, text=msg, parse_mode="HTML")
            except TelegramError as e:
                log(f"Geofence alert failed: {e}")


# ============================================================================
# REMINDER SYSTEM
# ============================================================================
def get_pending_reminders(job: dict, st: dict) -> List[dict]:
    """Get all pending reminders for the current job."""
    reminders = []
    tz_name = (st.get("last_location") or {}).get("tz") or "UTC"
    sent = st.get("reminders_sent") or {}
    load_label = load_id_text(job)
    
    # Appointment reminders
    for threshold in sorted(REMINDER_THRESHOLDS_MIN, reverse=True):
        # PU
        pu = job.get("pu") or {}
        if pu.get("time") and not (pu.get("status") or {}).get("comp"):
            appt = parse_appointment_time(pu["time"], tz_name)
            if appt:
                mins = minutes_until(appt)
                key = f"appt:PU:{threshold}"
                if threshold - 5 < mins <= threshold and not sent.get(key):
                    reminders.append({
                        "key": key,
                        "message": f"⏰ <b>PU in ~{int(mins)} min</b>\n{h(pu['time'])}\n{h(short_place(pu.get('lines',[]), pu.get('addr','')))}\n{h(load_label)}"
                    })
        
        # DEL
        for i, d in enumerate(job.get("del") or []):
            if d.get("time") and not (d.get("status") or {}).get("comp"):
                appt = parse_appointment_time(d["time"], tz_name)
                if appt:
                    mins = minutes_until(appt)
                    key = f"appt:DEL:{i}:{threshold}"
                    if threshold - 5 < mins <= threshold and not sent.get(key):
                        reminders.append({
                            "key": key,
                            "message": f"⏰ <b>DEL {i+1} in ~{int(mins)} min</b>\n{h(d['time'])}\n{h(short_place(d.get('lines',[]), d.get('addr','')))}\n{h(load_label)}"
                        })
    
    # Document reminders (after arrival)
    pu = job.get("pu") or {}
    ps = pu.get("status") or {}
    pd = pu.get("docs") or {}
    
    if ps.get("arr") and not ps.get("comp"):
        try:
            arr_time = datetime.fromisoformat(ps["arr"])
            mins_since = (now_utc() - arr_time).total_seconds() / 60
            
            if mins_since >= REMINDER_DOC_AFTER_MIN:
                if not pd.get("pti") and not sent.get("doc:PU:pti"):
                    reminders.append({
                        "key": "doc:PU:pti",
                        "message": f"📋 <b>PTI Reminder</b>\nArrived {int(mins_since)} min ago\n{h(load_label)}"
                    })
                if not pd.get("bol") and not sent.get("doc:PU:bol"):
                    reminders.append({
                        "key": "doc:PU:bol",
                        "message": f"📋 <b>BOL Reminder</b>\nArrived {int(mins_since)} min ago\n{h(load_label)}"
                    })
        except (ValueError, TypeError):
            pass
    
    # DEL POD reminders
    for i, d in enumerate(job.get("del") or []):
        ds = d.get("status") or {}
        dd = d.get("docs") or {}
        
        if ds.get("arr") and not ds.get("comp"):
            try:
                arr_time = datetime.fromisoformat(ds["arr"])
                mins_since = (now_utc() - arr_time).total_seconds() / 60
                
                if mins_since >= REMINDER_DOC_AFTER_MIN and not dd.get("pod"):
                    key = f"doc:DEL:{i}:pod"
                    if not sent.get(key):
                        reminders.append({
                            "key": key,
                            "message": f"📋 <b>POD Reminder - DEL {i+1}</b>\nArrived {int(mins_since)} min ago\n{h(load_label)}"
                        })
            except (ValueError, TypeError):
                pass
    
    # Late warnings
    pu = job.get("pu") or {}
    if pu.get("time") and not (pu.get("status") or {}).get("arr"):
        appt = parse_appointment_time(pu["time"], tz_name)
        if appt:
            mins = minutes_until(appt)
            if -SCHEDULE_GRACE_MIN < mins < 0:
                key = f"late:PU:{int(abs(mins)//10)*10}"
                if not sent.get(key):
                    reminders.append({
                        "key": key,
                        "message": f"🚨 <b>LATE for PU by {int(abs(mins))} min!</b>\n{h(load_label)}"
                    })
    
    return reminders


async def send_reminders(ctx: ContextTypes.DEFAULT_TYPE, st: dict, reminders: List[dict]) -> None:
    """Send reminder messages."""
    if not reminders:
        return
    
    chats = get_broadcast_chats(st)
    sent = st.setdefault("reminders_sent", {})
    
    for r in reminders:
        for chat_id in chats:
            try:
                await ctx.bot.send_message(chat_id=chat_id, text=r["message"], parse_mode="HTML")
            except TelegramError as e:
                log(f"Reminder failed: {e}")
        sent[r["key"]] = True


# ============================================================================
# BACKGROUND JOBS
# ============================================================================
async def reminder_job(ctx: ContextTypes.DEFAULT_TYPE) -> None:
    """Periodic reminder check."""
    async with _state_lock:
        st = load_state()
        job = normalize_job(st.get("job"))
        if not job:
            return
        
        reminders = get_pending_reminders(job, st)
        if reminders:
            await send_reminders(ctx, st, reminders)
            save_state(st)


async def geofence_job(ctx: ContextTypes.DEFAULT_TYPE) -> None:
    """Periodic geofence check."""
    async with _state_lock:
        st = load_state()
        job = normalize_job(st.get("job"))
        loc = st.get("last_location")
        
        if not job or not loc:
            return
        
        # Check location freshness
        updated = loc.get("updated_at")
        if updated:
            try:
                age = (now_utc() - datetime.fromisoformat(updated)).total_seconds() / 60
                if age > 5:
                    return
            except (ValueError, TypeError):
                return
        
        try:
            lat, lon = float(loc["lat"]), float(loc["lon"])
        except (KeyError, ValueError, TypeError):
            return
        
        events = await check_geofence(st, job, lat, lon)
        if events:
            await process_geofence_events(ctx, st, job, events)
            st["job"] = job
            save_state(st)


# ============================================================================
# UI HELPERS
# ============================================================================
async def send_progress_alert(ctx: ContextTypes.DEFAULT_TYPE, chat_id: int, text: str) -> None:
    """Send auto-deleting alert."""
    try:
        m = await ctx.bot.send_message(chat_id=chat_id, text=text, parse_mode="HTML", disable_notification=True)
    except TelegramError:
        return

    if ALERT_TTL_SECONDS <= 0:
        return

    async def _delete():
        await asyncio.sleep(ALERT_TTL_SECONDS)
        try:
            await ctx.bot.delete_message(chat_id=chat_id, message_id=m.message_id)
        except TelegramError:
            pass

    task = asyncio.create_task(_delete())
    _background_tasks.add(task)
    task.add_done_callback(_background_tasks.discard)


def b(label: str, data: str) -> InlineKeyboardButton:
    return InlineKeyboardButton(label, callback_data=data)


def chk(on: bool, label: str) -> str:
    return ("✅ " + label) if on else label


def build_finished_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[b("📊 Catalog", "SHOW:CAT")]])


def build_keyboard(job: dict, st: dict) -> InlineKeyboardMarkup:
    stage, i = focus(job, st)
    pu = job.get("pu") or {}
    ps = pu.get("status") or {}
    pd = pu.get("docs") or {}
    rows = []

    if stage == "PU":
        rows.append([
            b(chk(bool(ps.get("arr")), "Arrived"), "PU:A"),
            b(chk(bool(ps.get("load")), "Loaded"), "PU:L"),
            b(chk(bool(ps.get("dep")), "Departed"), "PU:D"),
        ])
        rows.append([
            b(chk(bool(pd.get("pti")), "PTI"), "DOC:PTI"),
            b(chk(bool(pd.get("bol")), "BOL"), "DOC:BOL"),
            b(chk(bool(ps.get("comp")), "PU Done"), "PU:C"),
        ])
    else:
        dels = job.get("del") or []
        d = dels[i] if i < len(dels) else {}
        ds = d.get("status") or {}
        dd = d.get("docs") or {}
        lbl = f"{i+1}/{len(dels)}"

        rows.append([
            b(chk(bool(ds.get("arr")), f"Arr {lbl}"), "DEL:A"),
            b(chk(bool(ds.get("del")), "Delivered"), "DEL:DL"),
            b(chk(bool(ds.get("dep")), "Departed"), "DEL:D"),
        ])
        rows.append([
            b(chk(bool(dd.get("pod")), "POD"), "DOC:POD"),
            b(chk(bool(ds.get("comp")), "Done"), "DEL:C"),
            b("Skip", "DEL:S"),
        ])
        
        if len(dels) > 1:
            nav = []
            if i > 0:
                nav.append(b("◀️", f"NAV:{i-1}"))
            if i < len(dels) - 1:
                nav.append(b("▶️", f"NAV:{i+1}"))
            if nav:
                rows.append(nav)

    rows.append([b("📍 ETA", "ETA:A"), b("📍 All", "ETA:ALL")])
    rows.append([b("📊 Catalog", "SHOW:CAT"), b("✅ Finish", "JOB:FIN")])
    return InlineKeyboardMarkup(rows)


# ============================================================================
# FINISH + HISTORY
# ============================================================================
def week_totals(hist: List[dict], wk: str) -> Tuple[int, float, float]:
    count, rate, miles = 0, 0.0, 0.0
    for r in hist:
        if r.get("week") != wk:
            continue
        count += 1
        if isinstance(r.get("rate"), (int, float)):
            rate += float(r["rate"])
        m = r.get("posted_miles") or r.get("est_miles")
        if isinstance(m, (int, float)):
            miles += float(m)
    return count, rate, miles


async def finish_load(update: Update, ctx: ContextTypes.DEFAULT_TYPE, source: str) -> Optional[Tuple[dict, dict]]:
    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            msg = "Owner only."
            if source == "cb" and update.callback_query:
                await update.callback_query.answer(msg, show_alert=True)
            else:
                await update.effective_message.reply_text(msg)
            return None

        job = normalize_job(st.get("job"))
        if not job:
            msg = "No active load."
            if source == "cb" and update.callback_query:
                await update.callback_query.answer(msg, show_alert=True)
            else:
                await update.effective_message.reply_text(msg)
            return None

        tz_name = (st.get("last_location") or {}).get("tz") or "UTC"
        
    est = await estimate_miles(st, job)
    
    async with _state_lock:
        st = load_state()
        job = normalize_job(st.get("job"))
        if not job:
            return None
            
        tz_name = (st.get("last_location") or {}).get("tz") or "UTC"
        dt = now_utc().astimezone(safe_tz(tz_name))
        wk = week_key(dt)

        meta = job.get("meta") or {}
        pu = job.get("pu") or {}
        dels = job.get("del") or []

        rec = {
            "week": wk,
            "completed": dt.strftime("%Y-%m-%d %H:%M"),
            "completed_utc": now_iso(),
            "tz": tz_name,
            "load_number": meta.get("load_number") or "",
            "job_id": job.get("id"),
            "load_date": meta.get("load_date"),
            "pu_time": pu.get("time"),
            "pickup": pu.get("addr") or "",
            "deliveries": " | ".join(d.get("addr", "") for d in dels),
            "del_times": " | ".join((d.get("time") or "-") for d in dels),
            "stops": len(dels),
            "rate": meta.get("rate"),
            "posted_miles": meta.get("miles"),
            "est_miles": est,
        }

        hist = list(st.get("history") or [])
        hist.append(rec)
        st["history"] = hist[-MAX_HISTORY_RECORDS:]
        st["hist"] = st["history"]
        st["last_finished"] = rec
        st["job"] = None
        st["focus_i"] = 0
        st["reminders_sent"] = {}
        st["geofence_state"] = {}
        
        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id:
            pm = st.get("panel_messages") or {}
            pm.pop(str(chat_id), None)
            st["panel_messages"] = pm
        
        save_state(st)

    cnt, tot_rate, tot_mi = week_totals(st["history"], wk)
    rec["_wk_count"] = cnt
    rec["_wk_rate"] = tot_rate
    rec["_wk_miles"] = tot_mi
    return rec, st


# ============================================================================
# EXCEL EXPORT
# ============================================================================
def try_parse_date(s) -> Optional[date]:
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(str(s).strip(), fmt).date()
        except ValueError:
            pass
    return None


def try_parse_dt(s) -> Optional[datetime]:
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M", "%m/%d/%y %H:%M"):
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(str(s).strip())
    except ValueError:
        return None


def autosize_cols(ws, min_w=10, max_w=60):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, 1):
            if v:
                widths[i] = max(widths.get(i, 0), len(str(v)))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = max(min_w, min(max_w, w + 2))


def safe_sheet_name(name: str, existing: Set[str]) -> str:
    name = name[:EXCEL_SHEET_NAME_MAX_LEN]
    for c in r'[]:*?/\\':
        name = name.replace(c, '_')
    base, counter = name, 1
    while name in existing:
        name = f"{base[:EXCEL_SHEET_NAME_MAX_LEN-3]}_{counter}"
        counter += 1
    return name


def write_sheet(wb: Workbook, wk: str, records: List[dict], existing: Set[str]):
    name = safe_sheet_name(wk, existing)
    existing.add(name)
    ws = wb.create_sheet(title=name)
    
    records = sorted(records, key=lambda r: try_parse_dt(r.get("completed_utc")) or datetime(1970,1,1))
    
    ws.append([f"Weekly Loads — {wk}"])
    ws["A1"].font = Font(bold=True, size=14)
    
    headers = ["Completed", "TZ", "Load #", "Job ID", "Load Date", "PU Time", 
               "Pickup", "DEL Times", "Deliveries", "Stops", "Rate", "Posted Mi", "Est Mi", "$/Mi"]
    ws.append(headers)
    for c in ws[2]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")

    sum_rate, sum_mi = 0.0, 0.0
    for r in records:
        rate = r.get("rate")
        posted = r.get("posted_miles")
        est = r.get("est_miles")
        rpm = (float(rate)/float(est)) if isinstance(rate,(int,float)) and isinstance(est,(int,float)) and est > 0 else None
        
        ws.append([
            try_parse_dt(r.get("completed")) or r.get("completed",""),
            r.get("tz",""), r.get("load_number",""), r.get("job_id",""),
            try_parse_date(r.get("load_date")) or r.get("load_date",""),
            try_parse_dt(r.get("pu_time")) or r.get("pu_time",""),
            r.get("pickup",""), r.get("del_times",""), r.get("deliveries",""),
            r.get("stops",""),
            float(rate) if isinstance(rate,(int,float)) else None,
            float(posted) if isinstance(posted,(int,float)) else None,
            float(est) if isinstance(est,(int,float)) else None,
            rpm
        ])
        if isinstance(rate,(int,float)):
            sum_rate += float(rate)
        m = posted if isinstance(posted,(int,float)) else (est if isinstance(est,(int,float)) else None)
        if m:
            sum_mi += float(m)

    ws.append([])
    ws.append(["TOTAL","","","","","","","","","", sum_rate,"",sum_mi, sum_rate/sum_mi if sum_mi else None])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E2EFDA")

    for row in ws.iter_rows(min_row=3):
        if len(row)>0 and isinstance(row[0].value, datetime):
            row[0].number_format = "yyyy-mm-dd hh:mm"
        if len(row)>10 and row[10].value is not None:
            row[10].number_format = '"$"#,##0'
        if len(row)>13 and row[13].value is not None:
            row[13].number_format = '"$"0.00'

    ws.freeze_panes = "A3"
    autosize_cols(ws)


def make_xlsx(records: List[dict], wk: str) -> Tuple[bytes, str]:
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    
    existing: Set[str] = set()
    if wk == "ALL":
        by_wk: Dict[str, List[dict]] = {}
        for r in records:
            by_wk.setdefault(r.get("week","?"), []).append(r)
        for w in sorted(by_wk):
            write_sheet(wb, w, by_wk[w], existing)
        fn = "load_catalog_ALL.xlsx"
    else:
        write_sheet(wb, wk, records, existing)
        fn = f"load_catalog_{wk}.xlsx"
    
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), fn


def parse_catalog_arg(args: List[str], tz_name: str) -> str:
    wk = week_key(now_utc().astimezone(safe_tz(tz_name)))
    if not args:
        return wk
    a = args[0].strip().lower()
    if a == "all":
        return "ALL"
    if a in ("last", "prev"):
        return week_key(now_utc().astimezone(safe_tz(tz_name)) - timedelta(days=7))
    if re.fullmatch(r"\d{4}-w\d{2}", a):
        return a.upper().replace("w","W")
    return wk


async def send_catalog(update: Update, ctx: ContextTypes.DEFAULT_TYPE, from_cb=False):
    async with _state_lock:
        st = load_state()

    if not is_owner(update, st):
        msg = "Owner only."
        if from_cb and update.callback_query:
            await update.callback_query.answer(msg, show_alert=True)
        else:
            await update.effective_message.reply_text(msg)
        return

    hist = list(st.get("history") or [])
    if not hist:
        msg = "No finished loads yet."
        if from_cb and update.callback_query:
            await update.callback_query.answer(msg, show_alert=True)
        else:
            await update.effective_message.reply_text(msg)
        return

    tz_name = (st.get("last_location") or {}).get("tz") or "UTC"
    wk = parse_catalog_arg(getattr(ctx, "args", []) or [], tz_name)
    
    records = hist if wk == "ALL" else [r for r in hist if r.get("week") == wk]
    if not records:
        msg = "No records for that week."
        if from_cb and update.callback_query:
            await update.callback_query.answer(msg, show_alert=True)
        else:
            await update.effective_message.reply_text(msg)
        return

    xlsx, fn = make_xlsx(records, wk)
    buf = io.BytesIO(xlsx)
    buf.name = fn
    
    cnt, tot_rate, tot_mi = week_totals(records, wk if wk != "ALL" else "")
    if wk == "ALL":
        cnt = len(records)
        tot_rate = sum(r.get("rate",0) or 0 for r in records if isinstance(r.get("rate"),(int,float)))
        tot_mi = sum((r.get("posted_miles") or r.get("est_miles") or 0) for r in records 
                     if isinstance(r.get("posted_miles"),(int,float)) or isinstance(r.get("est_miles"),(int,float)))
    
    await ctx.bot.send_document(
        chat_id=update.effective_chat.id,
        document=buf, filename=fn,
        caption=f"📊 <b>{wk}</b>\n{cnt} loads · {money(tot_rate)} · {int(tot_mi)} mi",
        parse_mode="HTML"
    )


# ============================================================================
# TELEGRAM COMMANDS
# ============================================================================
async def start_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msg = f"""🚚 <b>Trucker Dispatch Bot</b>
<i>v{BOT_VERSION}</i>

<b>Setup:</b>
1️⃣ DM: /claim &lt;code&gt;
2️⃣ DM: /update (share location)
3️⃣ Group: /allowhere

<b>Commands:</b>
• <code>eta</code> or <code>1717</code> — ETA
• /panel — Control panel
• /finish — Complete load
• /catalog [week|all] — Excel export
• /skip — Skip stop
• /status — Bot info

<b>Features:</b>
📍 Geofence: {GEOFENCE_MILES} mi radius
⏰ Appt alerts: {','.join(map(str,REMINDER_THRESHOLDS_MIN))} min
📋 Doc reminders: {REMINDER_DOC_AFTER_MIN} min after arrival
⏱️ Grace period: {SCHEDULE_GRACE_MIN} min

<b>Triggers:</b> {', '.join(sorted(TRIGGERS))}"""
    await update.effective_message.reply_text(msg, parse_mode="HTML")


async def ping_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.effective_message.reply_text(f"🏓 pong — {BOT_VERSION}")


async def status_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()

    uid = update.effective_user.id if update.effective_user else None
    loc = st.get("last_location")
    job = normalize_job(st.get("job"))
    
    loc_age = "N/A"
    if loc and loc.get("updated_at"):
        try:
            age = int((now_utc() - datetime.fromisoformat(loc["updated_at"])).total_seconds() / 60)
            loc_age = f"{age} min ago"
        except:
            pass

    await update.effective_message.reply_text(f"""<b>🚚 Status</b>
<b>Version:</b> {h(BOT_VERSION)}
<b>Your ID:</b> {h(uid)}
<b>Owner:</b> {h(st.get('owner_id'))}
<b>Allowed here:</b> {'✅' if chat_allowed(update, st) else '❌'}
<b>Location:</b> {'✅' if loc else '❌'} ({loc_age})
<b>Active load:</b> {h(load_id_text(job)) if job else '❌'}
<b>History:</b> {len(st.get('history') or [])} loads

<b>Settings:</b>
• Geofence: {GEOFENCE_MILES} mi
• Appt alerts: {REMINDER_THRESHOLDS_MIN}
• Doc reminder: {REMINDER_DOC_AFTER_MIN} min
• Grace: {SCHEDULE_GRACE_MIN} min""", parse_mode="HTML")


async def claim_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not update.effective_chat or update.effective_chat.type != "private":
        await update.effective_message.reply_text("⚠️ DM me: /claim <code>")
        return
    if not CLAIM_CODE:
        await update.effective_message.reply_text("⚠️ CLAIM_CODE not set.")
        return
    if " ".join(ctx.args or []).strip() != CLAIM_CODE:
        await update.effective_message.reply_text("❌ Wrong code.")
        return

    async with _state_lock:
        st = load_state()
        st["owner_id"] = st["owner"] = update.effective_user.id
        save_state(st)
    await update.effective_message.reply_text("✅ Owner set! Now /update to share location.")


async def allowhere_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            await update.effective_message.reply_text("⚠️ Owner only.")
            return
        chat = update.effective_chat
        if not chat or chat.type == "private":
            await update.effective_message.reply_text("⚠️ Run in a group.")
            return
        allowed = set(st.get("allowed_chats") or [])
        allowed.add(chat.id)
        st["allowed_chats"] = st["allowed"] = sorted(allowed)
        save_state(st)
    await update.effective_message.reply_text("✅ Group allowed!")


async def update_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            await update.effective_message.reply_text("⚠️ Owner only.")
            return
    if not update.effective_chat or update.effective_chat.type != "private":
        await update.effective_message.reply_text("📍 DM me /update")
        return

    kb = [[KeyboardButton("📍 Send Location", request_location=True)]]
    await update.effective_message.reply_text(
        "📍 Tap to share location.\n💡 Use <i>Live Location</i> for auto geofence!",
        reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True, one_time_keyboard=True),
        parse_mode="HTML"
    )


async def on_location(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msg = update.effective_message
    if not msg or not msg.location:
        return

    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            return
        
        lat, lon = msg.location.latitude, msg.location.longitude
        tz = TF.timezone_at(lat=lat, lng=lon) or "UTC"
        st["last_location"] = st["last"] = {"lat": lat, "lon": lon, "tz": tz, "updated_at": now_iso()}
        
        job = normalize_job(st.get("job"))
        if job:
            events = await check_geofence(st, job, lat, lon)
            if events:
                await process_geofence_events(ctx, st, job, events)
                st["job"] = job
        
        save_state(st)

    if update.effective_chat and update.effective_chat.type == "private":
        await msg.reply_text("✅ Location saved.", reply_markup=ReplyKeyboardRemove())


async def panel_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()

    if not chat_allowed(update, st):
        await update.effective_message.reply_text("⚠️ Run /allowhere first.")
        return

    job = normalize_job(st.get("job"))
    if not job:
        await update.effective_message.reply_text("📭 No active load. Forward a load sheet.")
        return

    stage, i = focus(job, st)
    meta = job.get("meta") or {}
    
    lines = [f"<b>{h(load_id_text(job))}</b>"]
    if meta.get("rate"):
        lines.append(f"💰 {money(meta['rate'])}")
    
    if stage == "PU":
        pu = job.get("pu") or {}
        lines.append(f"\n<b>📍 Pickup</b>")
        lines.append(h(short_place(pu.get("lines",[]), pu.get("addr",""))))
        if pu.get("time"):
            lines.append(f"⏰ {h(pu['time'])}")
    else:
        dels = job.get("del") or []
        if i < len(dels):
            d = dels[i]
            lines.append(f"\n<b>📍 Delivery {i+1}/{len(dels)}</b>")
            lines.append(h(short_place(d.get("lines",[]), d.get("addr",""))))
            if d.get("time"):
                lines.append(f"⏰ {h(d['time'])}")

    m = await update.effective_message.reply_text("\n".join(lines), parse_mode="HTML", reply_markup=build_keyboard(job, st))

    if update.effective_chat:
        async with _state_lock:
            st2 = load_state()
            pm = st2.setdefault("panel_messages", {})
            pm[str(update.effective_chat.id)] = m.message_id
            save_state(st2)


async def finish_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    out = await finish_load(update, ctx, "cmd")
    if not out:
        return
    rec, st = out

    rate_txt = money(rec.get("rate") if isinstance(rec.get("rate"),(int,float)) else None)
    id_txt = rec.get("load_number") or rec.get("job_id") or ""
    
    await send_progress_alert(ctx, update.effective_chat.id, f"✅ <b>Finished</b> {h(id_txt)} · {h(rate_txt)}")

    wk_cnt = int(rec.get("_wk_count",0))
    wk_rate = float(rec.get("_wk_rate",0))
    wk_mi = float(rec.get("_wk_miles",0))

    report = f"""✅ <b>Load Complete!</b>

<b>{h(id_txt)}</b> · {h(rate_txt)}

📊 <b>Week {h(rec.get('week'))}:</b>
• {wk_cnt} loads
• {money(wk_rate)} gross
• {int(wk_mi)} miles
• {money(wk_rate/wk_mi if wk_mi else 0)}/mi"""

    chat_id = update.effective_chat.id if update.effective_chat else None
    if chat_id:
        pm = st.get("panel_messages") or {}
        msg_id = pm.get(str(chat_id))
        if msg_id:
            try:
                await ctx.bot.edit_message_text(chat_id=chat_id, message_id=int(msg_id), 
                    text=report, parse_mode="HTML", reply_markup=build_finished_kb())
                return
            except TelegramError:
                pass
        await ctx.bot.send_message(chat_id=chat_id, text=report, parse_mode="HTML", reply_markup=build_finished_kb())


async def catalog_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await send_catalog(update, ctx, from_cb=False)


async def skip_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            await update.effective_message.reply_text("⚠️ Owner only.")
            return

        job = normalize_job(st.get("job"))
        if not job:
            await update.effective_message.reply_text("📭 No active load.")
            return

        stage, i = focus(job, st)
        if stage != "DEL":
            await update.effective_message.reply_text("⚠️ Complete PU first.")
            return

        dels = job.get("del") or []
        if not dels or i >= len(dels):
            await update.effective_message.reply_text("⚠️ No stops to skip.")
            return

        ds = dels[i].setdefault("status", {})
        ds["skip"] = True
        ds["comp"] = ds.get("comp") or now_iso()
        
        ni = next_incomplete(job, i + 1)
        if ni is not None:
            st["focus_i"] = ni

        st["job"] = job
        save_state(st)

    await send_progress_alert(ctx, update.effective_chat.id, f"⏭️ Skipped DEL {i+1}/{len(dels)}")
    await update.effective_message.reply_text("✅ Skipped. /panel to continue.")


async def deleteall_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            await update.effective_message.reply_text("⚠️ Owner only.")
            return

    chat = update.effective_chat
    if not chat or chat.type == "private":
        await update.effective_message.reply_text("Can't clear DM history.")
        return

    n = DELETEALL_DEFAULT
    if ctx.args:
        try:
            n = max(1, min(MAX_DELETEALL_MESSAGES, int(ctx.args[0])))
        except ValueError:
            pass

    notice = await update.effective_message.reply_text(f"🧹 Deleting up to {n} messages...")
    for mid in range(notice.message_id, max(1, notice.message_id - n), -1):
        try:
            await ctx.bot.delete_message(chat_id=chat.id, message_id=mid)
        except (Forbidden, BadRequest):
            break
        await asyncio.sleep(DELETE_MESSAGE_DELAY_SEC)


async def leave_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            await update.effective_message.reply_text("⚠️ Owner only.")
            return
        chat = update.effective_chat
        if not chat or chat.type == "private":
            await update.effective_message.reply_text("Run in a group.")
            return
        allowed = set(st.get("allowed_chats") or [])
        allowed.discard(chat.id)
        st["allowed_chats"] = st["allowed"] = sorted(allowed)
        save_state(st)

    await update.effective_message.reply_text("👋 Leaving...")
    try:
        await ctx.bot.leave_chat(chat.id)
    except TelegramError:
        pass


# ============================================================================
# ETA COMMAND
# ============================================================================
async def send_eta(update: Update, ctx: ContextTypes.DEFAULT_TYPE, which: str):
    async with _state_lock:
        st = load_state()

    if not chat_allowed(update, st):
        return

    loc = st.get("last_location")
    if not loc:
        await update.effective_message.reply_text("📍 No location. Owner: /update")
        return

    try:
        origin = (float(loc["lat"]), float(loc["lon"]))
    except (KeyError, ValueError, TypeError):
        await update.effective_message.reply_text("⚠️ Invalid location.")
        return
        
    tz_now = loc.get("tz") or "UTC"
    tz = safe_tz(tz_now)

    await ctx.bot.send_location(chat_id=update.effective_chat.id, latitude=origin[0], longitude=origin[1])

    job = normalize_job(st.get("job"))
    if not job:
        await update.effective_message.reply_text(
            f"⏱ <b>ETA</b>\n{h(datetime.now(tz).strftime('%H:%M'))} ({h(tz_now)})\n\n<i>No active load.</i>",
            parse_mode="HTML"
        )
        return

    which = (which or "AUTO").upper()

    if which == "ALL":
        lines = [f"<b>{h(load_id_text(job))}</b>", ""]
        pu = job.get("pu") or {}
        stops = [("PU", pu.get("addr",""), pu.get("lines",[]), pu.get("time"), bool((pu.get("status") or {}).get("comp")))]
        for j, d in enumerate((job.get("del") or [])[:ETA_ALL_MAX]):
            stops.append((f"D{j+1}", d.get("addr",""), d.get("lines",[]), d.get("time"), bool((d.get("status") or {}).get("comp"))))

        for lab, addr, addr_lines, appt, done in stops:
            if done:
                lines.append(f"✅ <b>{h(lab)}:</b> <s>{h(short_place(addr_lines, addr)[:30])}</s>")
                continue
            r = await eta_to(st, origin, lab, addr)
            place = short_place(addr_lines, addr)
            if r.get("ok"):
                arr = (now_utc().astimezone(tz) + timedelta(seconds=float(r["s"]))).strftime("%H:%M")
                tag = " ≈" if r.get("method") == "approx" else ""
                appt_txt = f" | {appt}" if appt else ""
                lines.append(f"<b>{h(lab)}:</b> {h(fmt_dur(r['s']))}{h(tag)} · {h(fmt_mi(r['m']))} · ~{h(arr)}{h(appt_txt)}")
            else:
                lines.append(f"<b>{h(lab)}:</b> ⚠️ {h(r.get('err'))}")

        await update.effective_message.reply_text("\n".join(lines), parse_mode="HTML", reply_markup=build_keyboard(job, st))
        return

    # Single ETA
    stage, i = focus(job, st)
    if stage == "PU":
        pu = job.get("pu") or {}
        addr, addr_lines, appt = pu.get("addr",""), pu.get("lines",[]), pu.get("time")
        stop_label = "PU"
    else:
        dels = job.get("del") or []
        d = dels[i] if i < len(dels) else {}
        addr, addr_lines, appt = d.get("addr",""), d.get("lines",[]), d.get("time")
        stop_label = f"DEL {i+1}/{len(dels)}" if dels else "DEL"

    r = await eta_to(st, origin, stop_label, addr)
    place = short_place(addr_lines, addr)

    if r.get("ok"):
        arr_time = now_utc().astimezone(tz) + timedelta(seconds=float(r["s"]))
        arr = arr_time.strftime("%H:%M")
        tag = " (approx)" if r.get("method") == "approx" else ""
        
        out = [
            f"⏱ <b>ETA: {h(fmt_dur(r['s']))}</b>{h(tag)}",
            "",
            f"<b>{h(stop_label)}</b> — {h(load_id_text(job))}",
            f"📍 {h(place)}",
            f"🚚 {h(fmt_mi(r['m']))} · ~{h(arr)} ({h(tz_now)})",
        ]
        if appt:
            appt_dt = parse_appointment_time(appt, tz_now)
            if appt_dt:
                diff = (appt_dt - arr_time).total_seconds() / 60
                if diff >= 0:
                    status = f"✅ On time (+{int(diff)}m early)"
                elif diff >= -SCHEDULE_GRACE_MIN:
                    status = f"⚠️ Close ({int(abs(diff))}m late)"
                else:
                    status = f"🚨 LATE {int(abs(diff))}m!"
                out.append(f"⏰ Appt: {h(appt)} — {status}")
            else:
                out.append(f"⏰ Appt: {h(appt)}")
        
        await update.effective_message.reply_text("\n".join(out), parse_mode="HTML", reply_markup=build_keyboard(job, st))
    else:
        await update.effective_message.reply_text(
            f"<b>{h(load_id_text(job))}</b>\n⏱ ⚠️ {h(r.get('err'))}\n📍 {h(place)}",
            parse_mode="HTML", reply_markup=build_keyboard(job, st)
        )


# ============================================================================
# CALLBACK HANDLER
# ============================================================================
async def on_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not q or not q.data:
        return

    data = q.data

    async with _state_lock:
        st = load_state()

    if not chat_allowed(update, st):
        await q.answer("Not allowed.", show_alert=False)
        return

    # ETA
    if data.startswith("ETA:"):
        await q.answer("Computing...")
        await send_eta(update, ctx, data.split(":",1)[1])
        return

    # Catalog
    if data == "SHOW:CAT":
        await q.answer()
        await send_catalog(update, ctx, from_cb=True)
        return

    # Navigation
    if data.startswith("NAV:"):
        try:
            new_i = int(data.split(":")[1])
            async with _state_lock:
                st2 = load_state()
                st2["focus_i"] = new_i
                save_state(st2)
                job = normalize_job(st2.get("job"))
            await q.answer(f"Stop {new_i + 1}")
            if job:
                try:
                    await q.edit_message_reply_markup(reply_markup=build_keyboard(job, st2))
                except TelegramError:
                    pass
        except (ValueError, IndexError):
            await q.answer("Invalid")
        return

    # Finish
    if data == "JOB:FIN":
        if not is_owner(update, st):
            await q.answer("Owner only.", show_alert=True)
            return
        await q.answer("Finishing...")
        out = await finish_load(update, ctx, "cb")
        if not out:
            return
        rec, _ = out

        rate_txt = money(rec.get("rate") if isinstance(rec.get("rate"),(int,float)) else None)
        id_txt = rec.get("load_number") or rec.get("job_id") or ""
        await send_progress_alert(ctx, update.effective_chat.id, f"✅ <b>Finished</b> {h(id_txt)} · {h(rate_txt)}")

        wk_cnt = int(rec.get("_wk_count",0))
        wk_rate = float(rec.get("_wk_rate",0))
        wk_mi = float(rec.get("_wk_miles",0))

        report = f"✅ <b>Load Complete!</b>\n\n<b>{h(id_txt)}</b> · {h(rate_txt)}\n\n📊 <b>Week {h(rec.get('week'))}:</b> {wk_cnt} loads · {money(wk_rate)} · {int(wk_mi)} mi"

        try:
            await q.edit_message_text(text=report, parse_mode="HTML", reply_markup=build_finished_kb())
        except TelegramError:
            pass
        return

    # Status buttons
    async with _state_lock:
        st2 = load_state()
        job = normalize_job(st2.get("job"))
        if not job:
            await q.answer("No active load.", show_alert=True)
            return

        stage, i = focus(job, st2)
        tz_name = (st2.get("last_location") or {}).get("tz") or "UTC"
        ts = local_stamp(tz_name)
        alert = None

        if data.startswith("PU:"):
            ps = job.get("pu",{}).get("status",{})
            if data == "PU:A":
                if toggle_ts(ps, "arr"):
                    alert = f"📍 <b>PU Arrived</b> — {h(ts)}"
            elif data == "PU:L":
                if toggle_ts(ps, "load"):
                    alert = f"📦 <b>Loaded</b> — {h(ts)}"
            elif data == "PU:D":
                if toggle_ts(ps, "dep"):
                    alert = f"🚚 <b>PU Departed</b> — {h(ts)}"
            elif data == "PU:C":
                if toggle_ts(ps, "comp"):
                    alert = f"✅ <b>PU COMPLETE</b> — {h(ts)}"
                    ni = next_incomplete(job, 0)
                    if ni is not None:
                        st2["focus_i"] = ni
            job["pu"]["status"] = ps

        elif data.startswith("DEL:"):
            if stage != "DEL":
                await q.answer("Complete PU first.", show_alert=False)
                return
            dels = job.get("del") or []
            if not dels or i >= len(dels):
                await q.answer("No stops.", show_alert=False)
                return

            ds = dels[i].get("status") or {}
            lbl = f"DEL {i+1}/{len(dels)}"

            if data == "DEL:A":
                if toggle_ts(ds, "arr"):
                    alert = f"📍 <b>Arrived {h(lbl)}</b> — {h(ts)}"
            elif data == "DEL:DL":
                if toggle_ts(ds, "del"):
                    alert = f"📦 <b>Delivered {h(lbl)}</b> — {h(ts)}"
            elif data == "DEL:D":
                if toggle_ts(ds, "dep"):
                    alert = f"🚚 <b>Departed {h(lbl)}</b> — {h(ts)}"
            elif data == "DEL:C":
                if toggle_ts(ds, "comp"):
                    alert = f"✅ <b>COMPLETE {h(lbl)}</b> — {h(ts)}"
                    ni = next_incomplete(job, i + 1)
                    if ni is not None:
                        st2["focus_i"] = ni
            elif data == "DEL:S":
                ds["skip"] = True
                ds["comp"] = ds.get("comp") or now_iso()
                alert = f"⏭️ <b>Skipped {h(lbl)}</b> — {h(ts)}"
                ni = next_incomplete(job, i + 1)
                if ni is not None:
                    st2["focus_i"] = ni

            dels[i]["status"] = ds
            job["del"] = dels

        elif data.startswith("DOC:"):
            if data == "DOC:PTI":
                job["pu"]["docs"]["pti"] = not job["pu"].get("docs",{}).get("pti")
            elif data == "DOC:BOL":
                job["pu"]["docs"]["bol"] = not job["pu"].get("docs",{}).get("bol")
            elif data == "DOC:POD":
                if stage != "DEL":
                    await q.answer("Complete PU first.")
                    return
                dels = job.get("del") or []
                if i < len(dels):
                    dd = dels[i].setdefault("docs", {})
                    dd["pod"] = not dd.get("pod")

        st2["job"] = job
        save_state(st2)

    await q.answer("Updated")
    if alert:
        await send_progress_alert(ctx, update.effective_chat.id, alert)

    try:
        await q.edit_message_reply_markup(reply_markup=build_keyboard(job, st2))
    except TelegramError:
        pass


# ============================================================================
# TEXT HANDLER
# ============================================================================
async def on_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msg = update.effective_message
    if not msg or not msg.text:
        return

    async with _state_lock:
        st = load_state()

    chat = update.effective_chat

    # Detect new loads in allowed groups
    if chat and chat.type in ("group", "supergroup"):
        if chat.id not in set(st.get("allowed_chats") or []):
            return
        
        job = parse_job(msg.text)
        if job:
            async with _state_lock:
                st2 = load_state()
                st2["job"] = job
                st2["focus_i"] = 0
                st2["reminders_sent"] = {}
                st2["geofence_state"] = {}
                save_state(st2)
            
            meta = job.get("meta") or {}
            lines = [f"📦 <b>New Load!</b>", f"<b>{h(load_id_text(job))}</b>"]
            if meta.get("rate"):
                lines.append(f"💰 {money(meta['rate'])}")
            if meta.get("miles"):
                lines.append(f"🛣️ {meta['miles']} mi")
            lines.append("\nType <code>eta</code> or /panel")
            
            await msg.reply_text("\n".join(lines), parse_mode="HTML")
            return

    # Triggers
    if not chat_allowed(update, st):
        return

    parts = msg.text.strip().split()
    if not parts:
        return

    first = re.sub(r"^[^\w]+|[^\w]+$", "", parts[0].lower())
    if first in TRIGGERS:
        rest = " ".join(parts[1:]).lower()
        which = "ALL" if "all" in rest else "AUTO"
        await send_eta(update, ctx, which)


# ============================================================================
# STARTUP
# ============================================================================
async def post_init(app: Application) -> None:
    try:
        await app.bot.delete_webhook(drop_pending_updates=True)
    except TelegramError as e:
        log(f"Webhook delete failed: {e}")
    
    try:
        me = await app.bot.get_me()
        log(f"Connected as @{me.username} (id {me.id})")
    except TelegramError as e:
        log(f"get_me failed: {e}")
    
    # Schedule background jobs
    jq = app.job_queue
    if jq:
        jq.run_repeating(reminder_job, interval=REMINDER_CHECK_INTERVAL_SEC, first=10)
        jq.run_repeating(geofence_job, interval=GEOFENCE_CHECK_INTERVAL_SEC, first=15)
        log(f"Background jobs scheduled")
    
    log("Ready!")


def main() -> None:
    if not TOKEN:
        raise RuntimeError("Missing TELEGRAM_TOKEN")

    app = ApplicationBuilder().token(TOKEN).post_init(post_init).build()

    # Commands
    app.add_handler(CommandHandler("start", start_cmd))
    app.add_handler(CommandHandler("help", start_cmd))
    app.add_handler(CommandHandler("ping", ping_cmd))
    app.add_handler(CommandHandler("status", status_cmd))
    app.add_handler(CommandHandler("claim", claim_cmd))
    app.add_handler(CommandHandler("allowhere", allowhere_cmd))
    app.add_handler(CommandHandler("update", update_cmd))
    app.add_handler(CommandHandler("panel", panel_cmd))
    app.add_handler(CommandHandler("finish", finish_cmd))
    app.add_handler(CommandHandler("catalog", catalog_cmd))
    app.add_handler(CommandHandler("skip", skip_cmd))
    app.add_handler(CommandHandler("deleteall", deleteall_cmd))
    app.add_handler(CommandHandler("leave", leave_cmd))

    # Handlers
    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.LOCATION, on_location))
    
    try:
        app.add_handler(MessageHandler(filters.UpdateType.EDITED_MESSAGE & filters.LOCATION, on_location))
    except Exception:
        pass

    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))

    log(f"Starting v{BOT_VERSION}...")
    log(f"Geofence: {GEOFENCE_MILES}mi | Alerts: {REMINDER_THRESHOLDS_MIN}min | Doc: {REMINDER_DOC_AFTER_MIN}min | Grace: {SCHEDULE_GRACE_MIN}min")
    
    app.run_polling(drop_pending_updates=True, allowed_updates=Update.ALL_TYPES, close_loop=False)


if __name__ == "__main__":
    main()
