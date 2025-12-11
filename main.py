"""
Telegram Trucker Dispatch Assistant Bot
Version: 2025-12-11_v5 (Beautiful Edition)

Features:
- Enhanced visual design with better formatting
- Cleaner UI with improved emoji usage
- Professional status displays
- Better organized information hierarchy
"""

import asyncio
import hashlib
import html
import io
import json
import math
import os
import re
import time
import traceback
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Set

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from timezonefinder import TimezoneFinder
from zoneinfo import ZoneInfo

from telegram import (
    Update,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.error import Forbidden, BadRequest, TelegramError
from telegram.ext import (
    Application,
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

BOT_VERSION = "5.0"

# ============================================================================
# CONFIGURATION
# ============================================================================
def _strip_quotes(s: str) -> str:
    s = (s or "").strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
        return s[1:-1].strip()
    return s

def env_str(name: str, default: str = "") -> str:
    v = os.environ.get(name)
    return _strip_quotes(v) if v else default

def env_int(name: str, default: int) -> int:
    try: return int(env_str(name, ""))
    except: return default

def env_float(name: str, default: float) -> float:
    try: return float(env_str(name, ""))
    except: return default

def env_bool(name: str, default: bool = False) -> bool:
    return env_str(name, "").lower() in ("1", "true", "yes", "on")

def env_list_int(name: str, default: List[int]) -> List[int]:
    v = env_str(name, "")
    if not v: return default
    try: return [int(x.strip()) for x in v.split(",") if x.strip()]
    except: return default

TOKEN = env_str("TELEGRAM_TOKEN")
CLAIM_CODE = env_str("CLAIM_CODE")
STATE_FILE = Path(env_str("STATE_FILE", "state.json"))
TRIGGERS = {t.strip().lower() for t in env_str("TRIGGERS", "eta,1717").split(",") if t.strip()}

NOMINATIM_USER_AGENT = env_str("NOMINATIM_USER_AGENT", "dispatch-bot/1.0")
NOMINATIM_MIN_INTERVAL = env_float("NOMINATIM_MIN_INTERVAL", 1.2)

ETA_ALL_MAX = env_int("ETA_ALL_MAX", 6)
ALERT_TTL_SECONDS = env_int("ALERT_TTL_SECONDS", 25)
DELETEALL_DEFAULT = env_int("DELETEALL_DEFAULT", 300)

GEOFENCE_MILES = env_float("GEOFENCE_MILES", 5.0)
REMINDER_DOC_AFTER_MIN = env_int("REMINDER_DOC_AFTER_MIN", 15)
REMINDER_THRESHOLDS_MIN = env_list_int("REMINDER_THRESHOLDS_MIN", [60, 30, 10])
SCHEDULE_GRACE_MIN = env_int("SCHEDULE_GRACE_MIN", 30)

DEBUG = env_bool("DEBUG", True)

def log(msg: str):
    ts = datetime.now().strftime('%H:%M:%S')
    print(f"[{ts}] {msg}", flush=True)

def log_debug(msg: str):
    if DEBUG:
        log(f"DEBUG: {msg}")

def log_error(msg: str, exc: Exception = None):
    log(f"ERROR: {msg}")
    if exc:
        log(f"  {type(exc).__name__}: {exc}")
        if DEBUG:
            traceback.print_exc()

# ============================================================================
# GLOBALS
# ============================================================================
TF = TimezoneFinder()

GEOCODE_SERVICES = [
    {
        "name": "Nominatim",
        "url": "https://nominatim.openstreetmap.org/search",
        "params": lambda q: {"q": q, "format": "jsonv2", "limit": 1, "countrycodes": "us"},
        "parse": lambda data: (float(data[0]["lat"]), float(data[0]["lon"])) if data else None,
    },
    {
        "name": "Photon",
        "url": "https://photon.komoot.io/api/",
        "params": lambda q: {"q": q, "limit": 1, "lang": "en"},
        "parse": lambda data: (
            data["features"][0]["geometry"]["coordinates"][1],
            data["features"][0]["geometry"]["coordinates"][0]
        ) if data.get("features") else None,
    },
]

OSRM_URL = "https://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}"

_state_lock = asyncio.Lock()
_geo_lock = asyncio.Lock()
_geo_last = 0.0
_tasks: Set[asyncio.Task] = set()

METERS_PER_MILE = 1609.344

# ============================================================================
# VISUAL DESIGN CONSTANTS
# ============================================================================
# Box drawing characters for clean layouts
LINE = "─" * 24
DIVIDER = "┄" * 24

# Status indicators
ICON_SUCCESS = "✅"
ICON_PENDING = "○"
ICON_ACTIVE = "◉"
ICON_WARNING = "⚠️"
ICON_ERROR = "❌"
ICON_TRUCK = "🚛"
ICON_PICKUP = "📦"
ICON_DELIVERY = "🏁"
ICON_CLOCK = "🕐"
ICON_MONEY = "💵"
ICON_MILES = "🛣"
ICON_LOCATION = "📍"
ICON_NAV = "🧭"
ICON_DOC = "📋"
ICON_FUEL = "⛽"
ICON_CALENDAR = "📅"

# ============================================================================
# HELPERS
# ============================================================================
def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def now_iso() -> str:
    return now_utc().isoformat()

def safe_tz(name: str):
    try: return ZoneInfo(name) if name else timezone.utc
    except: return timezone.utc

def h(x: Any) -> str:
    return html.escape(str(x) if x is not None else "", quote=False)

def money(x) -> str:
    try: 
        val = float(x)
        if val >= 1000:
            return f"${val:,.0f}"
        return f"${val:.0f}"
    except: 
        return "—"

def fmt_dur(secs: float) -> str:
    secs = max(0, int(secs))
    hours, mins = divmod(secs // 60, 60)
    if hours > 0:
        return f"{hours}h {mins:02d}m"
    return f"{mins}m"

def fmt_mi(meters: float) -> str:
    mi = meters / METERS_PER_MILE
    if mi < 10:
        return f"{mi:.1f} mi"
    return f"{int(mi)} mi"

def week_key(dt: datetime) -> str:
    iso = dt.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"

def local_stamp(tz_name: str) -> str:
    return now_utc().astimezone(safe_tz(tz_name)).strftime("%I:%M %p")

def local_date(tz_name: str) -> str:
    return now_utc().astimezone(safe_tz(tz_name)).strftime("%b %d")

def progress_bar(current: int, total: int, width: int = 10) -> str:
    """Create a visual progress bar."""
    if total == 0:
        return "░" * width
    filled = int((current / total) * width)
    return "▓" * filled + "░" * (width - filled)

# ============================================================================
# STATE MANAGEMENT
# ============================================================================
def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            st = json.loads(STATE_FILE.read_text())
            if isinstance(st, dict):
                return st
        except Exception as e:
            log_error(f"Failed to load state", e)
    return {}

def save_state(st: dict):
    try:
        STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
        tmp = STATE_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(st, indent=2))
        tmp.replace(STATE_FILE)
    except Exception as e:
        log_error("Failed to save state", e)

def is_owner(update: Update, st: dict) -> bool:
    u = update.effective_user
    return bool(u and st.get("owner_id") and u.id == st["owner_id"])

def chat_allowed(update: Update, st: dict) -> bool:
    chat = update.effective_chat
    if not chat: return False
    if chat.type == "private": return is_owner(update, st)
    return chat.id in (st.get("allowed_chats") or [])

def get_broadcast_chats(st: dict) -> List[int]:
    chats = list(st.get("allowed_chats") or [])
    if st.get("owner_id") and st["owner_id"] not in chats:
        chats.append(st["owner_id"])
    return chats

# ============================================================================
# ADDRESS NORMALIZATION
# ============================================================================
def normalize_address(addr: str) -> str:
    if not addr:
        return ""
    
    result = addr.upper().strip()
    
    num_match = re.search(r'(\d+\s+.+)', result)
    if num_match:
        result = num_match.group(1)
    
    replacements = [
        (r'\bSTREET\b', 'ST'),
        (r'\bAVENUE\b', 'AVE'),
        (r'\bBOULEVARD\b', 'BLVD'),
        (r'\bDRIVE\b', 'DR'),
        (r'\bROAD\b', 'RD'),
        (r'\bLANE\b', 'LN'),
        (r'\bCOURT\b', 'CT'),
        (r'\bPLACE\b', 'PL'),
        (r'\bCIRCLE\b', 'CIR'),
        (r'\bHIGHWAY\b', 'HWY'),
        (r'\bPARKWAY\b', 'PKWY'),
        (r'\bCENTER\b', 'CTR'),
        (r'\bNORTH\b', 'N'),
        (r'\bSOUTH\b', 'S'),
        (r'\bEAST\b', 'E'),
        (r'\bWEST\b', 'W'),
    ]
    
    for pattern, repl in replacements:
        result = re.sub(pattern, repl, result)
    
    result = re.sub(r'\s+', ' ', result)
    result = re.sub(r',\s*,', ',', result)
    result = result.strip(' ,')
    
    return result

def extract_address_components(addr: str) -> dict:
    components = {
        "street": None,
        "city": None,
        "state": None,
        "zip": None,
        "original": addr,
    }
    
    if not addr:
        return components
    
    zip_match = re.search(r'\b(\d{5})(?:-\d{4})?\b', addr)
    if zip_match:
        components["zip"] = zip_match.group(1)
    
    state_match = re.search(r'\b([A-Z]{2})\b(?:\s+\d{5})?(?:\s*,?\s*(?:USA|US)?)?$', addr.upper())
    if state_match:
        components["state"] = state_match.group(1)
    
    street_match = re.search(r'(\d+\s+[^,]+?)(?:,|\s+[A-Z]{2}\b)', addr, re.I)
    if street_match:
        components["street"] = street_match.group(1).strip()
    
    if components["state"]:
        city_match = re.search(r'([A-Za-z\s]+),?\s*' + components["state"], addr, re.I)
        if city_match:
            city = city_match.group(1).strip()
            if city and components["street"] and city not in components["street"]:
                components["city"] = city
            elif city and not components["street"]:
                components["city"] = city
    
    return components

def generate_address_variants(addr: str) -> List[str]:
    if not addr:
        return []
    
    variants = []
    
    normalized = normalize_address(addr)
    components = extract_address_components(normalized)
    
    log_debug(f"Address components: {components}")
    
    if normalized:
        variants.append(normalized)
        variants.append(f"{normalized}, USA")
    
    if all([components["street"], components["city"], components["state"]]):
        full = f"{components['street']}, {components['city']}, {components['state']}"
        variants.append(full)
        variants.append(f"{full}, USA")
        if components["zip"]:
            variants.append(f"{full} {components['zip']}")
    
    if components["street"] and components["state"]:
        variants.append(f"{components['street']}, {components['state']}")
        variants.append(f"{components['street']}, {components['state']}, USA")
    
    if components["city"] and components["state"]:
        variants.append(f"{components['city']}, {components['state']}")
        variants.append(f"{components['city']}, {components['state']}, USA")
    
    if components["zip"]:
        variants.append(components["zip"])
    
    if addr and "USA" not in addr.upper():
        variants.append(f"{addr}, USA")
    
    seen = set()
    unique = []
    for v in variants:
        v_clean = " ".join(v.split()).strip()
        v_key = v_clean.lower()
        if v_key and v_key not in seen and len(v_clean) >= 3:
            seen.add(v_key)
            unique.append(v_clean)
    
    log_debug(f"Generated {len(unique)} address variants")
    
    return unique

# ============================================================================
# GEOCODING
# ============================================================================
def haversine_miles(lat1, lon1, lat2, lon2) -> float:
    R = 3958.8
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat, dlon = lat2 - lat1, lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    return 2 * R * math.asin(math.sqrt(a))

def haversine_meters(lat1, lon1, lat2, lon2) -> float:
    return haversine_miles(lat1, lon1, lat2, lon2) * METERS_PER_MILE

async def geocode_with_service(client: httpx.AsyncClient, service: dict, query: str) -> Optional[Tuple[float, float]]:
    try:
        r = await client.get(service["url"], params=service["params"](query))
        log_debug(f"  {service['name']}: HTTP {r.status_code}")
        
        if r.status_code == 200:
            data = r.json()
            result = service["parse"](data)
            if result:
                log_debug(f"  {service['name']}: Found {result[0]:.4f}, {result[1]:.4f}")
                return result
        elif r.status_code == 429:
            log(f"  {service['name']}: Rate limited")
            
    except httpx.TimeoutException:
        log(f"  {service['name']}: Timeout")
    except Exception as e:
        log_error(f"  {service['name']}: Error", e)
    
    return None

async def geocode(addr: str, cache: dict) -> Optional[Tuple[float, float, str]]:
    if not addr:
        return None
    
    cache_key = addr.lower().strip()
    if cache_key in cache:
        c = cache[cache_key]
        return c["lat"], c["lon"], c.get("tz", "UTC")
    
    variants = generate_address_variants(addr)
    if not variants:
        return None
    
    headers = {
        "User-Agent": NOMINATIM_USER_AGENT or "DispatchBot/1.0",
        "Accept": "application/json",
    }
    
    async with httpx.AsyncClient(timeout=20, headers=headers, follow_redirects=True) as client:
        for variant in variants[:6]:
            log_debug(f"Geocode trying: {variant}")
            
            async with _geo_lock:
                global _geo_last
                wait = _geo_last + NOMINATIM_MIN_INTERVAL - time.monotonic()
                if wait > 0:
                    await asyncio.sleep(wait)
                _geo_last = time.monotonic()
            
            for service in GEOCODE_SERVICES:
                result = await geocode_with_service(client, service, variant)
                
                if result:
                    lat, lon = result
                    tz = TF.timezone_at(lat=lat, lng=lon) or "UTC"
                    cache[cache_key] = {"lat": lat, "lon": lon, "tz": tz}
                    log(f"Geocode SUCCESS: {lat:.4f}, {lon:.4f}")
                    return lat, lon, tz
                
                await asyncio.sleep(0.3)
    
    log(f"Geocode FAILED: '{addr[:40]}'")
    return None

async def get_route(origin: Tuple[float, float], dest: Tuple[float, float]) -> Optional[Tuple[float, float]]:
    url = OSRM_URL.format(lat1=origin[0], lon1=origin[1], lat2=dest[0], lon2=dest[1])
    
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(url, params={"overview": "false"})
            if r.status_code == 200:
                data = r.json()
                if data.get("code") == "Ok" and data.get("routes"):
                    rt = data["routes"][0]
                    return rt["distance"], rt["duration"]
    except Exception as e:
        log_error("Route error", e)
    
    return None

async def calc_eta(st: dict, origin: Tuple[float, float], addr: str) -> dict:
    if not addr:
        return {"ok": False, "err": "No address"}
    
    cache = st.setdefault("geocode_cache", {})
    geo = await geocode(addr, cache)
    
    if not geo:
        return {"ok": False, "err": f"Location not found"}
    
    dest = (geo[0], geo[1])
    route = await get_route(origin, dest)
    
    if route:
        return {
            "ok": True,
            "meters": route[0],
            "seconds": route[1],
            "tz": geo[2],
            "method": "route",
        }
    
    dist_m = haversine_meters(origin[0], origin[1], dest[0], dest[1])
    dist_mi = dist_m / METERS_PER_MILE
    
    speed_mph = 35 if dist_mi < 50 else (50 if dist_mi < 150 else 60)
    est_secs = (dist_mi / speed_mph) * 3600 * 1.2
    
    return {
        "ok": True,
        "meters": dist_m,
        "seconds": est_secs,
        "tz": geo[2],
        "method": "estimate",
    }

# ============================================================================
# LOAD PARSING
# ============================================================================
LOAD_NUM_PATTERN = re.compile(r"Load\s*#\s*:?\s*([A-Za-z0-9\-]+?)(?=BL|Reference|Pickup|Delivery|\s*\n|$)", re.I)
# Rate patterns: "$1,400.00" or "5050 $" or "Rate: $5050" or "Rate : 5050 $"
RATE_PATTERN = re.compile(r"Rate\s*:?\s*\$?\s*([\d,]+(?:\.\d{2})?)\s*\$?", re.I)
# Miles patterns: "Total mi : 561" or "Loaded mi : 2467"
MILES_PATTERN = re.compile(r"(?:Total|Loaded)\s*mi\s*:?\s*([\d,]+)", re.I)
PU_TIME_PATTERN = re.compile(r"PU\s*time\s*:\s*(.+?)(?=\s*PU\s*Address|$)", re.I)
DEL_TIME_PATTERN = re.compile(r"DEL\s*time\s*:\s*(.+?)(?=\s*DEL\s*Address|$)", re.I)
PU_ADDR_PATTERN = re.compile(r"PU\s*Address\s*:\s*(.+?)(?=DEL\s*time|DEL\s*Address|\n-{3,}|$)", re.I | re.S)
DEL_ADDR_PATTERN = re.compile(r"DEL\s*Address\s*:\s*(.+?)(?=-{3,}|Total\s*mi|Loaded\s*mi|Trailer|$)", re.I | re.S)

# Additional reference number patterns - stop at next field marker
BL_PATTERN = re.compile(r"BL\s*#\s*:?\s*(\d+)", re.I)
PO_PATTERN = re.compile(r"PO\s*#\s*:?\s*(\d+)", re.I)
DELIVERY_NUM_PATTERN = re.compile(r"Delivery\s*#\s*:?\s*([A-Za-z0-9\-]+?)(?=-{3,}|DK|ZN|\s*\n|$)", re.I)
REFERENCE_PATTERN = re.compile(r"Reference\s*#'?s?\s*:?\s*(.+?)(?=\n|PO|Delivery|$)", re.I)
PICKUP_NUM_PATTERN = re.compile(r"Pickup\s*#\s*:?\s*([A-Za-z0-9\-]+?)(?=ZN|DK|\s*\n|$)", re.I)

def parse_date_flexible(date_str: str) -> Optional[str]:
    """
    Parse various date formats and return normalized "Mon DD, YYYY" format.
    Handles:
    - "12/11/25" -> "Dec 11, 2025"
    - "12/11/2025" -> "Dec 11, 2025"
    - "Dec 11, 2025" -> "Dec 11, 2025"
    - "December 11, 2025" -> "Dec 11, 2025"
    """
    if not date_str:
        return None
    
    date_str = date_str.strip()
    
    # Already in "Mon DD, YYYY" format
    if re.match(r'\w{3}\s+\d{1,2},?\s+\d{4}', date_str):
        return date_str
    
    # MM/DD/YY or MM/DD/YYYY format
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', date_str)
    if m:
        month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if year < 100:
            year += 2000
        
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        if 1 <= month <= 12:
            return f"{months[month-1]} {day}, {year}"
    
    return date_str

def parse_time_window(time_str: str) -> dict:
    """
    Parse time strings like:
    - "Dec 10, 2025 08:00 -14:00 FCFS"
    - "Dec 11, 2025 00:01-23:59 FCFS"
    - "12/11/25 08:00"
    - "12/11/25 READY NOW"
    - "Dec 10, 2025 08:00"
    
    Returns: {
        "raw": original string,
        "date": "Dec 10, 2025",
        "early": "08:00",
        "late": "14:00" or None,
        "note": "FCFS" or "READY NOW" or None,
        "early_dt": datetime or None,
        "late_dt": datetime or None
    }
    """
    if not time_str:
        return {"raw": "", "date": None, "early": None, "late": None, "note": None}
    
    result = {
        "raw": time_str.strip(),
        "date": None,
        "early": None,
        "late": None,
        "note": None,
        "early_dt": None,
        "late_dt": None,
    }
    
    # Extract notes like "FCFS", "APPT", "READY NOW" etc
    note_match = re.search(r'\b(FCFS|APPT|BY APPT|APPOINTMENT|READY NOW|ASAP)\s*$', time_str, re.I)
    if note_match:
        result["note"] = note_match.group(1).upper()
        time_str = time_str[:note_match.start()].strip()
    
    # Try multiple date formats
    
    # Format 1: "Dec 10, 2025 08:00 -14:00" or "Dec 10, 2025 08:00-14:00"
    window_match = re.search(
        r'(\w{3}\s+\d{1,2},?\s+\d{4})\s+(\d{1,2}:\d{2})\s*[-–to]+\s*(\d{1,2}:\d{2})',
        time_str, re.I
    )
    
    if window_match:
        date_str = window_match.group(1).strip()
        date_str = re.sub(r',\s*', ' ', date_str)
        date_str = re.sub(r'(\w{3})\s+(\d{1,2})\s+(\d{4})', r'\1 \2, \3', date_str)
        result["date"] = date_str
        result["early"] = window_match.group(2)
        result["late"] = window_match.group(3)
        return result
    
    # Format 2: "12/11/25 08:00-14:00" (MM/DD/YY with window)
    window_match2 = re.search(
        r'(\d{1,2}/\d{1,2}/\d{2,4})\s+(\d{1,2}:\d{2})\s*[-–to]+\s*(\d{1,2}:\d{2})',
        time_str, re.I
    )
    
    if window_match2:
        result["date"] = parse_date_flexible(window_match2.group(1))
        result["early"] = window_match2.group(2)
        result["late"] = window_match2.group(3)
        return result
    
    # Format 3: "Dec 10, 2025 08:00" (single time, Mon DD YYYY)
    single_match = re.search(
        r'(\w{3}\s+\d{1,2},?\s+\d{4})\s+(\d{1,2}:\d{2})',
        time_str, re.I
    )
    if single_match:
        date_str = single_match.group(1).strip()
        date_str = re.sub(r',\s*', ' ', date_str)
        date_str = re.sub(r'(\w{3})\s+(\d{1,2})\s+(\d{4})', r'\1 \2, \3', date_str)
        result["date"] = date_str
        result["early"] = single_match.group(2)
        return result
    
    # Format 4: "12/11/25 08:00" (single time, MM/DD/YY)
    single_match2 = re.search(
        r'(\d{1,2}/\d{1,2}/\d{2,4})\s+(\d{1,2}:\d{2})',
        time_str, re.I
    )
    if single_match2:
        result["date"] = parse_date_flexible(single_match2.group(1))
        result["early"] = single_match2.group(2)
        return result
    
    # Format 5: "12/11/25" or "12/11/25 READY NOW" (date only, no time)
    date_only = re.search(r'(\d{1,2}/\d{1,2}/\d{2,4})', time_str, re.I)
    if date_only:
        result["date"] = parse_date_flexible(date_only.group(1))
        return result
    
    # Format 6: "Dec 10, 2025" (date only)
    date_only2 = re.search(r'(\w{3}\s+\d{1,2},?\s+\d{4})', time_str, re.I)
    if date_only2:
        date_str = date_only2.group(1).strip()
        date_str = re.sub(r',\s*', ' ', date_str)
        date_str = re.sub(r'(\w{3})\s+(\d{1,2})\s+(\d{4})', r'\1 \2, \3', date_str)
        result["date"] = date_str
        return result
    
    return result

def format_time_window(tw: dict, tz_name: str = "UTC") -> str:
    """Format a time window for display."""
    if not tw or not tw.get("date"):
        return tw.get("raw", "") if tw else ""
    
    # Convert 24h to 12h format
    def to_12h(t: str) -> str:
        if not t:
            return ""
        try:
            h, m = map(int, t.split(":"))
            suffix = "AM" if h < 12 else "PM"
            h = h % 12 or 12
            return f"{h}:{m:02d} {suffix}"
        except:
            return t
    
    date_part = tw["date"]
    
    # Try to shorten date (remove year if current year)
    try:
        if "2025" in date_part:
            date_part = date_part.replace(", 2025", "").replace(" 2025", "")
    except:
        pass
    
    if tw.get("late"):
        time_part = f"{to_12h(tw['early'])} - {to_12h(tw['late'])}"
    else:
        time_part = to_12h(tw.get("early", ""))
    
    result = f"{date_part} • {time_part}"
    
    if tw.get("note"):
        result += f" ({tw['note']})"
    
    return result

def get_appointment_deadline(tw: dict, tz_name: str) -> Optional[datetime]:
    """Get the latest acceptable arrival time (for ETA warnings)."""
    if not tw or not tw.get("date"):
        return None
    
    # Use late time if available, otherwise early time
    target_time = tw.get("late") or tw.get("early")
    if not target_time:
        return None
    
    try:
        # Parse date
        date_str = tw["date"]
        # Normalize format
        date_str = re.sub(r',?\s+', ' ', date_str).strip()
        
        dt_str = f"{date_str} {target_time}"
        
        # Try multiple formats
        for fmt in ["%b %d %Y %H:%M", "%B %d %Y %H:%M"]:
            try:
                dt = datetime.strptime(dt_str, fmt)
                return dt.replace(tzinfo=safe_tz(tz_name))
            except ValueError:
                continue
        
        return None
    except:
        return None

def check_eta_vs_window(eta_seconds: float, tw: dict, tz_name: str) -> dict:
    """
    Check if ETA meets the appointment window.
    Returns: {"status": "early"|"on_time"|"late"|"unknown", "message": str}
    """
    if not tw or not tw.get("date"):
        return {"status": "unknown", "message": ""}
    
    deadline = get_appointment_deadline(tw, tz_name)
    if not deadline:
        return {"status": "unknown", "message": ""}
    
    arrival = now_utc() + timedelta(seconds=eta_seconds)
    
    # Get early time for window start
    early_dt = None
    if tw.get("early") and tw.get("date"):
        try:
            date_str = re.sub(r',?\s+', ' ', tw["date"]).strip()
            dt_str = f"{date_str} {tw['early']}"
            for fmt in ["%b %d %Y %H:%M", "%B %d %Y %H:%M"]:
                try:
                    early_dt = datetime.strptime(dt_str, fmt).replace(tzinfo=safe_tz(tz_name))
                    break
                except ValueError:
                    continue
        except:
            pass
    
    # Calculate time differences
    mins_until_deadline = (deadline - arrival).total_seconds() / 60
    
    if mins_until_deadline < -15:
        return {
            "status": "late",
            "message": f"⚠️ {abs(int(mins_until_deadline))}m past deadline"
        }
    elif mins_until_deadline < 30:
        return {
            "status": "tight",
            "message": f"⏱ Cutting it close ({int(mins_until_deadline)}m buffer)"
        }
    elif early_dt and arrival < early_dt:
        mins_early = (early_dt - arrival).total_seconds() / 60
        return {
            "status": "early",
            "message": f"☕ {int(mins_early)}m early - window opens {tw['early']}"
        }
    else:
        return {
            "status": "on_time",
            "message": "✅ On time"
        }

def clean_address(addr: str) -> str:
    """Clean and normalize address from load sheet."""
    if not addr:
        return ""
    
    # Check if it's a comma-separated single line (like "Company,123 St,City,State ZIP,USA")
    # vs multi-line format
    addr = addr.strip()
    
    # If no newlines but has commas, it's already comma-separated
    if "\n" not in addr and "," in addr:
        # Just clean it up
        result = addr
    else:
        # Multi-line format - join with commas
        lines = []
        for ln in addr.strip().split("\n"):
            ln = ln.strip()
            if not ln or len(ln) < 2:
                continue
            
            ln_lower = ln.lower()
            if any(skip in ln_lower for skip in ["---", "===", "total mi", "loaded mi", "rate :", "rate:", "trailer", "failure"]):
                break
            
            lines.append(ln)
        
        if not lines:
            return ""
        
        result = ", ".join(lines)
    
    # Clean up
    result = re.sub(r'\s+', ' ', result)
    result = re.sub(r',\s*,', ',', result)
    result = re.sub(r'\s*,\s*', ', ', result)  # Normalize comma spacing
    
    # Remove trailing ,USA if present (geocoder doesn't need it)
    result = re.sub(r',\s*USA\s*$', '', result, flags=re.I)
    
    return result.strip(' ,')

def parse_load(text: str) -> Optional[dict]:
    if "pu address" not in text.lower() or "del address" not in text.lower():
        return None
    
    log("Parsing load...")
    
    # Primary load number
    load_num = None
    m = LOAD_NUM_PATTERN.search(text)
    if m:
        load_num = m.group(1).strip()
    
    # Additional reference numbers
    bl_num = None
    m = BL_PATTERN.search(text)
    if m:
        bl_num = m.group(1).strip()
    
    po_num = None
    m = PO_PATTERN.search(text)
    if m:
        po_num = m.group(1).strip()
    
    delivery_num = None
    m = DELIVERY_NUM_PATTERN.search(text)
    if m:
        delivery_num = m.group(1).strip()
    
    pickup_num = None
    m = PICKUP_NUM_PATTERN.search(text)
    if m:
        pickup_num = m.group(1).strip()
    
    rate = None
    m = RATE_PATTERN.search(text)
    if m:
        try:
            rate = float(m.group(1).replace(",", ""))
        except: pass
    
    miles = None
    m = MILES_PATTERN.search(text)
    if m:
        try:
            miles = int(m.group(1).replace(",", ""))
        except: pass
    
    pu_time_raw = None
    m = PU_TIME_PATTERN.search(text)
    if m:
        pu_time_raw = m.group(1).strip()
    pu_time = parse_time_window(pu_time_raw)
    
    del_time_raw = None
    m = DEL_TIME_PATTERN.search(text)
    if m:
        del_time_raw = m.group(1).strip()
    del_time = parse_time_window(del_time_raw)
    
    pu_addr = None
    m = PU_ADDR_PATTERN.search(text)
    if m:
        pu_addr = clean_address(m.group(1))
    
    del_addr = None
    m = DEL_ADDR_PATTERN.search(text)
    if m:
        del_addr = clean_address(m.group(1))
    
    if not pu_addr or not del_addr:
        return None
    
    job_id = hashlib.sha1(f"{load_num}|{pu_addr}|{del_addr}".encode()).hexdigest()[:10]
    
    # Build references dict with all found numbers
    refs = {}
    if bl_num:
        refs["bl"] = bl_num
    if po_num:
        refs["po"] = po_num
    if delivery_num:
        refs["delivery"] = delivery_num
    if pickup_num:
        refs["pickup"] = pickup_num
    
    job = {
        "id": job_id,
        "created_at": now_iso(),
        "meta": {
            "load_number": load_num,
            "rate": rate,
            "miles": miles,
            "refs": refs if refs else None,
        },
        "pu": {
            "addr": pu_addr,
            "time": pu_time,
            "num": pickup_num,  # Pickup # if different from load
            "status": {"arr": None, "load": None, "dep": None, "comp": None},
            "docs": {"pti": False, "bol": False},
        },
        "del": [{
            "addr": del_addr,
            "time": del_time,
            "num": delivery_num,  # Delivery #
            "status": {"arr": None, "del": None, "dep": None, "comp": None, "skip": False},
            "docs": {"pod": False},
        }],
    }
    
    # Log what we found
    refs_str = ", ".join(f"{k}={v}" for k, v in refs.items()) if refs else "none"
    log(f"Parsed: {load_num} | ${rate} | {miles}mi | refs: {refs_str}")
    return job

def get_job(st: dict) -> Optional[dict]:
    job = st.get("job")
    if not job or not isinstance(job, dict):
        return None
    if "pu" not in job or "del" not in job:
        return None
    return job

# ============================================================================
# WORKFLOW HELPERS
# ============================================================================
def is_pu_complete(job: dict) -> bool:
    return bool(job.get("pu", {}).get("status", {}).get("comp"))

def get_focus(job: dict, st: dict) -> Tuple[str, int]:
    if not is_pu_complete(job):
        return "PU", 0
    
    dels = job.get("del") or []
    idx = st.get("focus_i", 0)
    idx = max(0, min(idx, len(dels) - 1)) if dels else 0
    
    if dels and idx < len(dels):
        if dels[idx].get("status", {}).get("comp"):
            for i in range(idx + 1, len(dels)):
                if not dels[i].get("status", {}).get("comp"):
                    idx = i
                    break
    
    return "DEL", idx

def load_label(job: dict) -> str:
    meta = job.get("meta") or {}
    if meta.get("load_number"):
        return f"#{meta['load_number']}"
    return f"#{job.get('id', '?')[:8]}"

def short_addr(addr: str, max_len: int = 50) -> str:
    if not addr:
        return "—"
    if len(addr) <= max_len:
        return addr
    return addr[:max_len-3] + "..."

def toggle_timestamp(obj: dict, key: str) -> bool:
    if obj.get(key):
        obj[key] = None
        return False
    obj[key] = now_iso()
    return True

def get_completion_stats(job: dict) -> Tuple[int, int]:
    """Get (completed_steps, total_steps) for progress display."""
    total = 0
    completed = 0
    
    # PU steps: arrived, loaded, departed, complete
    pu_status = job.get("pu", {}).get("status", {})
    for key in ["arr", "load", "dep", "comp"]:
        total += 1
        if pu_status.get(key):
            completed += 1
    
    # DEL steps for each delivery
    for d in job.get("del", []):
        ds = d.get("status", {})
        for key in ["arr", "del", "dep", "comp"]:
            total += 1
            if ds.get(key):
                completed += 1
    
    return completed, total

# ============================================================================
# BEAUTIFUL KEYBOARDS
# ============================================================================
def btn(text: str, data: str) -> InlineKeyboardButton:
    return InlineKeyboardButton(text, callback_data=data)

def status_icon(done: bool) -> str:
    return "✓" if done else "○"

def build_panel_keyboard(job: dict, st: dict) -> InlineKeyboardMarkup:
    stage, idx = get_focus(job, st)
    rows = []
    
    if stage == "PU":
        pu = job.get("pu", {})
        ps = pu.get("status", {})
        pd = pu.get("docs", {})
        
        # Status row with cleaner icons
        rows.append([
            btn(f"{status_icon(ps.get('arr'))} Arrive", "PU:ARR"),
            btn(f"{status_icon(ps.get('load'))} Load", "PU:LOAD"),
            btn(f"{status_icon(ps.get('dep'))} Depart", "PU:DEP"),
        ])
        # Docs and complete
        rows.append([
            btn(f"{'📋' if pd.get('pti') else '○'} PTI", "DOC:PTI"),
            btn(f"{'📋' if pd.get('bol') else '○'} BOL", "DOC:BOL"),
            btn(f"{'✅' if ps.get('comp') else '⬜'} Complete", "PU:COMP"),
        ])
    else:
        dels = job.get("del", [])
        if idx < len(dels):
            d = dels[idx]
            ds = d.get("status", {})
            dd = d.get("docs", {})
            
            # Status row
            rows.append([
                btn(f"{status_icon(ds.get('arr'))} Arrive", "DEL:ARR"),
                btn(f"{status_icon(ds.get('del'))} Deliver", "DEL:DEL"),
                btn(f"{status_icon(ds.get('dep'))} Depart", "DEL:DEP"),
            ])
            # Docs, complete, skip
            rows.append([
                btn(f"{'📋' if dd.get('pod') else '○'} POD", "DOC:POD"),
                btn(f"{'✅' if ds.get('comp') else '⬜'} Complete", "DEL:COMP"),
                btn("⏭ Skip", "DEL:SKIP"),
            ])
            
            # Navigation for multi-stop
            if len(dels) > 1:
                nav = []
                if idx > 0:
                    nav.append(btn("◀ Prev", f"NAV:{idx-1}"))
                nav.append(btn(f"Stop {idx+1}/{len(dels)}", "NOOP"))
                if idx < len(dels) - 1:
                    nav.append(btn("Next ▶", f"NAV:{idx+1}"))
                rows.append(nav)
    
    # Bottom action row
    rows.append([
        btn("🧭 ETA", "ETA:ONE"),
        btn("📍 All ETAs", "ETA:ALL"),
    ])
    rows.append([
        btn("📊 Catalog", "CATALOG"),
        btn("🏁 Finish", "FINISH"),
    ])
    
    return InlineKeyboardMarkup(rows)

def build_done_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [btn("📊 View Catalog", "CATALOG")],
        [btn("🔄 New Load Ready", "NOOP")]
    ])

# ============================================================================
# ALERTS
# ============================================================================
async def send_alert(ctx: ContextTypes.DEFAULT_TYPE, chat_id: int, text: str):
    try:
        msg = await ctx.bot.send_message(chat_id, text, parse_mode="HTML", disable_notification=True)
        if ALERT_TTL_SECONDS > 0:
            async def delete_later():
                await asyncio.sleep(ALERT_TTL_SECONDS)
                try:
                    await ctx.bot.delete_message(chat_id, msg.message_id)
                except:
                    pass
            task = asyncio.create_task(delete_later())
            _tasks.add(task)
            task.add_done_callback(_tasks.discard)
    except Exception as e:
        log_error("Alert failed", e)

# ============================================================================
# EXCEL EXPORT - Enhanced
# ============================================================================
def make_catalog_xlsx(records: List[dict], week: str) -> Tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = week[:31] if week != "ALL" else "All Loads"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_align = Alignment(horizontal="center", vertical="center")
    
    headers = ["Date", "Load #", "Origin", "Destination", "Rate", "Miles", "$/Mi"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    # Data rows
    total_rate, total_miles = 0, 0
    money_format = '"$"#,##0'
    
    for r in sorted(records, key=lambda x: x.get("completed_utc", "")):
        rate = r.get("rate")
        miles = r.get("posted_miles") or r.get("est_miles")
        rpm = (rate / miles) if rate and miles else None
        
        row = [
            r.get("completed", "")[:10],  # Just date
            r.get("load_number", ""),
            r.get("pickup", "")[:40],
            r.get("deliveries", "")[:40],
            rate,
            miles,
            round(rpm, 2) if rpm else None,
        ]
        ws.append(row)
        
        if rate: total_rate += rate
        if miles: total_miles += miles
    
    # Totals row
    ws.append([])
    total_row = ["TOTAL", f"{len(records)} loads", "", "", total_rate, total_miles,
                 round(total_rate/total_miles, 2) if total_miles else None]
    ws.append(total_row)
    
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="ECF0F1")
    
    # Column widths
    widths = [12, 14, 35, 35, 10, 8, 8]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w
    
    buf = io.BytesIO()
    wb.save(buf)
    filename = f"loads_{week}.xlsx" if week != "ALL" else "loads_all.xlsx"
    return buf.getvalue(), filename

# ============================================================================
# BEAUTIFUL MESSAGE BUILDERS
# ============================================================================
def build_start_message() -> str:
    return f"""
<b>🚛 DISPATCH ASSISTANT</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

Welcome! I help track your loads, calculate ETAs, and manage documentation.

<b>⚡ QUICK START</b>
<code>┌────────────────────────┐</code>
<code>│</code> 1️⃣  DM me /claim code    <code>│</code>
<code>│</code> 2️⃣  DM me /update        <code>│</code>
<code>│</code> 3️⃣  Add me to group      <code>│</code>
<code>│</code> 4️⃣  Type /allowhere      <code>│</code>
<code>└────────────────────────┘</code>

<b>📱 COMMANDS</b>
• <code>eta</code> or <code>1717</code> — Get ETA
• /panel — Control panel
• /catalog — Weekly report
• /finish — Complete load

<code>━━━━━━━━━━━━━━━━━━━━━━</code>
<i>v{BOT_VERSION} • Powered by OSRM</i>
"""

def build_status_message(st: dict, update: Update) -> str:
    job = get_job(st)
    loc = st.get("last_location")
    gc = st.get("geocode_cache", {})
    history = st.get("history", [])
    
    # Location status
    loc_status = "❌ Not set"
    if loc:
        age_min = 0
        if loc.get("updated_at"):
            try:
                age_min = int((now_utc() - datetime.fromisoformat(loc["updated_at"])).total_seconds() / 60)
            except: pass
        if age_min < 5:
            loc_status = f"🟢 Live ({age_min}m)"
        elif age_min < 30:
            loc_status = f"🟡 Recent ({age_min}m)"
        else:
            loc_status = f"🔴 Stale ({age_min}m)"
    
    # Job status
    if job:
        completed, total = get_completion_stats(job)
        progress = progress_bar(completed, total, 8)
        job_status = f"🟢 {load_label(job)} [{progress}]"
    else:
        job_status = "⚪ No active load"
    
    # Weekly stats
    tz_name = loc.get("tz", "UTC") if loc else "UTC"
    week = week_key(now_utc().astimezone(safe_tz(tz_name)))
    week_loads = [r for r in history if r.get("week") == week]
    week_earnings = sum(r.get("rate", 0) or 0 for r in week_loads)
    
    return f"""
<b>📊 SYSTEM STATUS</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

<b>🔗 Connection</b>
   Location: {loc_status}
   Cache: {len(gc)} addresses

<b>📦 Current Load</b>
   {job_status}

<b>📈 This Week</b>
   Loads: {len(week_loads)}
   Earnings: {money(week_earnings)}

<code>━━━━━━━━━━━━━━━━━━━━━━</code>
<i>v{BOT_VERSION}</i>
"""

def build_new_load_message(job: dict) -> str:
    meta = job.get("meta", {})
    pu = job.get("pu", {})
    dels = job.get("del", [])
    refs = meta.get("refs", {}) or {}
    
    # Calculate rate per mile if available
    rpm = ""
    if meta.get("rate") and meta.get("miles"):
        rpm_val = meta["rate"] / meta["miles"]
        rpm = f"   <i>${rpm_val:.2f}/mi</i>\n"
    
    # Format time windows
    pu_time = pu.get("time", {})
    pu_time_str = ""
    if isinstance(pu_time, dict) and pu_time.get("date"):
        pu_time_str = f"\n🕐 {format_time_window(pu_time)}"
    elif isinstance(pu_time, str) and pu_time:
        pu_time_str = f"\n🕐 {pu_time}"
    
    del_time_str = ""
    if dels:
        del_time = dels[0].get("time", {})
        if isinstance(del_time, dict) and del_time.get("date"):
            del_time_str = f"\n🕐 {format_time_window(del_time)}"
        elif isinstance(del_time, str) and del_time:
            del_time_str = f"\n🕐 {del_time}"
    
    # Build reference numbers section
    refs_lines = []
    if refs.get("bl"):
        refs_lines.append(f"BL# {refs['bl']}")
    if refs.get("po"):
        refs_lines.append(f"PO# {refs['po']}")
    if refs.get("pickup"):
        refs_lines.append(f"PU# {refs['pickup']}")
    if refs.get("delivery"):
        refs_lines.append(f"DEL# {refs['delivery']}")
    
    refs_str = ""
    if refs_lines:
        refs_str = "\n<i>" + " • ".join(refs_lines) + "</i>\n"
    
    return f"""
<b>📦 NEW LOAD DETECTED</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

<b>{load_label(job)}</b>
{money(meta.get('rate'))}  •  {meta.get('miles') or '—'} mi
{rpm}{refs_str}
<b>📍 Pickup</b>
{h(short_addr(pu.get('addr', ''), 50))}{pu_time_str}

<b>🏁 Delivery</b>
{h(short_addr(dels[0].get('addr', '') if dels else '', 50))}{del_time_str}

<code>━━━━━━━━━━━━━━━━━━━━━━</code>
Type <code>eta</code> or tap /panel
"""

def build_panel_message(job: dict, st: dict) -> str:
    meta = job.get("meta", {})
    stage, idx = get_focus(job, st)
    completed, total = get_completion_stats(job)
    progress = progress_bar(completed, total, 12)
    
    # Header
    lines = [
        f"<b>🚛 {load_label(job)}</b>",
        f"<code>{progress}</code> {completed}/{total}",
        "",
    ]
    
    # Rate/miles info
    if meta.get("rate") or meta.get("miles"):
        info_parts = []
        if meta.get("rate"):
            info_parts.append(f"💵 {money(meta['rate'])}")
        if meta.get("miles"):
            info_parts.append(f"🛣 {meta['miles']} mi")
        if meta.get("rate") and meta.get("miles"):
            rpm = meta["rate"] / meta["miles"]
            info_parts.append(f"${rpm:.2f}/mi")
        lines.append(" • ".join(info_parts))
        lines.append("")
    
    # Current focus
    if stage == "PU":
        pu = job.get("pu", {})
        ps = pu.get("status", {})
        
        status_text = "⏳ In Progress"
        if ps.get("comp"):
            status_text = "✅ Complete"
        elif ps.get("dep"):
            status_text = "🚛 Departed"
        elif ps.get("load"):
            status_text = "📦 Loaded"
        elif ps.get("arr"):
            status_text = "📍 On Site"
        
        lines.append(f"<b>📦 PICKUP</b>  {status_text}")
        lines.append(f"<code>{h(short_addr(pu.get('addr', ''), 45))}</code>")
        
        # Time window display
        pu_time = pu.get("time", {})
        if isinstance(pu_time, dict) and pu_time.get("date"):
            lines.append(f"🕐 {format_time_window(pu_time)}")
        elif isinstance(pu_time, str) and pu_time:
            lines.append(f"🕐 {h(pu_time)}")
    else:
        dels = job.get("del", [])
        if idx < len(dels):
            d = dels[idx]
            ds = d.get("status", {})
            
            status_text = "⏳ In Progress"
            if ds.get("comp"):
                status_text = "✅ Complete"
            elif ds.get("skip"):
                status_text = "⏭ Skipped"
            elif ds.get("dep"):
                status_text = "🚛 Departed"
            elif ds.get("del"):
                status_text = "📦 Delivered"
            elif ds.get("arr"):
                status_text = "📍 On Site"
            
            stop_label = f"STOP {idx+1}/{len(dels)}" if len(dels) > 1 else "DELIVERY"
            lines.append(f"<b>🏁 {stop_label}</b>  {status_text}")
            lines.append(f"<code>{h(short_addr(d.get('addr', ''), 45))}</code>")
            
            # Time window display
            del_time = d.get("time", {})
            if isinstance(del_time, dict) and del_time.get("date"):
                lines.append(f"🕐 {format_time_window(del_time)}")
            elif isinstance(del_time, str) and del_time:
                lines.append(f"🕐 {h(del_time)}")
    
    return "\n".join(lines)

def build_eta_message(eta: dict, label: str, addr: str, appt: any, job: dict, tz_name: str = "UTC", age_warn: str = "") -> str:
    if not eta["ok"]:
        return f"""
<b>⚠️ ETA UNAVAILABLE</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

{eta['err']}

<b>📍 {label}</b>
<code>{h(short_addr(addr, 45))}</code>

<i>Try /clearcache and retry</i>
"""
    
    arr_time = now_utc() + timedelta(seconds=eta["seconds"])
    tz = safe_tz(eta.get("tz", tz_name))
    arr_str = arr_time.astimezone(tz).strftime("%I:%M %p")
    
    method_note = " <i>(estimated)</i>" if eta["method"] == "estimate" else ""
    
    lines = [
        f"<b>🧭 ETA: {fmt_dur(eta['seconds'])}</b>{method_note}",
        f"<code>━━━━━━━━━━━━━━━━━━━━━━</code>",
        "",
        f"<b>{label}</b>  •  {load_label(job)}",
        f"📍 <code>{h(short_addr(addr, 40))}</code>",
        "",
        f"🛣 {fmt_mi(eta['meters'])}",
        f"🕐 Arrive ~{arr_str}",
    ]
    
    # Handle time window
    if isinstance(appt, dict) and appt.get("date"):
        lines.append(f"📅 Window: {format_time_window(appt)}")
        
        # Check if on time
        window_check = check_eta_vs_window(eta["seconds"], appt, tz_name)
        if window_check.get("message"):
            lines.append(window_check["message"])
    elif isinstance(appt, str) and appt:
        lines.append(f"📅 Appt: {h(appt)}")
    
    if age_warn:
        lines.append("")
        lines.append(age_warn)
    
    return "\n".join(lines)

def build_all_etas_message(etas: List[dict], job: dict, age_warn: str = "") -> str:
    lines = [
        f"<b>📍 ALL ETAs</b>  •  {load_label(job)}",
        f"<code>━━━━━━━━━━━━━━━━━━━━━━</code>",
        "",
    ]
    
    for e in etas:
        if e.get("complete"):
            lines.append(f"✅ <s>{e['label']}</s>")
        elif e.get("ok"):
            method = "≈" if e.get("method") == "estimate" else ""
            lines.append(f"<b>{e['label']}</b>: {fmt_dur(e['seconds'])}{method} • {fmt_mi(e['meters'])}")
        else:
            lines.append(f"⚠️ <b>{e['label']}</b>: {e.get('err', 'Error')}")
    
    if age_warn:
        lines.append("")
        lines.append(age_warn)
    
    return "\n".join(lines)

def build_finish_message(job: dict, st: dict) -> str:
    meta = job.get("meta", {})
    tz_name = (st.get("last_location") or {}).get("tz", "UTC")
    wk = week_key(now_utc().astimezone(safe_tz(tz_name)))
    
    history = st.get("history", [])
    week_records = [r for r in history if r.get("week") == wk]
    wk_count = len(week_records)
    wk_rate = sum(r.get("rate", 0) or 0 for r in week_records)
    wk_miles = sum(r.get("posted_miles", 0) or 0 for r in week_records)
    wk_rpm = wk_rate / wk_miles if wk_miles else 0
    
    return f"""
<b>✅ LOAD COMPLETE</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

<b>{load_label(job)}</b>
{money(meta.get('rate'))}  •  {meta.get('miles') or '—'} mi

<b>📊 WEEK {wk}</b>
<code>┌────────────────────────┐</code>
<code>│</code>  Loads:     <b>{wk_count}</b>
<code>│</code>  Gross:     <b>{money(wk_rate)}</b>
<code>│</code>  Miles:     <b>{wk_miles}</b>
<code>│</code>  Avg $/mi:  <b>${wk_rpm:.2f}</b>
<code>└────────────────────────┘</code>

<i>Ready for next load!</i>
"""

def build_alert_message(icon: str, title: str, detail: str = "", timestamp: str = "") -> str:
    lines = [f"{icon} <b>{title}</b>"]
    if detail:
        lines.append(detail)
    if timestamp:
        lines.append(f"<i>{timestamp}</i>")
    return "\n".join(lines)

# ============================================================================
# COMMAND HANDLERS
# ============================================================================
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(build_start_message(), parse_mode="HTML")

async def cmd_ping(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"🏓 <b>Pong!</b>\n<i>v{BOT_VERSION} • Online</i>", parse_mode="HTML")

async def cmd_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    await update.message.reply_text(
        build_status_message(st, update),
        parse_mode="HTML"
    )

async def cmd_testgeo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        await update.message.reply_text("⚠️ Owner only")
        return
    
    if not ctx.args:
        await update.message.reply_text("Usage: /testgeo <address>")
        return
    
    addr = " ".join(ctx.args)
    await update.message.reply_text(f"🔍 Testing: <code>{h(addr)}</code>", parse_mode="HTML")
    
    cache = {}
    geo = await geocode(addr, cache)
    
    if geo:
        lat, lon, tz = geo
        await update.message.reply_text(
            f"""
<b>✅ GEOCODE SUCCESS</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

📍 {lat:.6f}, {lon:.6f}
🕐 Timezone: {tz}

<a href="https://www.google.com/maps?q={lat},{lon}">Open in Google Maps</a>
""",
            parse_mode="HTML"
        )
        try:
            await ctx.bot.send_location(update.effective_chat.id, lat, lon)
        except: pass
    else:
        variants = generate_address_variants(addr)
        variant_list = "\n".join(f"  • {v}" for v in variants[:5])
        await update.message.reply_text(
            f"""
<b>❌ GEOCODE FAILED</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

Tried variants:
{variant_list}

All services returned no results.
""",
            parse_mode="HTML"
        )

async def cmd_claim(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        await update.message.reply_text("⚠️ Please DM me: /claim <code>")
        return
    
    if not CLAIM_CODE:
        await update.message.reply_text("⚠️ CLAIM_CODE not configured")
        return
    
    code = " ".join(ctx.args) if ctx.args else ""
    if code != CLAIM_CODE:
        await update.message.reply_text("❌ Invalid code")
        return
    
    async with _state_lock:
        st = load_state()
        st["owner_id"] = update.effective_user.id
        save_state(st)
    
    log(f"Owner claimed: {update.effective_user.id}")
    await update.message.reply_text(
        """
<b>✅ OWNERSHIP CONFIRMED</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

You now control this bot!

<b>Next step:</b>
Send /update to share your location
""",
        parse_mode="HTML"
    )

async def cmd_allowhere(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        await update.message.reply_text("⚠️ Owner only")
        return
    
    if update.effective_chat.type == "private":
        await update.message.reply_text("⚠️ Run this in a group")
        return
    
    async with _state_lock:
        st = load_state()
        allowed = set(st.get("allowed_chats", []))
        allowed.add(update.effective_chat.id)
        st["allowed_chats"] = list(allowed)
        save_state(st)
    
    log(f"Group allowed: {update.effective_chat.id}")
    await update.message.reply_text(
        """
<b>✅ GROUP AUTHORIZED</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

This group can now receive load updates.

Forward a load sheet to get started!
""",
        parse_mode="HTML"
    )

async def cmd_update(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        await update.message.reply_text("⚠️ Owner only")
        return
    
    if update.effective_chat.type != "private":
        await update.message.reply_text("📍 Please DM me /update")
        return
    
    kb = [[KeyboardButton("📍 Share My Location", request_location=True)]]
    await update.message.reply_text(
        """
<b>📍 LOCATION UPDATE</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

Tap the button below to share.

<b>💡 Pro tip:</b> Use <i>Live Location</i> for automatic geofence alerts!
""",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True, one_time_keyboard=True)
    )

async def cmd_panel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not chat_allowed(update, st):
        await update.message.reply_text("⚠️ Not authorized")
        return
    
    job = get_job(st)
    if not job:
        await update.message.reply_text(
            """
<b>📭 NO ACTIVE LOAD</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

Forward a load sheet to get started.
""",
            parse_mode="HTML"
        )
        return
    
    msg = await update.message.reply_text(
        build_panel_message(job, st),
        parse_mode="HTML",
        reply_markup=build_panel_keyboard(job, st)
    )
    
    async with _state_lock:
        st = load_state()
        st.setdefault("panel_msgs", {})[str(update.effective_chat.id)] = msg.message_id
        save_state(st)

async def cmd_finish(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        await update.message.reply_text("⚠️ Owner only")
        return
    
    job = get_job(st)
    if not job:
        await update.message.reply_text("📭 No active load")
        return
    
    await do_finish_load(update, ctx, st, job)

async def cmd_catalog(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        await update.message.reply_text("⚠️ Owner only")
        return
    
    history = st.get("history", [])
    if not history:
        await update.message.reply_text(
            """
<b>📭 NO HISTORY</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

Complete some loads first!
""",
            parse_mode="HTML"
        )
        return
    
    tz_name = (st.get("last_location") or {}).get("tz", "UTC")
    current_week = week_key(now_utc().astimezone(safe_tz(tz_name)))
    
    arg = ctx.args[0].lower() if ctx.args else ""
    if arg == "all":
        week = "ALL"
        records = history
    else:
        week = current_week
        records = [r for r in history if r.get("week") == week]
    
    if not records:
        records = history
        week = "ALL"
    
    xlsx_data, filename = make_catalog_xlsx(records, week)
    total_rate = sum(r.get("rate", 0) or 0 for r in records)
    total_miles = sum((r.get("posted_miles") or 0) for r in records)
    
    await update.message.reply_document(
        document=io.BytesIO(xlsx_data),
        filename=filename,
        caption=f"""
<b>📊 LOAD CATALOG</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

Period: <b>{week}</b>
Loads: <b>{len(records)}</b>
Gross: <b>{money(total_rate)}</b>
Miles: <b>{total_miles}</b>
""",
        parse_mode="HTML"
    )

async def cmd_skip(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        return
    
    job = get_job(st)
    if not job:
        return
    
    stage, idx = get_focus(job, st)
    if stage != "DEL":
        await update.message.reply_text("⚠️ Complete pickup first")
        return
    
    async with _state_lock:
        st = load_state()
        job = get_job(st)
        if job:
            dels = job.get("del", [])
            if idx < len(dels):
                dels[idx]["status"]["skip"] = True
                dels[idx]["status"]["comp"] = now_iso()
                st["job"] = job
                save_state(st)
    
    await update.message.reply_text(f"⏭ Skipped stop {idx+1}")

async def cmd_deleteall(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        await update.message.reply_text("⚠️ Owner only")
        return
    
    chat = update.effective_chat
    if not chat or chat.type == "private":
        await update.message.reply_text("❌ Use in group")
        return
    
    n = DELETEALL_DEFAULT
    if ctx.args:
        try:
            n = int(ctx.args[0])
        except: pass
    n = max(1, min(n, 2000))
    
    cmd_msg_id = update.message.message_id
    status_msg = await update.message.reply_text(f"🧹 Cleaning up...")
    status_msg_id = status_msg.message_id
    
    deleted = 0
    failed = 0
    
    for mid in range(cmd_msg_id - 1, max(0, cmd_msg_id - n - 1), -1):
        try:
            await ctx.bot.delete_message(chat.id, mid)
            deleted += 1
            await asyncio.sleep(0.05)
        except BadRequest:
            failed += 1
            if failed > 30:
                break
        except Forbidden:
            await status_msg.edit_text("❌ No permission")
            return
        except:
            failed += 1
            if failed > 30:
                break
    
    try:
        await ctx.bot.delete_message(chat.id, cmd_msg_id)
    except: pass
    
    try:
        await status_msg.edit_text(f"✅ Removed {deleted} messages")
        await asyncio.sleep(3)
        await ctx.bot.delete_message(chat.id, status_msg_id)
    except: pass

async def cmd_leave(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        return
    
    chat = update.effective_chat
    if not chat or chat.type == "private":
        return
    
    async with _state_lock:
        st = load_state()
        allowed = set(st.get("allowed_chats", []))
        allowed.discard(chat.id)
        st["allowed_chats"] = list(allowed)
        save_state(st)
    
    await update.message.reply_text("👋 Goodbye!")
    try:
        await ctx.bot.leave_chat(chat.id)
    except: pass

async def cmd_clearcache(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        await update.message.reply_text("⚠️ Owner only")
        return
    
    async with _state_lock:
        st = load_state()
        old = len(st.get("geocode_cache", {}))
        st["geocode_cache"] = {}
        save_state(st)
    
    await update.message.reply_text(f"✅ Cleared {old} cached addresses")

# ============================================================================
# LOCATION HANDLER
# ============================================================================
async def on_location(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.location:
        return
    
    async with _state_lock:
        st = load_state()
    
    if not is_owner(update, st):
        return
    
    loc = update.message.location
    tz = TF.timezone_at(lat=loc.latitude, lng=loc.longitude) or "UTC"
    
    async with _state_lock:
        st = load_state()
        st["last_location"] = {
            "lat": loc.latitude,
            "lon": loc.longitude,
            "tz": tz,
            "updated_at": now_iso()
        }
        save_state(st)
    
    log(f"Location: {loc.latitude:.4f}, {loc.longitude:.4f}")
    
    if update.effective_chat.type == "private":
        await update.message.reply_text(
            f"""
<b>✅ LOCATION SAVED</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

📍 {loc.latitude:.4f}, {loc.longitude:.4f}
🕐 {tz}

You're all set!
""",
            parse_mode="HTML",
            reply_markup=ReplyKeyboardRemove()
        )

# ============================================================================
# TEXT HANDLER
# ============================================================================
async def on_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return
    
    text = update.message.text
    chat = update.effective_chat
    
    async with _state_lock:
        st = load_state()
    
    if chat and chat.type in ("group", "supergroup"):
        if chat.id not in (st.get("allowed_chats") or []):
            return
        
        job = parse_load(text)
        if job:
            async with _state_lock:
                st = load_state()
                st["job"] = job
                st["focus_i"] = 0
                st["reminders_sent"] = {}
                st["geofence_state"] = {}
                save_state(st)
            
            await update.message.reply_text(
                build_new_load_message(job),
                parse_mode="HTML"
            )
            return
    
    if not chat_allowed(update, st):
        return
    
    first_word = text.strip().split()[0].lower() if text.strip() else ""
    first_word = re.sub(r"[^\w]", "", first_word)
    
    if first_word in TRIGGERS:
        is_all = "all" in text.lower()
        await send_eta_response(update, ctx, st, all_stops=is_all)

# ============================================================================
# CALLBACK HANDLER
# ============================================================================
async def on_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    
    data = query.data
    
    # Handle NOOP buttons
    if data == "NOOP":
        await query.answer()
        return
    
    log_debug(f"Callback: {data}")
    
    await query.answer()
    
    async with _state_lock:
        st = load_state()
    
    if not chat_allowed(update, st):
        return
    
    try:
        if data.startswith("ETA:"):
            await handle_eta_callback(update, ctx, st, data)
        elif data == "CATALOG":
            await handle_catalog_callback(update, ctx, st)
        elif data == "FINISH":
            await handle_finish_callback(update, ctx, st)
        elif data.startswith("NAV:"):
            await handle_nav_callback(update, ctx, st, data)
        else:
            await handle_status_callback(update, ctx, st, data)
    except Exception as e:
        log_error(f"Callback error: {data}", e)

async def handle_eta_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE, st: dict, data: str):
    await send_eta_response(update, ctx, st, all_stops=(data == "ETA:ALL"))

async def handle_catalog_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE, st: dict):
    if not is_owner(update, st):
        return
    
    history = st.get("history", [])
    if not history:
        await update.effective_message.reply_text("📭 No completed loads")
        return
    
    tz_name = (st.get("last_location") or {}).get("tz", "UTC")
    week = week_key(now_utc().astimezone(safe_tz(tz_name)))
    records = [r for r in history if r.get("week") == week] or history
    
    xlsx_data, filename = make_catalog_xlsx(records, week if records else "ALL")
    total_rate = sum(r.get("rate", 0) or 0 for r in records)
    
    await ctx.bot.send_document(
        chat_id=update.effective_chat.id,
        document=io.BytesIO(xlsx_data),
        filename=filename,
        caption=f"📊 {len(records)} loads • {money(total_rate)}"
    )

async def handle_finish_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE, st: dict):
    if not is_owner(update, st):
        return
    
    job = get_job(st)
    if job:
        await do_finish_load(update, ctx, st, job)

async def handle_nav_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE, st: dict, data: str):
    try:
        new_idx = int(data.split(":")[1])
    except:
        return
    
    async with _state_lock:
        st = load_state()
        st["focus_i"] = new_idx
        save_state(st)
        job = get_job(st)
    
    if job:
        try:
            await update.callback_query.edit_message_text(
                build_panel_message(job, st),
                parse_mode="HTML",
                reply_markup=build_panel_keyboard(job, st)
            )
        except: pass

async def handle_status_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE, st: dict, data: str):
    async with _state_lock:
        st = load_state()
        job = get_job(st)
        if not job:
            return
        
        stage, idx = get_focus(job, st)
        tz_name = (st.get("last_location") or {}).get("tz", "UTC")
        ts = local_stamp(tz_name)
        alert_msg = None
        
        if data.startswith("PU:"):
            ps = job["pu"].setdefault("status", {})
            if data == "PU:ARR":
                if toggle_timestamp(ps, "arr"):
                    alert_msg = build_alert_message("📍", "ARRIVED AT PICKUP", "", ts)
            elif data == "PU:LOAD":
                if toggle_timestamp(ps, "load"):
                    alert_msg = build_alert_message("📦", "LOADED", "", ts)
            elif data == "PU:DEP":
                if toggle_timestamp(ps, "dep"):
                    alert_msg = build_alert_message("🚛", "DEPARTED PICKUP", "", ts)
            elif data == "PU:COMP":
                if toggle_timestamp(ps, "comp"):
                    alert_msg = build_alert_message("✅", "PICKUP COMPLETE", "", ts)
        
        elif data.startswith("DEL:"):
            if stage != "DEL":
                save_state(st)
                return
            
            dels = job.get("del", [])
            if idx >= len(dels):
                save_state(st)
                return
            
            ds = dels[idx].setdefault("status", {})
            lbl = f"Stop {idx+1}" if len(dels) > 1 else "Delivery"
            
            if data == "DEL:ARR":
                if toggle_timestamp(ds, "arr"):
                    alert_msg = build_alert_message("📍", f"ARRIVED AT {lbl.upper()}", "", ts)
            elif data == "DEL:DEL":
                if toggle_timestamp(ds, "del"):
                    alert_msg = build_alert_message("📦", "DELIVERED", "", ts)
            elif data == "DEL:DEP":
                if toggle_timestamp(ds, "dep"):
                    alert_msg = build_alert_message("🚛", f"DEPARTED {lbl.upper()}", "", ts)
            elif data == "DEL:COMP":
                if toggle_timestamp(ds, "comp"):
                    alert_msg = build_alert_message("✅", f"{lbl.upper()} COMPLETE", "", ts)
                    for i in range(idx + 1, len(dels)):
                        if not dels[i].get("status", {}).get("comp"):
                            st["focus_i"] = i
                            break
            elif data == "DEL:SKIP":
                ds["skip"] = True
                ds["comp"] = ds.get("comp") or now_iso()
                alert_msg = build_alert_message("⏭", f"SKIPPED {lbl.upper()}", "", ts)
                for i in range(idx + 1, len(dels)):
                    if not dels[i].get("status", {}).get("comp"):
                        st["focus_i"] = i
                        break
        
        elif data.startswith("DOC:"):
            if data == "DOC:PTI":
                job["pu"].setdefault("docs", {})["pti"] = not job["pu"].get("docs", {}).get("pti", False)
            elif data == "DOC:BOL":
                job["pu"].setdefault("docs", {})["bol"] = not job["pu"].get("docs", {}).get("bol", False)
            elif data == "DOC:POD" and stage == "DEL":
                dels = job.get("del", [])
                if idx < len(dels):
                    dels[idx].setdefault("docs", {})["pod"] = not dels[idx].get("docs", {}).get("pod", False)
        
        st["job"] = job
        save_state(st)
    
    if alert_msg:
        await send_alert(ctx, update.effective_chat.id, alert_msg)
    
    try:
        await update.callback_query.edit_message_text(
            build_panel_message(job, st),
            parse_mode="HTML",
            reply_markup=build_panel_keyboard(job, st)
        )
    except: pass

# ============================================================================
# ETA RESPONSE
# ============================================================================
async def send_eta_response(update: Update, ctx: ContextTypes.DEFAULT_TYPE, st: dict, all_stops: bool = False):
    loc = st.get("last_location")
    if not loc:
        await update.effective_message.reply_text(
            """
<b>📍 LOCATION NEEDED</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

Owner: DM me /update to share location
""",
            parse_mode="HTML"
        )
        return
    
    loc_age_min = 999
    if loc.get("updated_at"):
        try:
            loc_age_min = int((now_utc() - datetime.fromisoformat(loc["updated_at"])).total_seconds() / 60)
        except: pass
    
    origin = (loc["lat"], loc["lon"])
    tz_name = loc.get("tz", "UTC")
    tz = safe_tz(tz_name)
    
    try:
        await ctx.bot.send_location(update.effective_chat.id, origin[0], origin[1])
    except: pass
    
    job = get_job(st)
    if not job:
        local_time = now_utc().astimezone(tz).strftime("%I:%M %p")
        await update.effective_message.reply_text(
            f"""
<b>🕐 {local_time}</b>
<code>━━━━━━━━━━━━━━━━━━━━━━</code>

No active load.
Forward a load sheet to get started.
""",
            parse_mode="HTML"
        )
        return
    
    age_warn = ""
    if loc_age_min > 10:
        age_warn = f"⚠️ <i>Location is {loc_age_min}m old</i>"
    
    if all_stops:
        await send_all_etas(update, ctx, st, job, origin, tz, tz_name, age_warn)
    else:
        await send_single_eta(update, ctx, st, job, origin, tz, tz_name, age_warn)

async def send_all_etas(update, ctx, st, job, origin, tz, tz_name, age_warn=""):
    etas = []
    
    pu = job.get("pu", {})
    if pu.get("status", {}).get("comp"):
        etas.append({"label": "Pickup", "complete": True})
    else:
        eta = await calc_eta(st, origin, pu.get("addr", ""))
        etas.append({
            "label": "Pickup",
            "ok": eta["ok"],
            "seconds": eta.get("seconds"),
            "meters": eta.get("meters"),
            "method": eta.get("method"),
            "err": eta.get("err"),
        })
    
    for i, d in enumerate(job.get("del", [])[:ETA_ALL_MAX]):
        label = f"Stop {i+1}" if len(job.get("del", [])) > 1 else "Delivery"
        if d.get("status", {}).get("comp"):
            etas.append({"label": label, "complete": True})
        else:
            eta = await calc_eta(st, origin, d.get("addr", ""))
            etas.append({
                "label": label,
                "ok": eta["ok"],
                "seconds": eta.get("seconds"),
                "meters": eta.get("meters"),
                "method": eta.get("method"),
                "err": eta.get("err"),
            })
    
    async with _state_lock:
        st2 = load_state()
        st2["geocode_cache"] = st.get("geocode_cache", {})
        save_state(st2)
    
    await update.effective_message.reply_text(
        build_all_etas_message(etas, job, age_warn),
        parse_mode="HTML",
        reply_markup=build_panel_keyboard(job, st)
    )

async def send_single_eta(update, ctx, st, job, origin, tz, tz_name, age_warn=""):
    stage, idx = get_focus(job, st)
    
    if stage == "PU":
        addr = job.get("pu", {}).get("addr", "")
        appt = job.get("pu", {}).get("time")
        label = "📦 PICKUP"
    else:
        dels = job.get("del", [])
        d = dels[idx] if idx < len(dels) else {}
        addr = d.get("addr", "")
        appt = d.get("time")
        label = f"🏁 STOP {idx+1}" if len(dels) > 1 else "🏁 DELIVERY"
    
    eta = await calc_eta(st, origin, addr)
    
    async with _state_lock:
        st2 = load_state()
        st2["geocode_cache"] = st.get("geocode_cache", {})
        save_state(st2)
    
    await update.effective_message.reply_text(
        build_eta_message(eta, label, addr, appt, job, tz_name, age_warn),
        parse_mode="HTML",
        reply_markup=build_panel_keyboard(job, st)
    )

# ============================================================================
# FINISH LOAD
# ============================================================================
async def do_finish_load(update: Update, ctx: ContextTypes.DEFAULT_TYPE, st: dict, job: dict):
    tz_name = (st.get("last_location") or {}).get("tz", "UTC")
    dt = now_utc().astimezone(safe_tz(tz_name))
    wk = week_key(dt)
    
    meta = job.get("meta", {})
    pu = job.get("pu", {})
    dels = job.get("del", [])
    
    record = {
        "week": wk,
        "completed": dt.strftime("%Y-%m-%d %H:%M"),
        "completed_utc": now_iso(),
        "load_number": meta.get("load_number", ""),
        "pickup": pu.get("addr", ""),
        "deliveries": " | ".join(d.get("addr", "") for d in dels),
        "rate": meta.get("rate"),
        "posted_miles": meta.get("miles"),
    }
    
    async with _state_lock:
        st = load_state()
        history = st.setdefault("history", [])
        history.append(record)
        st["history"] = history[-1000:]
        st["job"] = None
        st["focus_i"] = 0
        st["reminders_sent"] = {}
        st["geofence_state"] = {}
        save_state(st)
    
    chat_id = update.effective_chat.id
    panel_id = st.get("panel_msgs", {}).get(str(chat_id))
    
    msg_text = build_finish_message(job, st)
    
    if panel_id:
        try:
            await ctx.bot.edit_message_text(
                chat_id=chat_id, message_id=panel_id,
                text=msg_text, parse_mode="HTML",
                reply_markup=build_done_keyboard()
            )
            return
        except: pass
    
    await ctx.bot.send_message(chat_id, msg_text, parse_mode="HTML", reply_markup=build_done_keyboard())

# ============================================================================
# BACKGROUND JOBS
# ============================================================================
async def reminder_job(ctx: ContextTypes.DEFAULT_TYPE):
    try:
        async with _state_lock:
            st = load_state()
        
        job = get_job(st)
        if not job:
            return
        
        tz_name = (st.get("last_location") or {}).get("tz", "UTC")
        sent = st.get("reminders_sent", {})
        chats = get_broadcast_chats(st)
        alerts = []
        
        for threshold in REMINDER_THRESHOLDS_MIN:
            pu = job.get("pu", {})
            pu_time = pu.get("time")
            if pu_time and not pu.get("status", {}).get("comp"):
                # Handle both dict (new) and string (legacy) time formats
                if isinstance(pu_time, dict):
                    deadline = get_appointment_deadline(pu_time, tz_name)
                else:
                    deadline = parse_appt_time(pu_time, tz_name)
                
                if deadline:
                    mins = (deadline - now_utc()).total_seconds() / 60
                    key = f"appt:pu:{threshold}"
                    if threshold - 5 < mins <= threshold and key not in sent:
                        alerts.append((key, build_alert_message("⏰", f"PICKUP DEADLINE IN ~{int(mins)}m", "", "")))
            
            # Also check delivery windows
            for i, d in enumerate(job.get("del", [])):
                del_time = d.get("time")
                if del_time and not d.get("status", {}).get("comp"):
                    if isinstance(del_time, dict):
                        deadline = get_appointment_deadline(del_time, tz_name)
                    else:
                        deadline = parse_appt_time(del_time, tz_name)
                    
                    if deadline:
                        mins = (deadline - now_utc()).total_seconds() / 60
                        key = f"appt:del{i}:{threshold}"
                        if threshold - 5 < mins <= threshold and key not in sent:
                            lbl = f"STOP {i+1}" if len(job.get("del", [])) > 1 else "DELIVERY"
                            alerts.append((key, build_alert_message("⏰", f"{lbl} DEADLINE IN ~{int(mins)}m", "", "")))
        
        if alerts:
            async with _state_lock:
                st = load_state()
                sent = st.setdefault("reminders_sent", {})
                for key, msg in alerts:
                    sent[key] = True
                    for chat_id in chats:
                        try:
                            await ctx.bot.send_message(chat_id, msg, parse_mode="HTML")
                        except: pass
                save_state(st)
    except Exception as e:
        log_error("Reminder job", e)

def parse_appt_time(time_str: str, tz_name: str) -> Optional[datetime]:
    if not time_str:
        return None
    
    m = re.search(r"(\w{3})\s+(\d{1,2}),?\s+(\d{4})\s+(\d{1,2}):(\d{2})", time_str)
    if m:
        try:
            dt_str = f"{m.group(1)} {m.group(2)} {m.group(3)} {m.group(4)}:{m.group(5)}"
            dt = datetime.strptime(dt_str, "%b %d %Y %H:%M")
            return dt.replace(tzinfo=safe_tz(tz_name))
        except: pass
    return None

async def geofence_job(ctx: ContextTypes.DEFAULT_TYPE):
    try:
        async with _state_lock:
            st = load_state()
        
        job = get_job(st)
        loc = st.get("last_location")
        
        if not job or not loc:
            return
        
        try:
            loc_time = datetime.fromisoformat(loc["updated_at"])
            if (now_utc() - loc_time).total_seconds() > 300:
                return
        except:
            return
        
        origin = (loc["lat"], loc["lon"])
        gf_state = st.get("geofence_state", {})
        chats = get_broadcast_chats(st)
        events = []
        cache = st.get("geocode_cache", {})
        
        pu = job.get("pu", {})
        if pu.get("addr") and not pu.get("status", {}).get("comp"):
            geo = await geocode(pu["addr"], cache)
            if geo:
                dist = haversine_miles(origin[0], origin[1], geo[0], geo[1])
                was_in = gf_state.get("pu", False)
                is_in = dist <= GEOFENCE_MILES
                
                if is_in and not was_in:
                    events.append(("pu", True, "PICKUP"))
                elif not is_in and was_in:
                    events.append(("pu", False, "PICKUP"))
                gf_state["pu"] = is_in
        
        for i, d in enumerate(job.get("del", [])):
            if d.get("addr") and not d.get("status", {}).get("comp"):
                geo = await geocode(d["addr"], cache)
                if geo:
                    dist = haversine_miles(origin[0], origin[1], geo[0], geo[1])
                    key = f"del{i}"
                    was_in = gf_state.get(key, False)
                    is_in = dist <= GEOFENCE_MILES
                    
                    lbl = f"STOP {i+1}" if len(job.get("del", [])) > 1 else "DELIVERY"
                    if is_in and not was_in:
                        events.append((key, True, lbl))
                    elif not is_in and was_in:
                        events.append((key, False, lbl))
                    gf_state[key] = is_in
        
        if events:
            async with _state_lock:
                st = load_state()
                job = get_job(st)
                if not job:
                    return
                
                for key, entered, label in events:
                    if entered:
                        msg = build_alert_message("📍", f"ARRIVING AT {label}", f"Within {GEOFENCE_MILES} miles", "")
                        if key == "pu":
                            job["pu"].setdefault("status", {})["arr"] = now_iso()
                        elif key.startswith("del"):
                            idx = int(key[3:])
                            if idx < len(job.get("del", [])):
                                job["del"][idx].setdefault("status", {})["arr"] = now_iso()
                    else:
                        msg = build_alert_message("🚛", f"DEPARTED {label}", "", "")
                        if key == "pu":
                            job["pu"].setdefault("status", {})["dep"] = now_iso()
                        elif key.startswith("del"):
                            idx = int(key[3:])
                            if idx < len(job.get("del", [])):
                                job["del"][idx].setdefault("status", {})["dep"] = now_iso()
                    
                    for chat_id in chats:
                        try:
                            await ctx.bot.send_message(chat_id, msg, parse_mode="HTML")
                        except: pass
                
                st["job"] = job
                st["geofence_state"] = gf_state
                st["geocode_cache"] = cache
                save_state(st)
    except Exception as e:
        log_error("Geofence job", e)

# ============================================================================
# MAIN
# ============================================================================
async def post_init(app: Application):
    try:
        await app.bot.delete_webhook(drop_pending_updates=True)
        me = await app.bot.get_me()
        log(f"Bot: @{me.username}")
    except Exception as e:
        log_error("Init", e)
    
    if app.job_queue:
        app.job_queue.run_repeating(reminder_job, interval=60, first=10)
        app.job_queue.run_repeating(geofence_job, interval=30, first=15)
    
    log(f"Ready! v{BOT_VERSION}")

def main():
    if not TOKEN:
        print("ERROR: TELEGRAM_TOKEN not set")
        return
    
    log(f"Starting v{BOT_VERSION}...")
    
    app = ApplicationBuilder().token(TOKEN).post_init(post_init).build()
    
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_start))
    app.add_handler(CommandHandler("ping", cmd_ping))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("testgeo", cmd_testgeo))
    app.add_handler(CommandHandler("claim", cmd_claim))
    app.add_handler(CommandHandler("allowhere", cmd_allowhere))
    app.add_handler(CommandHandler("update", cmd_update))
    app.add_handler(CommandHandler("panel", cmd_panel))
    app.add_handler(CommandHandler("finish", cmd_finish))
    app.add_handler(CommandHandler("catalog", cmd_catalog))
    app.add_handler(CommandHandler("skip", cmd_skip))
    app.add_handler(CommandHandler("deleteall", cmd_deleteall))
    app.add_handler(CommandHandler("leave", cmd_leave))
    app.add_handler(CommandHandler("clearcache", cmd_clearcache))
    
    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.LOCATION, on_location))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
