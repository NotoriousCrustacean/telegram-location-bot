import asyncio
import hashlib
import html
import io
import json
import math
import os
import re
import time
import weakref
from datetime import datetime, timezone, timedelta, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Set

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from timezonefinder import TimezoneFinder
from zoneinfo import ZoneInfo

from telegram import (
    Update,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.error import Forbidden, BadRequest, TelegramError
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

BOT_VERSION = "2025-12-11_improved_v1"


# ----------------------------
# Constants
# ----------------------------
EARTH_RADIUS_M = 6371000.0
METERS_PER_MILE = 1609.344
MAX_HISTORY_RECORDS = 1000
MAX_DELETEALL_MESSAGES = 2000
EXCEL_SHEET_NAME_MAX_LEN = 31
DELETE_MESSAGE_DELAY_SEC = 0.02


# ----------------------------
# Environment helpers
# ----------------------------
def _strip_quotes(s: str) -> str:
    """Remove surrounding quotes from a string."""
    s = (s or "").strip()
    if len(s) >= 2 and ((s[0] == s[-1] == '"') or (s[0] == s[-1] == "'")):
        return s[1:-1].strip()
    return s


def env_str(name: str, default: str = "") -> str:
    """Get string environment variable with optional quote stripping."""
    v = os.environ.get(name)
    if v is None:
        return default
    return _strip_quotes(v)


def env_int(name: str, default: int) -> int:
    """Get integer environment variable."""
    v = env_str(name, "")
    if not v:
        return default
    try:
        return int(v)
    except ValueError:
        return default


def env_float(name: str, default: float) -> float:
    """Get float environment variable."""
    v = env_str(name, "")
    if not v:
        return default
    try:
        return float(v)
    except ValueError:
        return default


def env_bool(name: str, default: bool = False) -> bool:
    """Get boolean environment variable."""
    v = env_str(name, "")
    if not v:
        return default
    return v.lower() in ("1", "true", "yes", "y", "on")


# ----------------------------
# ENV Configuration
# ----------------------------
TOKEN = env_str("TELEGRAM_TOKEN", "")
CLAIM_CODE = env_str("CLAIM_CODE", "")

STATE_FILE = Path(env_str("STATE_FILE", "state.json"))
STATE_FALLBACK = Path("/tmp/dispatch_bot_state.json")

TRIGGERS: Set[str] = {t.strip().lower() for t in env_str("TRIGGERS", "eta,1717").split(",") if t.strip()}

NOMINATIM_USER_AGENT = env_str("NOMINATIM_USER_AGENT", "dispatch-eta-bot/1.0")
NOMINATIM_MIN_INTERVAL = env_float("NOMINATIM_MIN_INTERVAL", 1.1)

ETA_ALL_MAX = env_int("ETA_ALL_MAX", 6)
DELETEALL_DEFAULT = env_int("DELETEALL_DEFAULT", 300)
ALERT_TTL_SECONDS = env_int("ALERT_TTL_SECONDS", 25)

DEBUG = env_bool("DEBUG", False)


def log(msg: str) -> None:
    """Debug logging."""
    if DEBUG:
        print(f"[bot {BOT_VERSION}] {msg}", flush=True)


# ----------------------------
# Globals
# ----------------------------
TF = TimezoneFinder()
NOM_URL = "https://nominatim.openstreetmap.org/search"
OSRM_URL = "https://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}"

_state_lock = asyncio.Lock()
_geo_lock = asyncio.Lock()
_geo_last = 0.0

# Track background tasks to prevent garbage collection
_background_tasks: Set[asyncio.Task] = set()


# ----------------------------
# Time helpers
# ----------------------------
def now_utc() -> datetime:
    """Get current UTC datetime."""
    return datetime.now(timezone.utc)


def now_iso() -> str:
    """Get current UTC datetime as ISO string."""
    return now_utc().isoformat()


def safe_tz(name: str) -> timezone:
    """Safely get a timezone, defaulting to UTC."""
    if not name:
        return timezone.utc
    try:
        return ZoneInfo(name)
    except Exception:
        return timezone.utc


def h(x: Any) -> str:
    """HTML-escape a value for Telegram messages."""
    return html.escape("" if x is None else str(x), quote=False)


def local_stamp(tz_name: str) -> str:
    """Get current local time as a formatted string."""
    tz = safe_tz(tz_name or "UTC")
    return now_utc().astimezone(tz).strftime("%Y-%m-%d %H:%M")


def week_key(dt: datetime) -> str:
    """Get ISO week key (e.g., '2025-W50')."""
    iso = dt.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def money(x: Optional[float]) -> str:
    """Format a monetary value."""
    if x is None:
        return "-"
    try:
        return f"${float(x):,.0f}"
    except (ValueError, TypeError):
        return str(x)


# ----------------------------
# State load/save
# ----------------------------
def _migrate_state(st: dict) -> Tuple[dict, bool]:
    """Migrate old state formats to current format."""
    changed = False

    # Owner aliases
    if st.get("owner_id") is None and st.get("owner") is not None:
        st["owner_id"] = st.get("owner")
        changed = True
    if st.get("owner") is None and st.get("owner_id") is not None:
        st["owner"] = st.get("owner_id")
        changed = True

    # Allowed chats aliases
    if (not st.get("allowed_chats")) and st.get("allowed"):
        st["allowed_chats"] = st.get("allowed")
        changed = True
    if (not st.get("allowed")) and st.get("allowed_chats"):
        st["allowed"] = st.get("allowed_chats")
        changed = True

    # Last location aliases
    if st.get("last_location") is None and st.get("last") is not None:
        ll = st.get("last") or {}
        st["last_location"] = {
            "lat": ll.get("lat"),
            "lon": ll.get("lon"),
            "tz": ll.get("tz"),
            "updated_at": ll.get("at") or ll.get("updated_at") or ll.get("timestamp"),
        }
        changed = True
    if st.get("last") is None and st.get("last_location") is not None:
        ll = st.get("last_location") or {}
        st["last"] = {
            "lat": ll.get("lat"),
            "lon": ll.get("lon"),
            "tz": ll.get("tz"),
            "at": ll.get("updated_at") or ll.get("at"),
        }
        changed = True

    # Geocode cache aliases
    if (not st.get("geocode_cache")) and st.get("gc"):
        st["geocode_cache"] = st.get("gc")
        changed = True
    if (not st.get("gc")) and st.get("geocode_cache"):
        st["gc"] = st.get("geocode_cache")
        changed = True

    # History aliases
    if (not st.get("history")) and st.get("hist"):
        st["history"] = st.get("hist")
        changed = True
    if (not st.get("hist")) and st.get("history"):
        st["hist"] = st.get("history")
        changed = True

    if st.get("focus_i") is None and st.get("del_index") is not None:
        st["focus_i"] = st.get("del_index")
        changed = True

    # Set defaults
    st.setdefault("owner_id", None)
    st.setdefault("allowed_chats", [])
    st.setdefault("last_location", None)
    st.setdefault("job", None)
    st.setdefault("focus_i", 0)
    st.setdefault("geocode_cache", {})
    st.setdefault("history", [])
    st.setdefault("last_finished", None)
    st.setdefault("panel_messages", {})

    # Mirror legacy keys
    st["owner"] = st.get("owner_id")
    st["allowed"] = st.get("allowed_chats")
    st["gc"] = st.get("geocode_cache")
    st["hist"] = st.get("history")

    return st, changed


def load_state() -> dict:
    """Load state from file with fallback."""
    global STATE_FILE

    if (not STATE_FILE.exists()) and STATE_FALLBACK.exists():
        STATE_FILE = STATE_FALLBACK

    if STATE_FILE.exists():
        try:
            st = json.loads(STATE_FILE.read_text(encoding="utf-8"))
            if not isinstance(st, dict):
                st = {}
        except (json.JSONDecodeError, OSError) as e:
            log(f"Error loading state: {e}")
            st = {}
    else:
        st = {}

    st, changed = _migrate_state(st)
    if changed:
        try:
            save_state(st)
        except OSError:
            pass
    return st


def save_state(st: dict) -> None:
    """Save state to file with fallback."""
    global STATE_FILE

    def _write(path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = path.with_suffix(".tmp")
        tmp.write_text(json.dumps(st, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(path)

    try:
        _write(STATE_FILE)
    except OSError as e:
        log(f"save_state failed at {STATE_FILE}: {e}. Falling back to {STATE_FALLBACK}")
        STATE_FILE = STATE_FALLBACK
        _write(STATE_FILE)


def is_owner(update: Update, st: dict) -> bool:
    """Check if the current user is the bot owner."""
    u = update.effective_user
    return bool(u and st.get("owner_id") and u.id == st["owner_id"])


def chat_allowed(update: Update, st: dict) -> bool:
    """Check if the current chat is allowed."""
    chat = update.effective_chat
    if not chat:
        return False
    if chat.type == "private":
        return is_owner(update, st)
    allowed_set = set(st.get("allowed_chats") or [])
    return chat.id in allowed_set


# ----------------------------
# Geocode/routing
# ----------------------------
def hav_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate Haversine distance in meters."""
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * EARTH_RADIUS_M * math.asin(math.sqrt(a))


def fallback_seconds(dist_m: float) -> float:
    """Estimate travel time based on distance."""
    km = dist_m / 1000.0
    # Speed varies by distance: local, regional, highway
    if km < 80:
        speed_kmh = 55
    elif km < 320:
        speed_kmh = 85
    else:
        speed_kmh = 105
    return (km / speed_kmh) * 3600.0


def fmt_dur(seconds: float) -> str:
    """Format duration in hours and minutes."""
    seconds = max(0, int(seconds))
    m = seconds // 60
    h_ = m // 60
    m = m % 60
    return f"{h_}h {m}m" if h_ else f"{m}m"


def fmt_mi(meters: float) -> str:
    """Format distance in miles."""
    mi = meters / METERS_PER_MILE
    return f"{mi:.1f} mi" if mi < 10 else f"{mi:.0f} mi"


def addr_variants(addr: str) -> List[str]:
    """Generate address variants for geocoding."""
    a = " ".join((addr or "").split())
    if not a:
        return []
    
    out: List[str] = [a]
    parts = [p.strip() for p in a.split(",") if p.strip()]
    
    if len(parts) >= 2:
        out.append(", ".join(parts[1:]))
    
    # Remove suite/unit numbers
    cleaned = re.sub(r"\b(?:suite|ste|unit|#)\s*[\w\-]+\b", "", a, flags=re.I).strip()
    if cleaned:
        out.append(cleaned)
    
    if len(parts) >= 2:
        out.append(", ".join(parts[-2:]))
    
    if "usa" not in a.lower():
        out.append(a + ", USA")

    # Deduplicate while preserving order
    seen: Set[str] = set()
    res: List[str] = []
    for x in out:
        x = " ".join(x.split())
        if x and x not in seen:
            seen.add(x)
            res.append(x)
    return res


async def geocode_cached(st: dict, addr: str) -> Optional[Tuple[float, float, str]]:
    """Geocode an address with caching."""
    cache = st.get("geocode_cache") or {}
    
    # Check cache first
    if addr in cache and isinstance(cache[addr], dict):
        try:
            v = cache[addr]
            return float(v["lat"]), float(v["lon"]), (v.get("tz") or "UTC")
        except (KeyError, ValueError, TypeError):
            pass

    if not NOMINATIM_USER_AGENT:
        return None

    headers = {"User-Agent": NOMINATIM_USER_AGENT}
    
    async with httpx.AsyncClient(timeout=15, headers=headers) as client:
        for q in addr_variants(addr):
            async with _geo_lock:
                global _geo_last
                wait = (_geo_last + NOMINATIM_MIN_INTERVAL) - time.monotonic()
                if wait > 0:
                    await asyncio.sleep(wait)
                
                try:
                    r = await client.get(NOM_URL, params={"q": q, "format": "jsonv2", "limit": 1})
                except httpx.RequestError as e:
                    log(f"Geocode request error: {e}")
                    continue
                finally:
                    _geo_last = time.monotonic()

            if r.status_code >= 400:
                continue
            
            try:
                js = r.json() or []
            except json.JSONDecodeError:
                continue
                
            if not js:
                continue

            try:
                lat, lon = float(js[0]["lat"]), float(js[0]["lon"])
            except (KeyError, ValueError, TypeError, IndexError):
                continue
                
            tz = TF.timezone_at(lat=lat, lng=lon) or "UTC"
            cache_entry = {"lat": lat, "lon": lon, "tz": tz}

            # Update state with new cache entry
            async with _state_lock:
                st2 = load_state()
                gc = st2.get("geocode_cache") or {}
                gc[addr] = cache_entry
                st2["geocode_cache"] = gc
                st2["gc"] = gc  # Mirror alias
                save_state(st2)

            # Also update the passed-in state dict for immediate use
            st["geocode_cache"] = st.get("geocode_cache") or {}
            st["geocode_cache"][addr] = cache_entry

            return lat, lon, tz

    return None


async def route(origin: Tuple[float, float], dest: Tuple[float, float]) -> Optional[Tuple[float, float]]:
    """Get driving route from OSRM."""
    url = OSRM_URL.format(lon1=origin[1], lat1=origin[0], lon2=dest[1], lat2=dest[0])
    
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(url, params={"overview": "false"})
            if r.status_code >= 400:
                return None
            js = r.json() or {}
            routes = js.get("routes") or []
            if not routes:
                return None
            return float(routes[0]["distance"]), float(routes[0]["duration"])
    except (httpx.RequestError, json.JSONDecodeError, KeyError, ValueError, TypeError, IndexError) as e:
        log(f"Route error: {e}")
        return None


async def eta_to(st: dict, origin: Tuple[float, float], label: str, addr: str) -> dict:
    """Calculate ETA to an address."""
    g = await geocode_cached(st, addr)
    if not g:
        return {"ok": False, "err": f"Couldn't locate {label}."}
    
    dest = (g[0], g[1])
    r = await route(origin, dest)
    
    if r:
        return {"ok": True, "m": r[0], "s": r[1], "method": "osrm", "tz": g[2]}
    
    dist = hav_m(origin[0], origin[1], dest[0], dest[1])
    return {"ok": True, "m": dist, "s": fallback_seconds(dist), "method": "approx", "tz": g[2]}


async def estimate_miles(st: dict, job: dict) -> Optional[float]:
    """Estimate total miles for a job."""
    pu = job.get("pu")
    if not pu or not pu.get("addr"):
        return None
        
    addrs = [pu["addr"]] + [d["addr"] for d in (job.get("del") or []) if d.get("addr")]
    
    coords: List[Tuple[float, float]] = []
    for a in addrs:
        g = await geocode_cached(st, a)
        if not g:
            return None
        coords.append((g[0], g[1]))
    
    if len(coords) < 2:
        return 0.0
    
    total_m = 0.0
    for a, b in zip(coords, coords[1:]):
        r = await route(a, b)
        total_m += r[0] if r else hav_m(a[0], a[1], b[0], b[1])
    
    return total_m / METERS_PER_MILE


# ----------------------------
# Load parsing
# ----------------------------
RATE_RE = re.compile(r"\b(?:RATE|PAY)\b\s*:\s*\$?\s*([0-9][0-9,]*(?:\.[0-9]{1,2})?)", re.I)
MILES_RE = re.compile(r"\b(?:LOADED|MILES)\b\s*:\s*([0-9][0-9,]*)", re.I)

PU_TIME_RE = re.compile(r"^\s*PU time:\s*(.+)$", re.I)
DEL_TIME_RE = re.compile(r"^\s*DEL time:\s*(.+)$", re.I)
PU_ADDR_RE = re.compile(r"^\s*PU Address\s*:\s*(.*)$", re.I)
DEL_ADDR_RE = re.compile(r"^\s*DEL Address(?:\s*\d+)?\s*:\s*(.*)$", re.I)

LOAD_NUM_RE = re.compile(r"^\s*Load Number\s*:\s*(.+)$", re.I)
LOAD_DATE_RE = re.compile(r"^\s*Load Date\s*:\s*(.+)$", re.I)
PICKUP_RE = re.compile(r"^\s*Pickup\s*:\s*(.+)$", re.I)
DELIVERY_RE = re.compile(r"^\s*Delivery\s*:\s*(.+)$", re.I)
TIMEISH = re.compile(r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{2}-\d{2}|\d{1,2}:\d{2})\b")


def extract_rate_miles(text: str) -> Tuple[Optional[float], Optional[int]]:
    """Extract rate and miles from load text."""
    rate: Optional[float] = None
    miles: Optional[int] = None
    
    m = RATE_RE.search(text)
    if m:
        try:
            rate = float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    
    m = MILES_RE.search(text)
    if m:
        try:
            miles = int(m.group(1).replace(",", ""))
        except ValueError:
            pass
    
    return rate, miles


def take_block(lines: List[str], i: int, first: str) -> Tuple[List[str], int]:
    """Extract a block of address lines."""
    out: List[str] = []
    if first.strip():
        out.append(first.strip())
    
    j = i + 1
    while j < len(lines):
        s = lines[j].strip()
        if not s:
            break
        low = s.lower()
        if low.startswith(("pu time:", "del time:", "pu address", "del address", "pickup:", "delivery:")):
            break
        if set(s) <= {"-"} or set(s) <= {"="}:
            break
        out.append(s)
        j += 1
    
    return out, j


def init_job(job: dict) -> dict:
    """Initialize job with default status structures."""
    job.setdefault("meta", {})
    
    pu = job.setdefault("pu", {})
    pu.setdefault("addr", "")
    pu.setdefault("lines", [])
    pu.setdefault("time", None)
    pu.setdefault("status", {"arr": None, "load": None, "dep": None, "comp": None})
    pu.setdefault("docs", {"pti": False, "bol": False})

    dels = job.setdefault("del", [])
    for d in dels:
        d.setdefault("addr", "")
        d.setdefault("lines", [])
        d.setdefault("time", None)
        d.setdefault("status", {"arr": None, "del": None, "dep": None, "comp": None, "skip": False})
        d.setdefault("docs", {"pod": False})
    
    return job


def normalize_job(job: Optional[dict]) -> Optional[dict]:
    """Normalize and validate a job dictionary."""
    if not job or not isinstance(job, dict):
        return None
    if "pu" not in job or "del" not in job:
        return None
    return init_job(job)


def parse_detailed(text: str) -> Optional[dict]:
    """Parse detailed load format with PU Address/DEL Address."""
    low = text.lower()
    if "pu address" not in low or "del address" not in low:
        return None

    lines = [ln.rstrip() for ln in text.splitlines()]
    pu_time: Optional[str] = None
    cur_del_time: Optional[str] = None
    pu_addr: Optional[str] = None
    pu_lines: Optional[List[str]] = None
    dels: List[dict] = []
    load_num: Optional[str] = None
    load_date: Optional[str] = None

    for i, ln in enumerate(lines):
        m = LOAD_NUM_RE.match(ln)
        if m:
            load_num = m.group(1).strip()
            continue

        m = LOAD_DATE_RE.match(ln)
        if m:
            load_date = m.group(1).strip()
            continue

        m = PU_TIME_RE.match(ln)
        if m:
            pu_time = m.group(1).strip()
            continue

        m = DEL_TIME_RE.match(ln)
        if m:
            cur_del_time = m.group(1).strip()
            continue

        m = PU_ADDR_RE.match(ln)
        if m and not pu_addr:
            blk, _ = take_block(lines, i, m.group(1))
            if blk:
                pu_lines = blk
                pu_addr = ", ".join(blk)
            continue

        m = DEL_ADDR_RE.match(ln)
        if m:
            blk, _ = take_block(lines, i, m.group(1))
            if blk:
                dels.append({"addr": ", ".join(blk), "lines": blk, "time": cur_del_time})

    if not pu_addr or not dels:
        return None

    rate, miles = extract_rate_miles(text)
    meta: Dict[str, Any] = {"rate": rate, "miles": miles}
    if load_num:
        meta["load_number"] = load_num
    if load_date:
        meta["load_date"] = load_date

    jid = hashlib.sha1((pu_addr + "|" + "|".join(d["addr"] for d in dels)).encode()).hexdigest()[:10]
    job = {
        "id": jid,
        "meta": meta,
        "pu": {"addr": pu_addr, "lines": pu_lines or [pu_addr], "time": pu_time},
        "del": dels,
    }
    return init_job(job)


def parse_summary(text: str) -> Optional[dict]:
    """Parse summary load format with Pickup:/Delivery:."""
    low = text.lower()
    if "pickup:" not in low or "delivery:" not in low:
        return None

    meta: Dict[str, Any] = {}
    pu_addr: Optional[str] = None
    pu_time: Optional[str] = None
    load_date: Optional[str] = None
    dels: List[dict] = []
    pending: Optional[dict] = None

    for ln in [x.strip() for x in text.splitlines() if x.strip()]:
        m = LOAD_NUM_RE.match(ln)
        if m:
            meta["load_number"] = m.group(1).strip()
            continue

        m = LOAD_DATE_RE.match(ln)
        if m:
            load_date = m.group(1).strip()
            continue

        m = PICKUP_RE.match(ln)
        if m:
            v = m.group(1).strip()
            if TIMEISH.search(v):
                pu_time = v
            else:
                pu_addr = v
            continue

        m = DELIVERY_RE.match(ln)
        if m:
            v = m.group(1).strip()
            if TIMEISH.search(v):
                if pending and not pending.get("time"):
                    pending["time"] = v
                    pending = None
            else:
                pending = {"addr": v, "lines": [v], "time": None}
                dels.append(pending)

    if not pu_addr or not dels:
        return None

    rate, miles = extract_rate_miles(text)
    if rate is not None:
        meta["rate"] = rate
    if miles is not None:
        meta["miles"] = miles
    if load_date:
        meta["load_date"] = load_date

    jid = hashlib.sha1(
        (str(meta.get("load_number", "")) + "|" + pu_addr + "|" + "|".join(d["addr"] for d in dels)).encode()
    ).hexdigest()[:10]
    
    job = {
        "id": jid,
        "meta": meta,
        "pu": {"addr": pu_addr, "lines": [pu_addr], "time": pu_time},
        "del": dels,
    }
    return init_job(job)


def parse_job(text: str) -> Optional[dict]:
    """Parse a load from text using available formats."""
    return parse_detailed(text) or parse_summary(text)


# ----------------------------
# Workflow helpers + UI
# ----------------------------
def pu_complete(job: dict) -> bool:
    """Check if pickup is complete."""
    pu = job.get("pu")
    if not pu:
        return False
    status = pu.get("status")
    if not status:
        return False
    return bool(status.get("comp"))


def next_incomplete(job: dict, start: int = 0) -> Optional[int]:
    """Find the next incomplete delivery stop."""
    dels = job.get("del") or []
    for i, d in enumerate(dels):
        if i < start:
            continue
        status = d.get("status") or {}
        if not status.get("comp"):
            return i
    return None


def focus(job: dict, st: dict) -> Tuple[str, int]:
    """Determine the current focus stage and index."""
    if not pu_complete(job):
        return "PU", 0
    
    dels = job.get("del") or []
    if not dels:
        return "DEL", 0
    
    i = int(st.get("focus_i") or 0)
    i = max(0, min(i, len(dels) - 1))
    
    # If current stop is complete, find next incomplete
    if dels[i].get("status", {}).get("comp"):
        ni = next_incomplete(job, i + 1)
        if ni is not None:
            i = ni
    
    return "DEL", i


def load_id_text(job: dict) -> str:
    """Get display text for load identification."""
    m = job.get("meta") or {}
    if m.get("load_number"):
        return f"Load {m.get('load_number')}"
    return f"Job {job.get('id', 'unknown')}"


def toggle_ts(obj: dict, key: str) -> bool:
    """Toggle a timestamp field on/off."""
    if obj.get(key):
        obj[key] = None
        return False
    obj[key] = now_iso()
    return True


async def send_progress_alert(ctx: ContextTypes.DEFAULT_TYPE, chat_id: int, text: str) -> None:
    """Send a short alert message that auto-deletes."""
    try:
        m = await ctx.bot.send_message(chat_id=chat_id, text=text, parse_mode="HTML", disable_notification=True)
    except TelegramError as e:
        log(f"Failed to send progress alert: {e}")
        return

    if ALERT_TTL_SECONDS <= 0:
        return

    async def _delete_later() -> None:
        await asyncio.sleep(ALERT_TTL_SECONDS)
        try:
            await ctx.bot.delete_message(chat_id=chat_id, message_id=m.message_id)
        except TelegramError:
            pass

    # Track the task to prevent garbage collection
    task = asyncio.create_task(_delete_later())
    _background_tasks.add(task)
    task.add_done_callback(_background_tasks.discard)


def short_place(lines: List[str], addr: str) -> str:
    """Get a short place name from address lines."""
    for x in reversed(lines or []):
        x = (x or "").strip()
        if x and len(x) <= 70:
            return x
    return (addr or "").strip()[:70]


# ----------------------------
# Buttons
# ----------------------------
def b(label: str, data: str) -> InlineKeyboardButton:
    """Create an inline keyboard button."""
    return InlineKeyboardButton(label, callback_data=data)


def chk(on: bool, label: str) -> str:
    """Add checkmark prefix to label if condition is true."""
    return ("✅ " + label) if on else label


def build_finished_keyboard() -> InlineKeyboardMarkup:
    """Build keyboard for finished load state."""
    return InlineKeyboardMarkup([[b("📊 Catalog", "SHOW:CAT")]])


def build_keyboard(job: dict, st: dict) -> InlineKeyboardMarkup:
    """Build the main control keyboard."""
    stage, i = focus(job, st)
    pu = job.get("pu") or {}
    ps = pu.get("status") or {}
    pd = pu.get("docs") or {}

    rows: List[List[InlineKeyboardButton]] = []

    if stage == "PU":
        rows.append([
            b(chk(bool(ps.get("arr")), "Arrived PU"), "PU:A"),
            b(chk(bool(ps.get("load")), "Loaded"), "PU:L"),
            b(chk(bool(ps.get("dep")), "Departed"), "PU:D"),
        ])
        rows.append([
            b(chk(bool(pd.get("pti")), "PTI"), "DOC:PTI"),
            b(chk(bool(pd.get("bol")), "BOL"), "DOC:BOL"),
            b(chk(bool(ps.get("comp")), "PU Complete"), "PU:C"),
        ])
    else:
        dels = job.get("del") or []
        d = dels[i] if i < len(dels) else {"addr": "", "lines": []}
        ds = d.get("status") or {}
        dd = d.get("docs") or {}
        lbl = f"DEL {i+1}/{len(dels)}" if dels else "DEL"

        rows.append([
            b(chk(bool(ds.get("arr")), f"Arrived {lbl}"), "DEL:A"),
            b(chk(bool(ds.get("del")), "Delivered"), "DEL:DL"),
            b(chk(bool(ds.get("dep")), "Departed"), "DEL:D"),
        ])
        rows.append([
            b(chk(bool(dd.get("pod")), "POD"), "DOC:POD"),
            b(chk(bool(ds.get("comp")), "Stop Complete"), "DEL:C"),
            b("Skip Stop", "DEL:S"),
        ])

    rows.append([b("ETA", "ETA:A"), b("ETA all", "ETA:ALL")])
    rows.append([b("📊 Catalog", "SHOW:CAT"), b("Finish Load", "JOB:FIN")])
    return InlineKeyboardMarkup(rows)


# ----------------------------
# Finish + weekly totals
# ----------------------------
def week_totals(hist: List[dict], wk: str) -> Tuple[int, float, float]:
    """Calculate weekly totals from history."""
    count = 0
    sum_rate = 0.0
    sum_miles = 0.0
    
    for r in hist:
        if (r.get("week") or "") != wk:
            continue
        count += 1
        
        rate = r.get("rate")
        if isinstance(rate, (int, float)):
            sum_rate += float(rate)
        
        pm = r.get("posted_miles")
        em = r.get("est_miles")
        use = pm if isinstance(pm, (int, float)) else (em if isinstance(em, (int, float)) else None)
        if use is not None:
            sum_miles += float(use)
    
    return count, sum_rate, sum_miles


async def finish_active_load(
    update: Update, ctx: ContextTypes.DEFAULT_TYPE, *, source: str
) -> Optional[Tuple[dict, dict]]:
    """Finish the active load and record it in history."""
    async with _state_lock:
        st = load_state()

        if not is_owner(update, st):
            if source == "callback" and update.callback_query:
                await update.callback_query.answer("Owner only. DM /claim <code>.", show_alert=True)
            else:
                await update.effective_message.reply_text("Owner only. DM me: /claim <code>")
            return None

        job = normalize_job(st.get("job"))
        if not job:
            if source == "callback" and update.callback_query:
                await update.callback_query.answer("No active load.", show_alert=True)
            else:
                await update.effective_message.reply_text("No active load.")
            return None

        # Calculate estimated miles (this is async but we still hold the lock - acceptable for short operation)
        loc = st.get("last_location") or {}
        tz_name = loc.get("tz") or "UTC"
        
    # Release lock for potentially slow geocoding
    est = await estimate_miles(st, job)
    
    async with _state_lock:
        # Re-load state in case it changed
        st = load_state()
        job = normalize_job(st.get("job"))
        if not job:
            return None
            
        loc = st.get("last_location") or {}
        tz_name = loc.get("tz") or "UTC"
        dt_local = now_utc().astimezone(safe_tz(tz_name))
        wk = week_key(dt_local)

        meta = job.get("meta") or {}
        pu = job.get("pu") or {}
        dels = job.get("del") or []
        del_times = " | ".join(((d.get("time") or "").strip() or "-") for d in dels)

        rec = {
            "week": wk,
            "completed": dt_local.strftime("%Y-%m-%d %H:%M"),
            "completed_utc": now_iso(),
            "tz": tz_name,
            "load_number": meta.get("load_number") or "",
            "job_id": job.get("id"),
            "load_date": meta.get("load_date"),
            "pu_time": pu.get("time"),
            "pickup": pu.get("addr") or "",
            "deliveries": " | ".join((d.get("addr") or "") for d in dels),
            "del_times": del_times,
            "stops": len(dels),
            "rate": meta.get("rate"),
            "posted_miles": meta.get("miles"),
            "est_miles": est,
        }

        chat_id = update.effective_chat.id if update.effective_chat else None

        hist = list(st.get("history") or [])
        hist.append(rec)
        st["history"] = hist[-MAX_HISTORY_RECORDS:]
        st["hist"] = st["history"]  # Mirror alias
        st["last_finished"] = rec
        st["job"] = None
        st["focus_i"] = 0
        
        # Clear stored panel message for this chat
        if chat_id is not None:
            pm = st.get("panel_messages") or {}
            pm.pop(str(chat_id), None)
            st["panel_messages"] = pm
        
        save_state(st)

    count, sum_rate, sum_miles = week_totals(st.get("history") or [], wk)
    rec["_wk_count"] = count
    rec["_wk_rate"] = sum_rate
    rec["_wk_miles"] = sum_miles

    return rec, st


# ----------------------------
# Excel catalog
# ----------------------------
def try_parse_date(s: Any) -> Optional[date]:
    """Try to parse a date from various formats."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def try_parse_dt(s: Any) -> Optional[datetime]:
    """Try to parse a datetime from various formats."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M", "%m/%d/%y %H:%M", "%m-%d-%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        return None


def autosize_columns(ws, min_w: int = 10, max_w: int = 60) -> None:
    """Auto-size worksheet columns based on content."""
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            if v is None:
                continue
            txt = str(v)
            widths[i] = max(widths.get(i, 0), len(txt))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = max(min_w, min(max_w, w + 2))


def safe_sheet_name(name: str, existing: Set[str]) -> str:
    """Create a safe, unique Excel sheet name."""
    # Truncate to max length
    name = name[:EXCEL_SHEET_NAME_MAX_LEN]
    
    # Remove invalid characters
    invalid_chars = r'[]:*?/\\'
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Ensure uniqueness
    base_name = name
    counter = 1
    while name in existing:
        suffix = f"_{counter}"
        name = base_name[:EXCEL_SHEET_NAME_MAX_LEN - len(suffix)] + suffix
        counter += 1
    
    return name


def write_week_sheet(wb: Workbook, wk: str, records: List[dict], existing_names: Set[str]) -> None:
    """Write a weekly sheet to the workbook."""
    name = safe_sheet_name(wk, existing_names)
    existing_names.add(name)
    ws = wb.create_sheet(title=name)

    def _sort_key(r: dict):
        d = try_parse_dt(r.get("completed_utc")) or try_parse_dt(r.get("completed"))
        return d or datetime(1970, 1, 1)

    records = sorted(records, key=_sort_key)

    ws.append([f"Weekly Loads — {wk}"])
    ws["A1"].font = Font(bold=True, size=14)

    headers = [
        "Completed (Local)",
        "TZ",
        "Load #",
        "Job ID",
        "Load Date",
        "PU Time",
        "Pickup",
        "Delivery Times",
        "Deliveries",
        "Stops",
        "Rate",
        "Posted Miles",
        "Est Miles",
        "Rate/EstMi",
    ]
    ws.append(headers)
    for c in ws[2]:
        c.font = Font(bold=True)

    sum_rate = 0.0
    sum_miles = 0.0

    for r in records:
        completed_dt = try_parse_dt(r.get("completed")) or try_parse_dt(r.get("completed_utc"))
        load_date = try_parse_date(r.get("load_date"))
        pu_time = try_parse_dt(r.get("pu_time"))

        rate = r.get("rate")
        posted = r.get("posted_miles")
        est = r.get("est_miles")

        rpm: Optional[float] = None
        if isinstance(rate, (int, float)) and isinstance(est, (int, float)) and float(est) > 0:
            rpm = float(rate) / float(est)

        ws.append([
            completed_dt if completed_dt else (r.get("completed") or ""),
            r.get("tz") or "",
            r.get("load_number") or "",
            r.get("job_id") or "",
            load_date if load_date else (r.get("load_date") or ""),
            pu_time if pu_time else (r.get("pu_time") or ""),
            r.get("pickup") or "",
            r.get("del_times") or "",
            r.get("deliveries") or "",
            r.get("stops") or "",
            float(rate) if isinstance(rate, (int, float)) else None,
            float(posted) if isinstance(posted, (int, float)) else None,
            float(est) if isinstance(est, (int, float)) else None,
            float(rpm) if rpm is not None else None,
        ])

        if isinstance(rate, (int, float)):
            sum_rate += float(rate)
        use = posted if isinstance(posted, (int, float)) else (est if isinstance(est, (int, float)) else None)
        if use is not None:
            sum_miles += float(use)

    ws.append([])
    ws.append([
        "TOTAL", "", "", "", "", "", "", "", "", "",
        sum_rate,
        "",
        sum_miles,
        (sum_rate / sum_miles) if sum_miles else None,
    ])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)

    # Apply number formats
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        if isinstance(row[0].value, datetime):
            row[0].number_format = "yyyy-mm-dd hh:mm"
        if len(row) > 4 and isinstance(row[4].value, date):
            row[4].number_format = "yyyy-mm-dd"
        if len(row) > 5 and isinstance(row[5].value, datetime):
            row[5].number_format = "yyyy-mm-dd hh:mm"
        if len(row) > 10 and row[10].value is not None:
            row[10].number_format = '"$"#,##0'
        if len(row) > 11 and row[11].value is not None:
            row[11].number_format = "0"
        if len(row) > 12 and row[12].value is not None:
            row[12].number_format = "0"
        if len(row) > 13 and row[13].value is not None:
            row[13].number_format = '"$"#,##0.00'

    ws.freeze_panes = "A3"
    autosize_columns(ws)


def make_xlsx_weekly(records: List[dict], wk: str) -> Tuple[bytes, str]:
    """Create an Excel file with weekly load data."""
    wb = Workbook()
    
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    existing_names: Set[str] = set()

    if wk == "ALL":
        by: Dict[str, List[dict]] = {}
        for r in records:
            k = r.get("week") or "UNKNOWN"
            by.setdefault(k, []).append(r)
        for wk2 in sorted(by.keys()):
            write_week_sheet(wb, wk2, by[wk2], existing_names)
        filename = "load_catalog_ALL.xlsx"
    else:
        write_week_sheet(wb, wk, records, existing_names)
        filename = f"load_catalog_{wk}.xlsx"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), filename


def parse_catalog_arg(args: List[str], tz_name: str) -> str:
    """Parse catalog command arguments."""
    wk = week_key(now_utc().astimezone(safe_tz(tz_name)))
    if not args:
        return wk
    
    a = args[0].strip().lower()
    if a == "all":
        return "ALL"
    if a in ("last", "prev", "previous"):
        return week_key(now_utc().astimezone(safe_tz(tz_name)) - timedelta(days=7))
    if re.fullmatch(r"\d{4}-w\d{2}", a):
        return a.upper().replace("w", "W")
    return wk


async def send_catalog(update: Update, ctx: ContextTypes.DEFAULT_TYPE, *, from_callback: bool = False):
    """Send the load catalog as an Excel file."""
    async with _state_lock:
        st = load_state()

    if not is_owner(update, st):
        if from_callback and update.callback_query:
            await update.callback_query.answer("Owner only.", show_alert=True)
        else:
            await update.effective_message.reply_text("Owner only.")
        return

    if not chat_allowed(update, st):
        if from_callback and update.callback_query:
            await update.callback_query.answer("Run /allowhere in this group.", show_alert=True)
        else:
            await update.effective_message.reply_text("This chat isn't allowed. Owner: run /allowhere here.")
        return

    hist = list(st.get("history") or [])
    if not hist:
        if from_callback and update.callback_query:
            await update.callback_query.answer("No finished loads yet.", show_alert=True)
        else:
            await update.effective_message.reply_text("No finished loads yet. Finish a load first.")
        return

    tz_name = (st.get("last_location") or {}).get("tz") or "UTC"
    wk = parse_catalog_arg(getattr(ctx, "args", []) or [], tz_name)

    records = hist if wk == "ALL" else [r for r in hist if r.get("week") == wk]
    if not records:
        if from_callback and update.callback_query:
            await update.callback_query.answer("No records for that week.", show_alert=True)
        else:
            await update.effective_message.reply_text("No records for that week.")
        return

    xlsx, filename = make_xlsx_weekly(records, wk)
    bio = io.BytesIO(xlsx)
    bio.name = filename
    await ctx.bot.send_document(
        chat_id=update.effective_chat.id,
        document=bio,
        filename=filename,
        caption=f"📊 Catalog ({wk})",
    )


# ----------------------------
# Telegram commands
# ----------------------------
async def start_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /start command."""
    await update.effective_message.reply_text(
        f"Dispatch Bot ({BOT_VERSION})\n"
        f"Triggers: {', '.join(sorted(TRIGGERS))}\n\n"
        "DM setup:\n"
        "1) /claim <code>\n"
        "2) /update (send location)\n\n"
        "Group setup:\n"
        "3) /allowhere (in the group)\n\n"
        "Use: eta / 1717 or /panel\n"
        "Finish: /finish (owner)\n"
        "Catalog: /catalog (owner)\n"
        "Tools: /skip • /leave • /deleteall\n"
        "Debug: /status • /ping"
    )


async def ping_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /ping command."""
    await update.effective_message.reply_text(f"pong ✅ ({BOT_VERSION})")


async def status_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /status command."""
    async with _state_lock:
        st = load_state()

    uid = update.effective_user.id if update.effective_user else None
    allowed_here = chat_allowed(update, st)
    loc = st.get("last_location")
    job = normalize_job(st.get("job"))

    lines = [
        f"<b>Status</b> ({h(BOT_VERSION)})",
        f"<b>Your user id:</b> {h(uid)}",
        f"<b>Owner id:</b> {h(st.get('owner_id'))}",
        f"<b>This chat allowed:</b> {h(allowed_here)}",
        f"<b>Allowed chats:</b> {h(len(st.get('allowed_chats') or []))}",
        f"<b>State file:</b> {h(str(STATE_FILE))}",
        f"<b>Location saved:</b> {'✅' if loc else '❌'}",
        f"<b>Active load:</b> {'✅' if job else '❌'}",
        f"<b>History rows:</b> {h(len(st.get('history') or []))}",
    ]
    await update.effective_message.reply_text("\n".join(lines), parse_mode="HTML")


async def claim_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /claim command."""
    if not update.effective_chat or update.effective_chat.type != "private":
        await update.effective_message.reply_text("DM me: /claim <code>")
        return
    
    if not CLAIM_CODE:
        await update.effective_message.reply_text("CLAIM_CODE is missing in environment variables.")
        return

    code = " ".join(ctx.args or []).strip()
    if code != CLAIM_CODE:
        await update.effective_message.reply_text("❌ Wrong claim code.")
        return

    async with _state_lock:
        st = load_state()
        st["owner_id"] = update.effective_user.id
        st["owner"] = st["owner_id"]  # Mirror alias
        save_state(st)

    await update.effective_message.reply_text("✅ Owner set. Now send /update to save your location.")


async def allowhere_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /allowhere command."""
    async with _state_lock:
        st = load_state()
        
        if not is_owner(update, st):
            await update.effective_message.reply_text("Owner only. DM /claim <code> first.")
            return

        chat = update.effective_chat
        if not chat or chat.type == "private":
            await update.effective_message.reply_text("Run /allowhere inside the group you want to allow.")
            return

        allowed = set(st.get("allowed_chats") or [])
        allowed.add(chat.id)
        st["allowed_chats"] = sorted(list(allowed))
        st["allowed"] = st["allowed_chats"]  # Mirror alias
        save_state(st)

    await update.effective_message.reply_text("✅ Group allowed.")


async def update_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /update command."""
    async with _state_lock:
        st = load_state()
        
        if not is_owner(update, st):
            await update.effective_message.reply_text("Owner only. DM /claim <code> first.")
            return

    if not update.effective_chat or update.effective_chat.type != "private":
        await update.effective_message.reply_text("Please DM me /update (best).")
        return

    kb = [[KeyboardButton("📍 Send my current location", request_location=True)]]
    await update.effective_message.reply_text(
        "Tap to send your location.\n"
        "Tip: Attach → Location → Share Live Location for ongoing updates.",
        reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True, one_time_keyboard=True),
    )


async def on_location(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle location messages."""
    msg = update.effective_message
    if not msg or not msg.location:
        return

    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            return
        
        lat, lon = msg.location.latitude, msg.location.longitude
        tz = TF.timezone_at(lat=lat, lng=lon) or "UTC"
        st["last_location"] = {"lat": lat, "lon": lon, "tz": tz, "updated_at": now_iso()}
        st["last"] = {  # Mirror alias
            "lat": lat, "lon": lon, "tz": tz, "at": st["last_location"]["updated_at"]
        }
        save_state(st)

    if update.effective_chat and update.effective_chat.type == "private":
        await msg.reply_text("✅ Location saved.", reply_markup=ReplyKeyboardRemove())


async def panel_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /panel command."""
    async with _state_lock:
        st = load_state()

    if not chat_allowed(update, st):
        if update.effective_chat and update.effective_chat.type != "private":
            await update.effective_message.reply_text("This chat isn't allowed yet. Owner: run /allowhere here.")
        return

    job = normalize_job(st.get("job"))
    if not job:
        await update.effective_message.reply_text("No active load detected yet.")
        return

    m = await update.effective_message.reply_text(
        f"<b>{h(load_id_text(job))}</b>\nTap buttons to update status.",
        parse_mode="HTML",
        reply_markup=build_keyboard(job, st),
    )

    # Remember panel message id for /finish
    if update.effective_chat:
        async with _state_lock:
            st2 = load_state()
            pm = st2.get("panel_messages") or {}
            pm[str(update.effective_chat.id)] = int(m.message_id)
            st2["panel_messages"] = pm
            save_state(st2)


async def finish_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /finish command."""
    out = await finish_active_load(update, ctx, source="command")
    if not out:
        return
    rec, st2 = out

    rate = rec.get("rate")
    rate_txt = money(rate if isinstance(rate, (int, float)) else None)
    id_txt = rec.get("load_number") or rec.get("job_id") or ""

    # Short, auto-deleting notification
    await send_progress_alert(ctx, update.effective_chat.id, f"✅ <b>Load finished</b> {h(id_txt)} · {h(rate_txt)}")

    wk_count = int(rec.get("_wk_count") or 0)
    wk_rate = float(rec.get("_wk_rate") or 0.0)
    wk_miles = float(rec.get("_wk_miles") or 0.0)

    report = "\n".join([
        f"✅ <b>Load finished</b>",
        f"{h(id_txt)} · {h(rate_txt)}",
        f"Week {h(rec.get('week'))}: {h(wk_count)} loads · {h(money(wk_rate))} · {h(int(round(wk_miles)))} mi",
        "📊 Tap Catalog for Excel.",
    ])

    # Try to edit the last panel message in this chat
    chat_id = update.effective_chat.id if update.effective_chat else None
    if chat_id:
        pm = st2.get("panel_messages") or {}
        msg_id = pm.get(str(chat_id))
        if msg_id:
            try:
                await ctx.bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=int(msg_id),
                    text=report,
                    parse_mode="HTML",
                    reply_markup=build_finished_keyboard()
                )
                return
            except TelegramError:
                pass

        # Fallback: send a new compact report
        await ctx.bot.send_message(
            chat_id=chat_id,
            text=report,
            parse_mode="HTML",
            reply_markup=build_finished_keyboard()
        )


async def catalog_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /catalog command."""
    await send_catalog(update, ctx, from_callback=False)


async def skip_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /skip command."""
    async with _state_lock:
        st = load_state()

        if not is_owner(update, st):
            await update.effective_message.reply_text("Owner only.")
            return

        job = normalize_job(st.get("job"))
        if not job:
            await update.effective_message.reply_text("No active load.")
            return

        stage, i = focus(job, st)
        if stage != "DEL":
            await update.effective_message.reply_text("Can't skip yet — finish PU first.")
            return

        dels = job.get("del") or []
        if not dels:
            await update.effective_message.reply_text("No delivery stops.")
            return

        if i >= len(dels):
            await update.effective_message.reply_text("Invalid delivery index.")
            return

        dd = dels[i]
        ds = dd.get("status") or {}
        ds["skip"] = True
        if not ds.get("comp"):
            ds["comp"] = now_iso()
        dd["status"] = ds
        dels[i] = dd
        job["del"] = dels

        ni = next_incomplete(job, i + 1)
        if ni is not None:
            st["focus_i"] = ni

        st["job"] = job
        save_state(st)

    await send_progress_alert(ctx, update.effective_chat.id, f"⏭️ <b>Skipped</b> stop {h(i+1)}/{h(len(dels))}")
    await update.effective_message.reply_text("Skipped. Use /panel to refresh buttons.")


# ----------------------------
# ETA
# ----------------------------
async def send_eta(update: Update, ctx: ContextTypes.DEFAULT_TYPE, which: str):
    """Send ETA information."""
    async with _state_lock:
        st = load_state()

    if not chat_allowed(update, st):
        return

    loc = st.get("last_location")
    if not loc:
        await update.effective_message.reply_text("No saved location yet. Owner: DM /update.")
        return

    try:
        origin = (float(loc["lat"]), float(loc["lon"]))
    except (KeyError, ValueError, TypeError):
        await update.effective_message.reply_text("Invalid location data. Owner: DM /update.")
        return
        
    tz_now = loc.get("tz") or "UTC"
    tz = safe_tz(tz_now)

    await ctx.bot.send_location(chat_id=update.effective_chat.id, latitude=origin[0], longitude=origin[1])

    job = normalize_job(st.get("job"))
    if not job:
        await update.effective_message.reply_text(
            f"<b>⏱ ETA</b>\nNow: {h(datetime.now(tz).strftime('%Y-%m-%d %H:%M'))} ({h(tz_now)})\n\n<i>No active load yet.</i>",
            parse_mode="HTML",
        )
        return

    which = (which or "AUTO").upper()

    if which == "ALL":
        lines: List[str] = [f"<b>{h(load_id_text(job))}</b>"]
        pu = job.get("pu") or {}
        stops: List[Tuple[str, str, List[str], Optional[str]]] = [
            ("PU", pu.get("addr") or "", pu.get("lines") or [], pu.get("time"))
        ]
        for j, d in enumerate((job.get("del") or [])[:ETA_ALL_MAX]):
            stops.append((f"D{j+1}", d.get("addr") or "", d.get("lines") or [], d.get("time")))

        for lab, addr, addr_lines, appt in stops:
            r = await eta_to(st, origin, lab, addr)
            place = short_place(addr_lines, addr)
            if r.get("ok"):
                arr = (now_utc().astimezone(tz) + timedelta(seconds=float(r["s"]))).strftime("%H:%M")
                tag = " (approx)" if r.get("method") == "approx" else ""
                appt_txt = f" · Appt: {appt}" if appt else ""
                lines.append(
                    f"<b>{h(lab)}:</b> <b>{h(fmt_dur(r['s']))}</b>{h(tag)} · {h(fmt_mi(r['m']))} · ~{h(arr)}{h(appt_txt)} — {h(place)}"
                )
            else:
                lines.append(f"<b>{h(lab)}:</b> ⚠️ {h(r.get('err'))} — {h(place)}")

        await update.effective_message.reply_text(
            "\n".join(lines),
            parse_mode="HTML",
            reply_markup=build_keyboard(job, st)
        )
        return

    stage, i = focus(job, st)
    if stage == "PU":
        pu = job.get("pu") or {}
        addr = pu.get("addr") or ""
        addr_lines = pu.get("lines") or []
        appt = pu.get("time")
        stop_label = "PU"
    else:
        dels = job.get("del") or []
        d = dels[i] if i < len(dels) else {"addr": "", "lines": [], "time": None}
        addr = d.get("addr") or ""
        addr_lines = d.get("lines") or []
        appt = d.get("time")
        stop_label = f"DEL {i+1}/{len(dels)}" if dels else "DEL"

    r = await eta_to(st, origin, stop_label, addr)
    place = short_place(addr_lines, addr)

    if r.get("ok"):
        arr = (now_utc().astimezone(tz) + timedelta(seconds=float(r["s"]))).strftime("%H:%M")
        tag = " (approx)" if r.get("method") == "approx" else ""
        out = [
            f"<b>⏱ ETA: {h(fmt_dur(r['s']))}</b>{h(tag)} — {h(stop_label)}",
            f"{h(load_id_text(job))} · {h(place)}",
            f"Arrive ~ {h(arr)} ({h(tz_now)}) · {h(fmt_mi(r['m']))}",
        ]
        if appt:
            out.append(f"Appt: {h(appt)}")
        await update.effective_message.reply_text(
            "\n".join(out),
            parse_mode="HTML",
            reply_markup=build_keyboard(job, st)
        )
    else:
        await update.effective_message.reply_text(
            f"<b>{h(load_id_text(job))}</b>\n<b>⏱ ETA:</b> ⚠️ {h(r.get('err'))}\n<b>Target:</b> {h(place)}",
            parse_mode="HTML",
            reply_markup=build_keyboard(job, st),
        )


# ----------------------------
# Admin tools
# ----------------------------
async def deleteall_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /deleteall command."""
    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            await update.effective_message.reply_text("Owner only.")
            return

    chat = update.effective_chat
    if not chat or chat.type == "private":
        await update.effective_message.reply_text("Bots can't clear a DM history. Delete the chat from your side.")
        return

    n = DELETEALL_DEFAULT
    if ctx.args:
        try:
            n = max(1, min(MAX_DELETEALL_MESSAGES, int(ctx.args[0])))
        except ValueError:
            pass

    notice = await update.effective_message.reply_text(f"🧹 Deleting up to {n} messages… (bot must be admin)")
    start_id = notice.message_id

    for mid in range(start_id, max(1, start_id - n + 1) - 1, -1):
        try:
            await ctx.bot.delete_message(chat_id=chat.id, message_id=mid)
        except (Forbidden, BadRequest):
            break
        await asyncio.sleep(DELETE_MESSAGE_DELAY_SEC)


async def leave_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /leave command."""
    async with _state_lock:
        st = load_state()
        if not is_owner(update, st):
            await update.effective_message.reply_text("Owner only.")
            return

        chat = update.effective_chat
        if not chat or chat.type == "private":
            await update.effective_message.reply_text("Run /leave inside the group you want the bot to leave.")
            return

        allowed = set(st.get("allowed_chats") or [])
        allowed.discard(chat.id)
        st["allowed_chats"] = sorted(list(allowed))
        st["allowed"] = st["allowed_chats"]  # Mirror alias
        save_state(st)

    await update.effective_message.reply_text("👋 Leaving this chat…")
    try:
        await ctx.bot.leave_chat(chat.id)
    except TelegramError as e:
        log(f"Failed to leave chat: {e}")


# ----------------------------
# Callback handler
# ----------------------------
async def on_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle inline keyboard callbacks."""
    q = update.callback_query
    if not q or not q.data:
        return

    data = q.data

    async with _state_lock:
        st = load_state()

    if not chat_allowed(update, st):
        await q.answer("Not allowed here.", show_alert=False)
        return

    if data.startswith("ETA:"):
        await q.answer("Computing ETA…", show_alert=False)
        await send_eta(update, ctx, data.split(":", 1)[1])
        return

    if data == "SHOW:CAT":
        await q.answer()
        await send_catalog(update, ctx, from_callback=True)
        return

    if data == "JOB:FIN":
        if not is_owner(update, st):
            await q.answer("Owner only. DM /claim <code>.", show_alert=True)
            return

        await q.answer("Finishing…", show_alert=False)
        out = await finish_active_load(update, ctx, source="callback")
        if not out:
            return
        rec, _ = out

        rate = rec.get("rate")
        rate_txt = money(rate if isinstance(rate, (int, float)) else None)
        id_txt = rec.get("load_number") or rec.get("job_id") or ""
        await send_progress_alert(ctx, update.effective_chat.id, f"✅ <b>Load finished</b> {h(id_txt)} · {h(rate_txt)}")

        wk_count = int(rec.get("_wk_count") or 0)
        wk_rate = float(rec.get("_wk_rate") or 0.0)
        wk_miles = float(rec.get("_wk_miles") or 0.0)

        report = "\n".join([
            f"✅ <b>Load finished</b>",
            f"{h(id_txt)} · {h(rate_txt)}",
            f"Week {h(rec.get('week'))}: {h(wk_count)} loads · {h(money(wk_rate))} · {h(int(round(wk_miles)))} mi",
            "📊 Tap Catalog for Excel.",
        ])

        try:
            await q.edit_message_text(text=report, parse_mode="HTML", reply_markup=build_finished_keyboard())
        except TelegramError:
            try:
                await q.edit_message_reply_markup(reply_markup=build_finished_keyboard())
            except TelegramError:
                pass
        return

    # Handle progress buttons
    async with _state_lock:
        st2 = load_state()
        job = normalize_job(st2.get("job"))
        if not job:
            await q.answer("No active load.", show_alert=True)
            try:
                await q.edit_message_reply_markup(reply_markup=build_finished_keyboard())
            except TelegramError:
                pass
            return

        stage, i = focus(job, st2)
        tz_name = (st2.get("last_location") or {}).get("tz") or "UTC"
        ts = local_stamp(tz_name)
        load_label = load_id_text(job)

        progress_broadcast: Optional[str] = None

        if data.startswith("PU:"):
            pu = job.get("pu") or {}
            ps = pu.get("status") or {}
            
            if data == "PU:A":
                on = toggle_ts(ps, "arr")
                if on:
                    progress_broadcast = f"📍 <b>PU Arrived</b> — {h(ts)} — {h(load_label)}"
            elif data == "PU:L":
                on = toggle_ts(ps, "load")
                if on:
                    progress_broadcast = f"📦 <b>Loaded</b> — {h(ts)} — {h(load_label)}"
            elif data == "PU:D":
                on = toggle_ts(ps, "dep")
                if on:
                    progress_broadcast = f"🚚 <b>Departed PU</b> — {h(ts)} — {h(load_label)}"
            elif data == "PU:C":
                on = toggle_ts(ps, "comp")
                if on:
                    progress_broadcast = f"✅ <b>PU COMPLETE</b> — {h(ts)} — {h(load_label)}"
                    ni = next_incomplete(job, 0)
                    if ni is not None:
                        st2["focus_i"] = ni
            
            pu["status"] = ps
            job["pu"] = pu

        elif data.startswith("DEL:"):
            if stage != "DEL":
                await q.answer("Complete PU first.", show_alert=False)
                return

            dels = job.get("del") or []
            if not dels or i >= len(dels):
                await q.answer("No deliveries.", show_alert=False)
                return

            dd = dels[i]
            ds = dd.get("status") or {}
            lbl = f"DEL {i+1}/{len(dels)}"

            if data == "DEL:A":
                on = toggle_ts(ds, "arr")
                if on:
                    progress_broadcast = f"📍 <b>Arrived {h(lbl)}</b> — {h(ts)} — {h(load_label)}"
            elif data == "DEL:DL":
                on = toggle_ts(ds, "del")
                if on:
                    progress_broadcast = f"📦 <b>Delivered {h(lbl)}</b> — {h(ts)} — {h(load_label)}"
            elif data == "DEL:D":
                on = toggle_ts(ds, "dep")
                if on:
                    progress_broadcast = f"🚚 <b>Departed {h(lbl)}</b> — {h(ts)} — {h(load_label)}"
            elif data == "DEL:C":
                on = toggle_ts(ds, "comp")
                if on:
                    progress_broadcast = f"✅ <b>STOP COMPLETE {h(lbl)}</b> — {h(ts)} — {h(load_label)}"
                    ni = next_incomplete(job, i + 1)
                    if ni is not None:
                        st2["focus_i"] = ni
            elif data == "DEL:S":
                ds["skip"] = True
                if not ds.get("comp"):
                    ds["comp"] = now_iso()
                progress_broadcast = f"⏭️ <b>SKIPPED {h(lbl)}</b> — {h(ts)} — {h(load_label)}"
                ni = next_incomplete(job, i + 1)
                if ni is not None:
                    st2["focus_i"] = ni

            dd["status"] = ds
            dels[i] = dd
            job["del"] = dels

        elif data.startswith("DOC:"):
            pu = job.get("pu") or {}
            pd = pu.get("docs") or {}
            
            if data == "DOC:PTI":
                pd["pti"] = not bool(pd.get("pti"))
            elif data == "DOC:BOL":
                pd["bol"] = not bool(pd.get("bol"))
            elif data == "DOC:POD":
                if stage != "DEL":
                    await q.answer("Complete PU first.", show_alert=False)
                    return
                dels = job.get("del") or []
                if not dels or i >= len(dels):
                    await q.answer("No deliveries.", show_alert=False)
                    return
                dd = dels[i].get("docs") or {}
                dd["pod"] = not bool(dd.get("pod"))
                dels[i]["docs"] = dd
                job["del"] = dels
            
            pu["docs"] = pd
            job["pu"] = pu

        st2["job"] = job
        save_state(st2)

    await q.answer("Updated.", show_alert=False)

    if progress_broadcast:
        await send_progress_alert(ctx, update.effective_chat.id, progress_broadcast)

    try:
        await q.edit_message_reply_markup(reply_markup=build_keyboard(job, st2))
    except TelegramError:
        pass


# ----------------------------
# Text handler (new load + triggers)
# ----------------------------
async def on_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle text messages for load detection and triggers."""
    msg = update.effective_message
    if not msg or not msg.text:
        return

    async with _state_lock:
        st = load_state()

    chat = update.effective_chat

    # Detect new loads only in allowed groups
    if chat and chat.type in ("group", "supergroup"):
        allowed_set = set(st.get("allowed_chats") or [])
        if chat.id not in allowed_set:
            return
        
        job = parse_job(msg.text)
        if job:
            async with _state_lock:
                st2 = load_state()
                st2["job"] = job
                st2["focus_i"] = 0
                save_state(st2)
            await msg.reply_text("📦 New load detected. Type eta / 1717 or /panel.")
            return

    # Triggers in allowed chats
    if not chat_allowed(update, st):
        return

    parts = msg.text.strip().split()
    if not parts:
        return

    first = re.sub(r"^[^\w]+|[^\w]+$", "", parts[0].lower())
    if first in TRIGGERS:
        rest = " ".join(parts[1:]).lower()
        which = "ALL" if "all" in rest else "AUTO"
        await send_eta(update, ctx, which)


# ----------------------------
# Startup: disable webhook so polling works
# ----------------------------
async def _post_init(app):
    """Post-initialization hook."""
    try:
        await app.bot.delete_webhook(drop_pending_updates=True)
    except TelegramError as e:
        log(f"Failed to delete webhook: {e}")
    
    try:
        me = await app.bot.get_me()
        log(f"Connected as @{me.username} (id {me.id})")
    except TelegramError as e:
        log(f"get_me failed: {e}")
    
    log("Ready.")


# ----------------------------
# Main
# ----------------------------
def main() -> None:
    """Main entry point."""
    if not TOKEN:
        raise RuntimeError("Missing TELEGRAM_TOKEN environment variable")

    builder = ApplicationBuilder().token(TOKEN)
    
    try:
        builder = builder.post_init(_post_init)
    except Exception as e:
        log(f"Failed to set post_init: {e}")

    app = builder.build()

    # Register command handlers
    app.add_handler(CommandHandler("start", start_cmd))
    app.add_handler(CommandHandler("ping", ping_cmd))
    app.add_handler(CommandHandler("status", status_cmd))
    app.add_handler(CommandHandler("claim", claim_cmd))
    app.add_handler(CommandHandler("allowhere", allowhere_cmd))
    app.add_handler(CommandHandler("update", update_cmd))
    app.add_handler(CommandHandler("panel", panel_cmd))
    app.add_handler(CommandHandler("finish", finish_cmd))
    app.add_handler(CommandHandler("catalog", catalog_cmd))
    app.add_handler(CommandHandler("skip", skip_cmd))
    app.add_handler(CommandHandler("deleteall", deleteall_cmd))
    app.add_handler(CommandHandler("leave", leave_cmd))

    # Register other handlers
    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.LOCATION, on_location))
    
    try:
        app.add_handler(MessageHandler(filters.UpdateType.EDITED_MESSAGE & filters.LOCATION, on_location))
    except Exception as e:
        log(f"Failed to add edited location handler: {e}")

    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))

    log("Starting polling…")
    app.run_polling(drop_pending_updates=True, allowed_updates=Update.ALL_TYPES, close_loop=False)


if __name__ == "__main__":
    main()
